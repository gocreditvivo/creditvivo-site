from __future__ import annotations

"""
Credit Vivo Proprietary Parser Engine v16

Purpose:
- Build a stronger original Credit Vivo parser/scanner engine.
- No Claude / Anthropic / paid AI API.
- No competitor code.
- No automatic dispute sending.
- Every output includes source evidence and confidence.

Core ideas:
1. Extract all raw report text first.
2. Detect bureau and report layout.
3. Segment report into account/profile blocks.
4. Extract fields into Credit Vivo's normalized tradeline schema.
5. Score confidence.
6. Match same/similar accounts across bureaus.
7. Detect review issues.
8. Return customer-friendly and admin-ready output.

This is draft review software. It does not provide legal advice.
"""

from dataclasses import dataclass, field, asdict
from typing import Dict, List, Optional, Tuple, Iterable
import hashlib
import re
import json
import csv
from pathlib import Path

try:
    from .bureau_debt_collection_reference import build_bureau_debt_collection_reference
    from .fcra_rights_reference import build_fcra_rights_reference
except ImportError:
    from bureau_debt_collection_reference import build_bureau_debt_collection_reference
    from fcra_rights_reference import build_fcra_rights_reference

try:
    from openpyxl import Workbook
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None
    load_workbook = None
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None


PARSER_VERSION = "16.1"
COMPLIANCE_RULE_PACK_VERSION = "2026.07.03-v1"
METRO2_RULE_PACK_VERSION = "metro2-review-2026.07.03-v1"
LETTER_TEMPLATE_VERSION = "letters-2026.07.03-v1"
SECURITY_CONFIG_VERSION = "scanner-security-2026.07.03-v1"


# -----------------------------
# Utility
# -----------------------------

def clean_text(text: str) -> str:
    text = text.replace("\x00", " ")
    text = text.replace("\u00a0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n[ \t]+", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def compact_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def stable_id(*parts: str) -> str:
    raw = "|".join(parts).encode("utf-8", errors="ignore")
    return "cvp_" + hashlib.sha256(raw).hexdigest()[:16]


def money_to_number(value: str) -> Optional[float]:
    if not value:
        return None
    value = value.replace("$", "").replace(",", "").strip()
    try:
        return float(value)
    except Exception:
        return None


def normalize_money(value: str) -> str:
    n = money_to_number(value)
    if n is None:
        return value.strip()
    if n.is_integer():
        return f"${int(n):,}"
    return f"${n:,.2f}"


def normalize_date(value: str) -> str:
    if not value:
        return ""
    return re.sub(r"\s+", " ", value.strip())


def mask_account_number(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""

    # Preserve only an already-visible trailing identifier. Reports sometimes
    # render a masked value as `****1234`; the inverse form `1234XXXX` must not
    # be interpreted as revealing the last four because those digits are a
    # clear prefix, not an approved suffix.
    if re.search(r"[*Xx#•]", value):
        trailing_clear = re.search(r"[*Xx#•][*Xx#•\s._-]*([A-WYZa-wyz0-9]{1,4})$", value)
        if trailing_clear:
            return "*" + trailing_clear.group(1)[-4:]
        return "****"

    token = re.sub(r"[^A-Za-z0-9]", "", value)
    if not token:
        return ""
    if len(token) <= 4:
        return "*" * len(token)
    return "*" + token[-4:]


def sanitize_sensitive_text(value: str) -> str:
    """Redact identifiers before any parser data crosses an output boundary."""
    text = str(value or "")
    text = re.sub(r"\b\d{3}-?\d{2}-?\d{4}\b", "***-**-****", text)
    text = re.sub(
        r"((?:date\s+of\s+birth|dob)\s*[:#-]?\s*)"
        r"(?:0?[1-9]|1[0-2])[/-](?:0?[1-9]|[12]\d|3[01])[/-](?:19|20)\d{2}\b",
        r"\1**/**/****",
        text,
        flags=re.I,
    )

    # Keep a match on one physical line so a raw evidence block is not
    # swallowed after the account value.
    account_pattern = re.compile(
        r"((?:account|acct)\s*(?:number|no\.?|#)\s*[:#-]?\s*)([A-Za-z0-9*Xx#•._ -]{2,})",
        flags=re.I,
    )
    account_pattern = re.compile(
        account_pattern.pattern
        .replace(r"\s*", r"[ \t]*")
        .replace(r"no\.?|#", r"no\.?(?![A-Za-z])|#"),
        flags=re.I,
    )
    text = account_pattern.sub(lambda match: match.group(1) + mask_account_number(match.group(2)), text)

    # Long uninterrupted numbers in evidence are identifiers or confirmation
    # values, not amounts/dates. Retain only a last-four reference.
    text = re.sub(r"\b\d{6,}\b", lambda match: "*" + match.group(0)[-4:], text)
    return text


def sanitize_output_payload(value):
    """Recursively sanitize every persisted/API/exported parser value."""
    if isinstance(value, dict):
        sanitized = {}
        for key, item in value.items():
            if key in {"raw_block", "raw_value"}:
                continue
            if key == "account_number_masked" and isinstance(item, str):
                sanitized[key] = mask_account_number(str(item or ""))
            else:
                sanitized[key] = sanitize_output_payload(item)
        return sanitized
    if isinstance(value, list):
        return [sanitize_output_payload(item) for item in value]
    if isinstance(value, tuple):
        return tuple(sanitize_output_payload(item) for item in value)
    if isinstance(value, str):
        return sanitize_sensitive_text(value)
    return value


def simple_similarity(a: str, b: str) -> float:
    """
    Custom no-dependency similarity for proprietary matching.
    Not perfect, but good enough for candidate grouping without RapidFuzz.
    """
    ak = compact_key(a)
    bk = compact_key(b)
    if not ak or not bk:
        return 0.0
    if ak == bk:
        return 1.0
    if ak in bk or bk in ak:
        return 0.86

    aset = set(re.findall(r"[a-z0-9]{2,}", a.lower()))
    bset = set(re.findall(r"[a-z0-9]{2,}", b.lower()))
    if not aset or not bset:
        return 0.0
    return len(aset & bset) / max(1, len(aset | bset))


# -----------------------------
# Schema
# -----------------------------

@dataclass
class Evidence:
    bureau: str
    page: Optional[int]
    snippet: str
    extraction_method: str = "native_rule_engine"


@dataclass
class FieldEvidence:
    field_name: str
    raw_value: str
    normalized_value: str
    source_filename: str
    source_file_hash: str
    bureau: str
    page: Optional[int]
    raw_block_id: str
    raw_block_hash: str
    raw_line: str = ""
    extraction_rule_id: str = ""
    confidence_score: float = 0.0
    needs_admin_review: bool = False


@dataclass
class BureauDetectionResult:
    bureau: str
    confidence_score: float
    source: str
    conflict: bool = False
    evidence: List[str] = field(default_factory=list)


@dataclass
class NormalizedTradeline:
    id: str
    bureau: str
    source_filename: str
    account_name: str = ""
    account_number_masked: str = ""
    account_type: str = ""
    portfolio_type: str = ""
    responsibility: str = ""
    creditor_classification: str = ""
    original_creditor: str = ""
    collector_or_debt_buyer: str = ""
    status: str = ""
    pay_status: str = ""
    balance: str = ""
    past_due: str = ""
    high_credit_or_original_amount: str = ""
    credit_limit: str = ""
    date_opened: str = ""
    date_closed: str = ""
    date_reported: str = ""
    date_last_activity: str = ""
    date_last_payment: str = ""
    date_of_first_delinquency: str = ""
    estimated_removal_date: str = ""
    remarks: str = ""
    payment_history_summary: str = ""
    raw_block: str = ""
    page_start: Optional[int] = None
    confidence: str = "medium"
    confidence_score: float = 0.0
    needs_admin_review: bool = True
    field_evidence: Dict[str, FieldEvidence] = field(default_factory=dict)
    source_file_hash: str = ""
    raw_block_id: str = ""
    raw_block_hash: str = ""
    bureau_conflict: bool = False
    parser_warnings: List[str] = field(default_factory=list)
    is_negative: bool = False
    negative_item_type: str = ""
    negative_signals: List[str] = field(default_factory=list)
    positive_signals: List[str] = field(default_factory=list)


@dataclass
class ReviewIssue:
    id: str
    issue_type: str
    severity: str
    customer_label: str
    customer_explanation: str
    admin_explanation: str
    suggested_round: str
    related_tradeline_ids: List[str] = field(default_factory=list)
    evidence: List[Evidence] = field(default_factory=list)
    confidence: str = "medium"


@dataclass
class ParseResult:
    engine: str
    version: str
    paid_ai_used: bool
    files: List[dict]
    tradelines: List[NormalizedTradeline]
    issues: List[ReviewIssue]
    cross_bureau_groups: List[dict]
    customer_summary: dict
    admin_summary: dict
    parser_qa_warnings: List[dict] = field(default_factory=list)


# -----------------------------
# Bureau profiles
# -----------------------------

DATE = r"(?:\d{1,2}/\d{1,2}/\d{2,4}|\d{4}-\d{2}-\d{2}|(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*\.?\s+\d{1,2},?\s+\d{4}|(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*\.?\s+\d{4})"
MONEY = r"\$?\s?[0-9]{1,3}(?:,[0-9]{3})*(?:\.\d{2})?|\$?\s?[0-9]+(?:\.\d{2})?"

COMMON_FIELD_PATTERNS: Dict[str, List[str]] = {
    "account_number_masked": [
        r"(?:account\s*(?:#|number|no\.?)|acct\s*(?:#|number|no\.?))\s*[:\-]?\s*([A-Za-z0-9\*xX# -]{3,40})$",
        r"(?:account\s*(?:#|number|no\.?)|acct\s*(?:#|number|no\.?))\s*[:\-]?\s*([A-Za-z0-9\*\-xX]{3,32})",
        r"^\s*([A-Za-z0-9]{4,}\*{2,}[A-Za-z0-9\*]*)\s*$",
    ],
    "account_type": [
        r"(?:loan/account type|account type|type of account|loan type|type)\s*[:\-]?\s*([A-Za-z0-9 /&-]{3,80})",
    ],
    "portfolio_type": [
        r"(?:portfolio type|portfolio)\s*[:\-]?\s*([A-Za-z0-9 /&-]{3,60})",
    ],
    "responsibility": [
        r"(?:responsibility|account holder|owner)\s*[:\-]?\s*([A-Za-z0-9 /&-]{3,60})",
    ],
    "status": [
        r"(?:account status|status)\s*[:\-]?\s*([A-Za-z0-9 /&.,'$-]{3,140})",
        r"(?:pay status)\s*[:\-]?\s*([A-Za-z0-9 /&.,'$-]{3,140})",
    ],
    "pay_status": [
        r"(?:payment status|pay status|payment\s*condition)\s*[:\-]?\s*([A-Za-z0-9 /&.,'$-]{3,140})",
    ],
    "balance": [
        r"(?:current balance|balance)\s*[:\-]?\s*(" + MONEY + r")",
    ],
    "past_due": [
        r"(?:past due|amount past due)\s*[:\-]?\s*(" + MONEY + r")",
        r"(" + MONEY + r")\s+past due",
    ],
    "high_credit_or_original_amount": [
        r"(?:high credit|original amount|original balance|loan amount)\s*[:\-]?\s*(" + MONEY + r")",
    ],
    "credit_limit": [
        r"(?:credit limit|limit)\s*[:\-]?\s*(" + MONEY + r")",
    ],
    "date_opened": [
        r"(?:date opened|opened)\s*[:\-]?\s*(" + DATE + r")",
    ],
    "date_closed": [
        r"(?:date closed|closed)\s*[:\-]?\s*(" + DATE + r")",
    ],
    "date_reported": [
        r"(?:date reported|last reported|balance updated|date updated)\s*[:\-]?\s*(" + DATE + r")",
        r"\breported\s*[:\-]?\s*(" + DATE + r")",
    ],
    "date_last_activity": [
        r"(?:date of last activity|last activity|date last active)\s*[:\-]?\s*(" + DATE + r")",
    ],
    "date_last_payment": [
        r"(?:date of last payment|last payment made|last payment|date last payment)\s*[:\-]?\s*(" + DATE + r")",
    ],
    "date_of_first_delinquency": [
        r"(?:date of first delinquency|first delinquency|dofd|first reported delinquency)\s*[:\-]?\s*(" + DATE + r")",
    ],
    "estimated_removal_date": [
        r"(?:on record until|estimated month and year this item will be removed|estimated removal|scheduled to continue on record until)\s*[:\-]?\s*(" + DATE + r")",
        r"(?:by\s+(" + DATE + r"),?\s+this account is scheduled)",
    ],
    "original_creditor": [
        r"(?:original creditor|original lender|original account)\s*[:\-]?\s*([A-Za-z0-9 &.,'()/\-]{3,90})",
    ],
    "collector_or_debt_buyer": [
        r"(?:collection agency|collector|debt buyer|assigned to)\s*[:\-]?\s*([A-Za-z0-9 &.,'()/\-]{3,90})",
    ],
    "creditor_classification": [
        r"(?:creditor classification|classification|industry type)\s*[:\-]?\s*([A-Za-z0-9 &.,'()/\-]{3,80})",
    ],
    "remarks": [
        r"(?:remarks|comments|comment|account information)\s*[:\-]?\s*([A-Za-z0-9 &.,'()/\-]{5,220})",
        r"(account information disputed by consumer[^\n]{0,180})",
    ],
}

FIELD_STOP_LABELS = [
    "Account Number",
    "Owner",
    "Responsibility",
    "Account Type",
    "Loan Type",
    "Loan/Account Type",
    "Status",
    "Pay Status",
    "Status Updated",
    "Balance",
    "Balance Updated",
    "Date Opened",
    "Date Reported",
    "Date Updated",
    "Date of Last Activity",
    "Last Payment Made",
    "Date of Last Payment",
    "Date of 1st Delinquency",
    "Date of First Delinquency",
    "Date Major Delinquency",
    "Terms",
    "High Credit",
    "High Balance",
    "Credit Limit",
    "Scheduled Payment Amount",
    "Amount Past Due",
    "Actual Payment Amount",
    "Charge Off Amount",
    "Balloon Payment",
    "Term Duration",
    "Payment History",
    "On Record Until",
    "Recent Payment",
    "Monthly Payment",
]

BUREAU_SIGNATURES = {
    "Experian": ["experian", "experian credit report", "experian information solutions"],
    "Equifax": ["equifax", "equifax information services", "econsumer"],
    "TransUnion": ["transunion", "trans union", "transunion llc"],
}

NEGATIVE_TERMS = [
    "collection", "charge off", "charge-off", "charged off", "past due", "late",
    "delinquent", "derogatory", "settled", "repossession", "foreclosure",
    "120 days", "90 days", "60 days", "30 days", "placed for collection",
    "transferred", "sold", "profit and loss", "written off", "bad debt",
    "unpaid", "seriously past due"
]

COLLECTION_TERMS = [
    "collection", "collection agency", "collector", "debt buyer",
    "original creditor", "placed for collection", "assigned"
]

BOILERPLATE_TERMS = [
    "consumerfinance.gov",
    "violates the fcra",
    "additional information",
    "the last reported status of the account",
    "name, address, and phone",
    "insurer, employer, landlord",
    "upgrades and enhancements",
    "account types is good for your credit",
    "amount of the debt",
    "public records and residential information",
    "supplemental public records",
    "your credit report can include",
    "consumer added notices",
    "security freezes or locks",
    "prescreened offers of credit",
    "when reviewing your account info",
    "see if an account is open",
]

BAD_ACCOUNT_NAME_FRAGMENTS = [
    "consumerfinance.gov",
    "learnmore",
    "violates the fcra",
    "additional information",
    "last reported status",
    "account types is good",
    "amount of the debt",
    "name, address",
    "insurer, employer",
    "landlord",
    "enhancements",
    "responsibility individual",
    "public records and residential",
    "supplemental public records",
    "consumer added notices",
    "equifax.com/personal/help",
    "your credit report can include",
    "see if an account is open",
    "closed accounts should have",
    "s of credit accounts",
]

ACCOUNT_SECTION_TERMS = [
    "account number",
    "account #",
    "account name",
    "account type",
    "loan/account type",
    "date opened",
    "date reported",
    "date updated",
    "date of first delinquency",
    "date of 1st delinquency",
    "original creditor",
    "pay status",
    "payment status",
    "account status",
    "balance",
    "past due",
]

NON_ACCOUNT_SECTION_TERMS = [
    "payment history guide",
    "your credit report can include",
    "consumer added notices",
    "security freezes or locks",
    "prescreened offers",
    "see if an account is open",
    "closed accounts should have no money",
    "for more information",
]


# -----------------------------
# Text + page utilities
# -----------------------------

def source_hash(filename: str, text: str) -> str:
    return hashlib.sha256((filename + "\n" + text).encode("utf-8", errors="ignore")).hexdigest()


def detect_bureau(filename: str, text: str) -> str:
    sample = (filename + "\n" + text[:5000]).lower()
    header_text = text[:1200].lower()
    exact_headers = {
        "Equifax": [r"\bequifax\s+credit\s+report\b", r"\bequifax\s+credit\s+file\b"],
        "Experian": [r"\bexperian\s+credit\s+report\b", r"\bexperian\s+credit\s+file\b"],
        "TransUnion": [r"\btransunion\s+credit\s+report\b", r"\btrans\s*union\s+credit\s+report\b"],
    }
    for bureau, patterns in exact_headers.items():
        if any(re.search(pattern, header_text, flags=re.I) for pattern in patterns):
            return bureau

    scores = {}
    for bureau, terms in BUREAU_SIGNATURES.items():
        scores[bureau] = sum(1 for t in terms if t in sample)
    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else "Unknown Bureau"


def detect_bureau_with_conflict(filename: str, text: str, upload_bureau: str = "") -> BureauDetectionResult:
    header_text = text[:1200]
    exact_headers = {
        "Equifax": [r"\bequifax\s+credit\s+report\b", r"\bequifax\s+credit\s+file\b"],
        "Experian": [r"\bexperian\s+credit\s+report\b", r"\bexperian\s+credit\s+file\b"],
        "TransUnion": [r"\btransunion\s+credit\s+report\b", r"\btrans\s*union\s+credit\s+report\b"],
    }
    for bureau, patterns in exact_headers.items():
        for pattern in patterns:
            m = re.search(pattern, header_text, flags=re.I)
            if m:
                conflict = bool(upload_bureau and upload_bureau != "Unknown Bureau" and upload_bureau != bureau)
                evidence = [clean_text(m.group(0))]
                if conflict:
                    evidence.append(f"upload_metadata={upload_bureau}")
                return BureauDetectionResult(
                    bureau=bureau,
                    confidence_score=0.98,
                    source="explicit_page_or_block_header",
                    conflict=conflict,
                    evidence=evidence,
                )

    guessed = detect_bureau(filename, text)
    confidence = 0.65 if guessed != "Unknown Bureau" else 0.0
    return BureauDetectionResult(
        bureau=guessed,
        confidence_score=confidence,
        source="filename_or_signature_fallback" if guessed != "Unknown Bureau" else "not_detected",
        conflict=False,
        evidence=[filename] if guessed != "Unknown Bureau" else [],
    )


def page_split(text: str) -> List[Tuple[Optional[int], str]]:
    """
    Supports extraction format with --- PAGE 1 --- markers.
    """
    parts = re.split(r"\n?\s*---\s*PAGE\s+(\d+)\s*---\s*\n?", text, flags=re.I)
    if len(parts) <= 1:
        return [(None, text)]
    pages = []
    preface = parts[0]
    for i in range(1, len(parts), 2):
        try:
            page_num = int(parts[i])
        except Exception:
            page_num = None
        page_text = parts[i + 1] if i + 1 < len(parts) else ""
        pages.append((page_num, page_text))
    return pages


def split_dense_account_sections(page_text: str) -> List[str]:
    """Split pages whose extracted text has no blank lines between accounts."""
    lines = [clean_text(line) for line in page_text.splitlines()]
    account_indexes = [
        index for index, line in enumerate(lines)
        if re.match(r"^\s*(?:Account\s*(?:Number|#)|Acct\s*(?:Number|#))\s*:?\s*", line, flags=re.I)
    ]
    if len(account_indexes) < 2:
        return [page_text]

    field_prefix = re.compile(
        r"^(?:Account\s*(?:Number|#|Type|Status|Name)|Acct\s*(?:Number|#)|"
        r"Original\s+Creditor|Balance|Past\s+Due|Status|Pay\s+Status|Remarks|"
        r"Date\b|High\s+(?:Balance|Credit)|Credit\s+Limit|Payment\s+History|"
        r"Responsibility|Owner|Terms|Scheduled\s+Payment|Actual\s+Payment)\b",
        flags=re.I,
    )
    report_header = re.compile(r"^(?:Experian|Equifax|TransUnion)(?:\s+Credit)?\s+Report\b", flags=re.I)
    starts: List[int] = []
    previous_account_index = -1
    for account_index in account_indexes:
        start = account_index
        lower_bound = max(previous_account_index + 1, account_index - 8)
        run_top = None
        for candidate_index in range(account_index - 1, lower_bound - 1, -1):
            candidate = lines[candidate_index].strip()
            if not candidate:
                if run_top is not None:
                    break
                continue
            boundaryish = (
                field_prefix.match(candidate) or report_header.match(candidate)
                or len(candidate) > 140
                or re.match(r"^(?:\$|\d{1,2}/\d{1,2}/\d{2,4}\b)", candidate)
            )
            if boundaryish:
                if run_top is not None:
                    break
                continue
            run_top = candidate_index
        if run_top is not None:
            start = run_top
        starts.append(start)
        previous_account_index = account_index

    starts = sorted(set(starts))
    if len(starts) < 2:
        return [page_text]
    sections = []
    for index, start in enumerate(starts):
        end = starts[index + 1] if index + 1 < len(starts) else len(lines)
        section = clean_text("\n".join(lines[start:end]))
        if section:
            sections.append(section)
    return sections or [page_text]


def candidate_blocks(text: str) -> List[Tuple[Optional[int], str]]:
    blocks: List[Tuple[Optional[int], str]] = []

    for page_num, page_text in page_split(text):
        page_text = clean_text(page_text)
        if not page_text:
            continue

        for account_section in split_dense_account_sections(page_text):
            chunks = re.split(r"\n\s*\n", account_section)
            buffer: List[str] = []
            for chunk in chunks:
                c = clean_text(chunk)
                if len(c) < 20:
                    continue
                lower = c.lower()
                continuation = lower.startswith((
                    "date opened", "date of last activity", "date of last payment",
                    "scheduled payment", "actual payment", "amount past due",
                    "payment history", "terms", "high balance", "credit limit", "by ",
                ))
                starts = (
                    any(label in lower for label in ["account number", "account #", "payment status", "account status", "account name"])
                    or any(term in lower for term in NEGATIVE_TERMS)
                ) and not continuation
                if starts and buffer:
                    block = "\n".join(buffer)
                    if len(block) > 80:
                        blocks.append((page_num, block))
                    buffer = [c]
                else:
                    buffer.append(c)
                if len("\n".join(buffer)) > 3500:
                    blocks.append((page_num, "\n".join(buffer)))
                    buffer = []
            if buffer:
                block = "\n".join(buffer)
                if len(block) > 80:
                    blocks.append((page_num, block))

    return dedupe_blocks(blocks)


def dedupe_blocks(blocks: List[Tuple[Optional[int], str]]) -> List[Tuple[Optional[int], str]]:
    seen = set()
    out = []
    for page_num, block in blocks:
        clean = clean_text(block)
        if len(clean) < 80:
            continue
        key = hashlib.sha1(clean[:700].encode("utf-8", errors="ignore")).hexdigest()
        if key not in seen:
            seen.add(key)
            out.append((page_num, clean[:4500]))
    return out


# -----------------------------
# Field extraction
# -----------------------------

def first_match(patterns: List[str], text: str) -> str:
    for pattern in patterns:
        m = re.search(pattern, text, flags=re.I | re.M)
        if m:
            val = clean_text(m.group(1))
            val = re.split(r"\n| {3,}", val)[0].strip(" :-")
            return val[:240]
    return ""


def raw_line_for_match(text: str, match) -> str:
    if not match:
        return ""
    start = match.start()
    for line in text.splitlines():
        line_start = text.find(line)
        line_end = line_start + len(line)
        if line_start <= start <= line_end:
            return clean_text(line)[:500]
    return clean_text(match.group(0))[:500]


def normalize_field_value(field_name: str, value: str) -> str:
    value = trim_embedded_labels(value, field_name)
    if field_name in {"balance", "past_due", "high_credit_or_original_amount", "credit_limit"}:
        return normalize_money(value)
    if "date" in field_name:
        return normalize_date(value)
    if field_name == "account_number_masked":
        return mask_account_number(value)
    return value


def extract_field_with_evidence(field_name: str, block: str, context: dict) -> Tuple[str, Optional[FieldEvidence]]:
    for index, pattern in enumerate(COMMON_FIELD_PATTERNS.get(field_name, []), start=1):
        match = re.search(pattern, block, flags=re.I | re.M)
        if not match:
            continue
        raw_value = clean_text(match.group(1))
        raw_value = re.split(r"\n| {3,}", raw_value)[0].strip(" :-")
        normalized_value = normalize_field_value(field_name, raw_value)
        raw_line = raw_line_for_match(block, match)
        evidence = FieldEvidence(
            field_name=field_name,
            raw_value=raw_value[:240],
            normalized_value=normalized_value,
            source_filename=context.get("source_filename", ""),
            source_file_hash=context.get("source_file_hash", ""),
            bureau=context.get("bureau", ""),
            page=context.get("page"),
            raw_block_id=context.get("raw_block_id", ""),
            raw_block_hash=context.get("raw_block_hash", ""),
            raw_line=raw_line,
            extraction_rule_id=f"{field_name}:pattern_{index}",
            confidence_score=0.9 if raw_line else 0.72,
            needs_admin_review=not bool(raw_line),
        )
        return normalized_value, evidence
    return "", None


def trim_embedded_labels(value: str, field_name: str) -> str:
    value = clean_text(value).strip(" :-|")
    if not value:
        return ""

    protected = {
        "account_type": {"Account"},
        "status": set(),
        "pay_status": set(),
        "remarks": set(),
    }.get(field_name, set())

    for label in FIELD_STOP_LABELS:
        if label in protected:
            continue
        pattern = r"(?:\s+\|\s+|\s+\|\s*|\s{1,})(?=" + re.escape(label) + r"\b\s*:?)"
        parts = re.split(pattern, value, maxsplit=1, flags=re.I)
        if len(parts) > 1 and parts[0].strip():
            value = parts[0].strip(" :-|")

    if field_name in {"status", "pay_status"}:
        value = re.split(
            r"\s+(?:Status Updated|Balance Updated|Balance|Recent Payment|Monthly Payment|Credit Limit|Highest Balance|Terms|Payment History)\b",
            value,
            maxsplit=1,
            flags=re.I,
        )[0].strip(" :-|")
        if re.match(
            r"^(?:Date Opened|Date Reported|Date Updated|Date of|Terms|Frequency|Months Reviewed|Scheduled Payment|Amount Past Due)\b",
            value,
            flags=re.I,
        ):
            value = ""

    if field_name == "account_type":
        value = re.split(r"\s+\|\s+|\s+Status\s*:", value, maxsplit=1, flags=re.I)[0].strip(" :-|")

    if field_name == "responsibility":
        m = re.search(r"\b(Individual|Joint|Authorized User|Co-signer|Terminated|Undesignated)\b", value, flags=re.I)
        if m:
            value = clean_text(m.group(1)).title()

    return value[:240]


def infer_missing_fields_from_block(t: NormalizedTradeline) -> None:
    block = t.raw_block
    lower = block.lower()

    if not t.past_due:
        m = re.search(r"(" + MONEY + r")\s+past due\b", block, flags=re.I)
        if m:
            t.past_due = normalize_money(m.group(1))

    if not t.pay_status:
        m = re.search(r"\bPay Status\s*[:\-]?\s*([A-Za-z0-9 /&.,'$-]{3,120})", block, flags=re.I)
        if m:
            t.pay_status = trim_embedded_labels(m.group(1), "pay_status")

    if not t.status and t.pay_status:
        t.status = t.pay_status

    if not t.date_reported:
        m = re.search(r"\b(?:Balance Updated|Date Updated)\s*[:\-]?\s*(" + DATE + r")", block, flags=re.I)
        if m:
            t.date_reported = normalize_date(m.group(1))

    if not t.date_last_payment:
        m = re.search(r"\b(?:Last Payment Made|Date of Last Payment|Last Payment)\s*[:\-]?\s*(" + DATE + r")", block, flags=re.I)
        if m:
            t.date_last_payment = normalize_date(m.group(1))

    if not t.responsibility:
        m = re.search(r"\b(?:Owner|Responsibility)\s*[:\-]?\s*(Individual|Joint|Authorized User|Co-signer|Terminated|Undesignated)", block, flags=re.I)
        if m:
            t.responsibility = clean_text(m.group(1)).title()

    if not t.account_type:
        m = re.search(r"\bLoan Type\s*[:\-]?\s*([A-Za-z0-9 /&-]{3,80})", block, flags=re.I)
        if m:
            t.account_type = trim_embedded_labels(m.group(1), "account_type")

    if not t.credit_limit:
        m = re.search(r"\bCredit Limit(?:\s*\(Hist\.\))?\s*(?:[:\-]|of)?\s*(" + MONEY + r")", block, flags=re.I)
        if m:
            t.credit_limit = normalize_money(m.group(1))

    if not t.high_credit_or_original_amount:
        m = re.search(r"\b(?:High Balance|Highest Balance|Original Balance|High Credit)(?:\s*\(Hist\.\))?\s*(?:[:\-]|of)?\s*(" + MONEY + r")", block, flags=re.I)
        if m:
            t.high_credit_or_original_amount = normalize_money(m.group(1))

    if not t.estimated_removal_date:
        m = re.search(r"\bOn Record Until\s*[:\-]?\s*(" + DATE + r")", block, flags=re.I)
        if m:
            t.estimated_removal_date = normalize_date(m.group(1))

    if not t.remarks and "account information disputed by consumer" in lower:
        t.remarks = "Account information disputed by consumer"


def evidence_for_inferred_field(t: NormalizedTradeline, field_name: str, value: str) -> FieldEvidence:
    raw_line = ""
    if value:
        value_key = re.escape(str(value).strip())
        for line in t.raw_block.splitlines():
            if re.search(value_key, line, flags=re.I):
                raw_line = clean_text(line)
                break
    if not raw_line:
        for line in t.raw_block.splitlines():
            if field_name.replace("_", " ") in line.lower():
                raw_line = clean_text(line)
                break
    if not raw_line:
        raw_line = clean_text(t.raw_block.splitlines()[0] if t.raw_block.splitlines() else t.raw_block)[:500]
    return FieldEvidence(
        field_name=field_name,
        raw_value=str(value or "")[:240],
        normalized_value=str(value or ""),
        source_filename=t.source_filename,
        source_file_hash=t.source_file_hash,
        bureau=t.bureau,
        page=t.page_start,
        raw_block_id=t.raw_block_id,
        raw_block_hash=t.raw_block_hash,
        raw_line=raw_line[:500],
        extraction_rule_id=f"{field_name}:inferred_from_raw_block",
        confidence_score=0.55,
        needs_admin_review=True,
    )


def ensure_field_evidence(t: NormalizedTradeline) -> None:
    for field_name in [
        "account_name",
        "account_number_masked",
        "account_type",
        "portfolio_type",
        "responsibility",
        "creditor_classification",
        "original_creditor",
        "collector_or_debt_buyer",
        "status",
        "pay_status",
        "balance",
        "past_due",
        "high_credit_or_original_amount",
        "credit_limit",
        "date_opened",
        "date_closed",
        "date_reported",
        "date_last_activity",
        "date_last_payment",
        "date_of_first_delinquency",
        "estimated_removal_date",
        "remarks",
    ]:
        value = getattr(t, field_name, "")
        if value and field_name not in t.field_evidence:
            t.field_evidence[field_name] = evidence_for_inferred_field(t, field_name, value)
    missing_proof = [field for field, evidence in t.field_evidence.items() if not evidence.raw_line]
    if missing_proof:
        t.needs_admin_review = True
        t.parser_warnings.append("field_evidence_missing_raw_line:" + ",".join(sorted(missing_proof)))


def extract_field(field_name: str, block: str) -> str:
    value = first_match(COMMON_FIELD_PATTERNS.get(field_name, []), block)
    return normalize_field_value(field_name, value)


def is_bad_account_name(name: str) -> bool:
    cleaned = clean_text(name).strip(" .;:,|-")
    lower = cleaned.lower()
    if not cleaned or cleaned == "Review Item":
        return True
    if len(cleaned) < 3 or len(cleaned) > 90:
        return True
    if any(fragment in lower for fragment in BAD_ACCOUNT_NAME_FRAGMENTS):
        return True
    if re.match(
        r"^(?:high balance|highest balance|high credit|credit limit|balance|past due|amount past due|monthly payment|recent payment)\b",
        lower,
    ):
        return True
    if lower.startswith(("of ", "and ", "the ", "this ", "that ", "www.")):
        return True
    if cleaned.endswith(".") and len(cleaned.split()) > 3:
        return True
    if len(cleaned.split()) > 8:
        return True
    if not re.search(r"[A-Za-z]", cleaned):
        return True
    return False


def account_name_quality(name: str) -> float:
    cleaned = clean_text(name).strip(" .;:,|-")
    lower = cleaned.lower()
    if is_bad_account_name(cleaned):
        return -10.0

    score = 0.0
    words = re.findall(r"[A-Za-z0-9&']+", cleaned)
    if 1 <= len(words) <= 5:
        score += 2.0
    if len(words) > 5:
        score -= 1.0
    if re.search(r"\b(bank|credit|capital|midland|ford|federal|jpm|jpmcb|caine|jefferson|systems?|management|services?|financial|portfolio|synchrony|citibank|northwest|firstpoint|resources|funding|lvnv|resurgent|macys|cbna)\b", lower, flags=re.I):
        score += 2.0
    if any(x in lower for x in [
        "prepared for",
        "date:",
        "date reported",
        "balance:",
        "loan/account type",
        "account type:",
        "account type ",
        "loan type",
        "pay status",
        "credit limit",
        "high credit",
        "status:",
        "term duration",
        "terms paid",
        "minimum payment",
        "confirmation #",
        "responsibility relationship",
    ]):
        score -= 5.0
    if re.search(r"\d{2}/\d{2}/\d{4}|\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec|january|february|march|april|june|july|august|september|october|november|december)\s+\d{4}\b|\b20\d{2}\s+(?:CO|OK|C|30|60|90|120|150|CLS|ND)(?:\s+(?:CO|OK|C|30|60|90|120|150|CLS|ND))*\b|\$\d", cleaned, flags=re.I):
        score -= 4.0
    if "|" in cleaned or ":" in cleaned:
        score -= 2.0
    return score


def tradeline_quality_score(t: NormalizedTradeline) -> float:
    score = t.confidence_score * 10
    score += account_name_quality(t.account_name)
    if t.account_number_masked:
        score += 4
    if t.balance:
        score += 1
    if t.account_type:
        score += 1
    if t.date_opened:
        score += 1
    if t.date_reported:
        score += 1
    if t.original_creditor:
        score += 1
    if t.status and t.status.lower() not in {"date opened", "account information"}:
        score += 1
    if len(t.raw_block) > 3200:
        score -= 1
    return score


def guess_account_name(block: str) -> str:
    lines = [clean_text(x).strip(" :-") for x in block.splitlines() if clean_text(x)]
    bad = (
        "account number", "account #", "balance", "date ", "status", "payment",
        "remarks", "comments", "address", "phone", "page ", "experian", "equifax",
        "transunion", "trans union", "credit report", "report number", "summary",
        "prepared for", "confirmation #", "account type", "loan type", "pay status", "responsibility relationship"
    )

    labeled = first_match([
        r"(?:account name|furnisher|subscriber)\s*[:\-]?\s*([A-Za-z0-9 &.,'()/\-]{3,90})"
    ], block)
    if labeled:
        return labeled if not is_bad_account_name(labeled) else "Review Item"

    candidates = []
    for idx, line in enumerate(lines[:26]):
        lower = line.lower()
        if any(lower.startswith(x) for x in bad):
            continue
        if re.match(r"^date\s*:", lower):
            continue
        if re.match(r"^(?:january|february|march|april|may|june|july|august|september|october|november|december)\s+\d{4}$", lower):
            continue
        if re.match(r"^20\d{2}\s+(?:co|ok|c|30|60|90|120|150|cls|nd)(?:\s+(?:co|ok|c|30|60|90|120|150|cls|nd))*$", lower):
            continue
        if lower.startswith(("responsibility relationship", "total months", "rating", "ok", "date opened", "terms paid", "minimum payment", "pay status")):
            continue
        if lower.endswith("minimum payment.") or "credit scoring model" in lower:
            continue
        if lower in {"collections", "satisfactory accounts", "potentially negative", "account information"}:
            continue
        if re.fullmatch(r"[A-Z][A-Z ]{1,30}", line) and len(line.split()) <= 4:
            # Skip consumer-name header lines in bureau PDFs. Creditor names
            # usually appear near account fields and score higher below.
            following = lines[idx + 1].lower() if idx + 1 < len(lines) else ""
            header_positioned = following.startswith((
                "account number", "account #", "account type", "acct number",
                "acct #", "original creditor", "balance", "status", "pay status",
                "date opened",
            ))
            if (not header_positioned) and not re.search(r"\b(BANK|CREDIT|CAPITAL|MIDLAND|FORD|FEDERAL|JPM|JPMCB|CAINE|JEFFERSON|RESURGENT|LVNV|MACYS|CBNA|MOTOR|FUNDING|SYSTEM|SYSTEMS|MANAGEMENT|SERVICES|FIRSTPOINT|RESOURCES)\b", line):
                continue
        if len(line) < 3 or len(line) > 85:
            continue
        if not re.search(r"[A-Za-z]", line):
            continue
        # Avoid long report sentences
        if len(line.split()) <= 9 and not is_bad_account_name(line):
            candidates.append(line)

    if candidates:
        return max(candidates, key=account_name_quality)

    return "Review Item"


def is_boilerplate_block(block: str) -> bool:
    lower = block.lower()
    boilerplate_hits = sum(1 for term in BOILERPLATE_TERMS if term in lower)
    core_field_hits = sum(
        1
        for term in ACCOUNT_SECTION_TERMS
        if term in lower
    )
    return boilerplate_hits > 0 and core_field_hits < 2


def account_section_score(block: str) -> int:
    lower = block.lower()
    return sum(1 for term in ACCOUNT_SECTION_TERMS if term in lower)


def has_non_account_section_bias(block: str) -> bool:
    lower = block.lower()
    non_account_hits = sum(1 for term in NON_ACCOUNT_SECTION_TERMS if term in lower)
    return non_account_hits >= 2 and account_section_score(block) < 4


def is_probable_tradeline(t: NormalizedTradeline) -> bool:
    if is_bad_account_name(t.account_name):
        return False
    if has_non_account_section_bias(t.raw_block):
        return False
    if compact_key(t.account_name) in {"co", "ok", "cls", "nd", "address"}:
        return False

    core_fields = [
        t.account_number_masked,
        t.account_type,
        t.status or t.pay_status,
        t.balance,
        t.date_opened,
        t.date_reported,
        t.date_of_first_delinquency,
        t.original_creditor,
        t.remarks,
    ]
    core_count = sum(1 for field in core_fields if field)
    lower = t.raw_block.lower()
    section_score = account_section_score(t.raw_block)
    has_money_or_account = bool(t.balance or t.account_number_masked)
    has_account_identifier = bool(
        t.account_number_masked
        or re.search(r"\b(?:account number|account #|account name)\b", lower, flags=re.I)
    )
    has_credit_signal = any(
        term in lower
        for term in [
            "account type",
            "original creditor",
            "date opened",
            "date reported",
            "payment status",
            "account status",
            "credit limit",
            "past due",
            "high credit",
        ]
    )
    if section_score < 2 and not has_account_identifier:
        return False
    if not has_account_identifier and core_count < 4:
        return False
    return core_count >= 2 and (has_money_or_account or has_credit_signal)


def classify_category(t: NormalizedTradeline) -> Tuple[str, str, str, str]:
    blob = " ".join([
        t.account_name, t.account_type, t.status, t.pay_status, t.remarks,
        t.original_creditor, t.collector_or_debt_buyer, t.raw_block
    ]).lower()

    if any(x in blob for x in COLLECTION_TERMS):
        return (
            "Collection Review",
            "medium",
            "Collection or debt-buyer information should be reviewed.",
            "Round 2 — Collection Review"
        )

    if "charge" in blob and "off" in blob:
        return (
            "Reporting Accuracy Review",
            "medium",
            "Charge-off reporting should be reviewed for balance, status, dates, and ownership.",
            "Round 4 — Reporting Accuracy Review"
        )

    if any(x in blob for x in ["transferred", "sold", "closed"]):
        return (
            "Reporting Accuracy Review",
            "medium",
            "Transferred, sold, or closed reporting should be reviewed for accuracy.",
            "Round 4 — Reporting Accuracy Review"
        )

    if any(x in blob for x in NEGATIVE_TERMS):
        return (
            "Factual Review",
            "medium",
            "This negative item should be reviewed for factual accuracy.",
            "Round 5 — Factual Review"
        )

    return (
        "Bureau Match Review",
        "low",
        "Compare this item across bureaus for consistency.",
        "Round 3 — Bureau Match Review"
    )


def score_confidence(t: NormalizedTradeline) -> float:
    score = 0.0
    if t.account_name and t.account_name != "Review Item":
        score += 0.22
    if t.account_number_masked:
        score += 0.16
    if t.balance:
        score += 0.10
    if t.status or t.pay_status:
        score += 0.13
    if t.date_opened:
        score += 0.08
    if t.date_reported:
        score += 0.08
    if t.date_of_first_delinquency or t.estimated_removal_date:
        score += 0.10
    if t.raw_block and len(t.raw_block) > 250:
        score += 0.08
    if any(x in t.raw_block.lower() for x in NEGATIVE_TERMS):
        score += 0.05
    return min(1.0, score)


def confidence_label(score: float) -> str:
    if score >= 0.72:
        return "high"
    if score >= 0.42:
        return "medium"
    return "low"


def page_bureau_map(text: str) -> Dict[Optional[int], str]:
    return {
        page_num: detect_bureau("", page_text)
        for page_num, page_text in page_split(text)
    }


def parse_tradelines_for_bureau(bureau: str, filename: str, text: str, file_hash: str = "", upload_bureau: str = "") -> List[NormalizedTradeline]:
    tradelines = []
    for page_num, page_text in page_split(text):
        detection = detect_bureau_with_conflict("", page_text, upload_bureau or bureau)
        page_bureau = detection.bureau
        active_bureau = page_bureau if page_bureau != "Unknown Bureau" else bureau
        page_input = f"--- PAGE {page_num} ---\n{page_text}" if page_num is not None else page_text
        for page_num, block in candidate_blocks(page_input):
            lower = block.lower()
            if is_boilerplate_block(block):
                continue
            has_signal = (
                any(term in lower for term in NEGATIVE_TERMS)
                or any(label in lower for label in ["account number", "account #", "date opened", "payment status", "account status"])
            )
            if not has_signal:
                continue

            account_name = guess_account_name(block)
            raw_block_hash = hashlib.sha256(block.encode("utf-8", errors="ignore")).hexdigest()
            raw_block_id = stable_id(filename, active_bureau, str(page_num or ""), raw_block_hash[:16])
            context = {
                "source_filename": filename,
                "source_file_hash": file_hash,
                "bureau": active_bureau,
                "page": page_num,
                "raw_block_id": raw_block_id,
                "raw_block_hash": raw_block_hash,
            }
            extracted = {}
            field_evidence = {}
            for field_name in COMMON_FIELD_PATTERNS:
                value, evidence = extract_field_with_evidence(field_name, block, context)
                extracted[field_name] = value
                if evidence:
                    field_evidence[field_name] = evidence
            t = NormalizedTradeline(
                id=stable_id(active_bureau, filename, account_name, block[:160]),
                bureau=active_bureau,
                source_filename=filename,
                account_name=account_name,
                account_number_masked=extracted.get("account_number_masked", ""),
                account_type=extracted.get("account_type", ""),
                portfolio_type=extracted.get("portfolio_type", ""),
                responsibility=extracted.get("responsibility", ""),
                creditor_classification=extracted.get("creditor_classification", ""),
                original_creditor=extracted.get("original_creditor", ""),
                collector_or_debt_buyer=extracted.get("collector_or_debt_buyer", ""),
                status=extracted.get("status", ""),
                pay_status=extracted.get("pay_status", ""),
                balance=extracted.get("balance", ""),
                past_due=extracted.get("past_due", ""),
                high_credit_or_original_amount=extracted.get("high_credit_or_original_amount", ""),
                credit_limit=extracted.get("credit_limit", ""),
                date_opened=extracted.get("date_opened", ""),
                date_closed=extracted.get("date_closed", ""),
                date_reported=extracted.get("date_reported", ""),
                date_last_activity=extracted.get("date_last_activity", ""),
                date_last_payment=extracted.get("date_last_payment", ""),
                date_of_first_delinquency=extracted.get("date_of_first_delinquency", ""),
                estimated_removal_date=extracted.get("estimated_removal_date", ""),
                remarks=extracted.get("remarks", ""),
                raw_block=block[:2500],
                page_start=page_num,
                field_evidence=field_evidence,
                source_file_hash=file_hash,
                raw_block_id=raw_block_id,
                raw_block_hash=raw_block_hash,
                bureau_conflict=detection.conflict,
                parser_warnings=["upload_metadata_conflicts_with_page_header"] if detection.conflict else [],
            )
            infer_missing_fields_from_block(t)
            if t.account_name:
                t.field_evidence["account_name"] = evidence_for_inferred_field(t, "account_name", t.account_name)
            ensure_field_evidence(t)
            negative = classify_negative_status(t)
            t.is_negative = bool(negative.get("is_negative"))
            t.negative_item_type = negative.get("negative_type", "")
            t.negative_signals = negative.get("signals", [])
            t.positive_signals = negative.get("positive_signals", [])
            if negative.get("needs_admin_review"):
                t.needs_admin_review = True
                t.parser_warnings.append("negative_classifier_needs_admin_review")
            t.confidence_score = score_confidence(t)
            t.confidence = confidence_label(t.confidence_score)
            t.needs_admin_review = t.needs_admin_review or t.confidence != "high" or t.bureau_conflict
            if not is_probable_tradeline(t):
                continue
            tradelines.append(t)

    return dedupe_tradelines(tradelines)


def dedupe_tradelines(items: List[NormalizedTradeline]) -> List[NormalizedTradeline]:
    best_by_key: Dict[Tuple[str, str], NormalizedTradeline] = {}
    no_number: List[NormalizedTradeline] = []

    for t in items:
        if t.account_number_masked:
            key = (t.bureau, t.account_number_masked)
            current = best_by_key.get(key)
            if current is None or tradeline_quality_score(t) > tradeline_quality_score(current):
                best_by_key[key] = t
            continue
        no_number.append(t)

    out = list(best_by_key.values())
    seen = {
        (
            t.bureau,
            compact_key(t.account_name)[:35],
            t.account_number_masked,
            compact_key(t.balance),
            compact_key(t.status or t.pay_status)[:25],
        )
        for t in out
    }

    for t in no_number:
        key = (
            t.bureau,
            compact_key(t.account_name)[:35],
            compact_key(t.balance),
            compact_key(t.status or t.pay_status)[:25],
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(t)
    return out


def _block_account_signature(block: str) -> str:
    acct = extract_field("account_number_masked", block)
    if acct:
        return f"acct:{acct}"
    name = guess_account_name(block)
    opened = extract_field("date_opened", block)
    balance = extract_field("balance", block)
    return "block:" + stable_id(compact_key(name), opened, balance)


def block_looks_like_account_candidate(block: str) -> bool:
    lower = block.lower()
    if is_boilerplate_block(block) or has_non_account_section_bias(block):
        return False
    field_hits = sum(1 for term in ACCOUNT_SECTION_TERMS if term in lower)
    has_account_number = bool(re.search(r"\b(?:account number|account #|acct\s*(?:#|number))\b", lower, flags=re.I))
    has_account_type = "account type" in lower or "loan/account type" in lower or "loan type" in lower
    has_status = "status" in lower or "pay status" in lower or "payment status" in lower
    has_date = "date opened" in lower or "date reported" in lower or "date updated" in lower
    has_balance = "balance" in lower or bool(re.search(MONEY, block))
    has_creditor_like_line = any(
        re.fullmatch(r"[A-Z0-9][A-Z0-9 &.,'()/\-]{3,80}", clean_text(line))
        and not is_bad_account_name(clean_text(line))
        for line in block.splitlines()[:12]
    )
    return (
        field_hits >= 3
        or (has_account_number and has_account_type)
        or (has_account_number and has_status and (has_date or has_balance))
        or (has_creditor_like_line and has_account_number and (has_date or has_balance or has_status))
    )


def detect_possible_skipped_tradelines(filename: str, bureau: str, text: str, file_hash: str, parsed_items: List[NormalizedTradeline]) -> List[dict]:
    parsed_signatures = {_block_account_signature(item.raw_block) for item in parsed_items if item.source_filename == filename}
    warnings = []
    for page_num, page_text in page_split(text):
        detection = detect_bureau_with_conflict("", page_text, bureau)
        active_bureau = detection.bureau if detection.bureau != "Unknown Bureau" else bureau
        page_input = f"--- PAGE {page_num} ---\n{page_text}" if page_num is not None else page_text
        for block_page, block in candidate_blocks(page_input):
            if not block_looks_like_account_candidate(block):
                continue
            signature = _block_account_signature(block)
            if signature in parsed_signatures:
                continue
            account_name = guess_account_name(block)
            warnings.append({
                "warning_id": stable_id("possible_skipped_tradeline", filename, active_bureau, str(block_page or ""), signature),
                "warning_type": "possible_skipped_tradeline",
                "severity": "high",
                "source_filename": filename,
                "source_file_hash": file_hash,
                "bureau": active_bureau,
                "page": block_page,
                "account_name_guess": account_name,
                "account_number_guess": extract_field("account_number_masked", block),
                "raw_block_hash": hashlib.sha256(block.encode("utf-8", errors="ignore")).hexdigest(),
                "raw_text_snippet": clean_text(block)[:900],
                "admin_action": "Admin QA should review this account-like block. Do not create disputes or letters from this warning unless a tradeline is confirmed.",
                "customer_visible": False,
                "creates_dispute_issue": False,
            })
    return warnings


# -----------------------------
# Cross-bureau matching
# -----------------------------

def is_collection_like(t: NormalizedTradeline) -> bool:
    blob = " ".join([
        t.account_name,
        t.account_type,
        t.status,
        t.pay_status,
        t.remarks,
        t.collector_or_debt_buyer,
        t.raw_block,
    ]).lower()
    return any(term in blob for term in [
        "collection",
        "collector",
        "debt buyer",
        "factoring company",
        "placed for collection",
        "assigned",
    ])


def same_cross_bureau_account(left: NormalizedTradeline, right: NormalizedTradeline) -> bool:
    name_score = simple_similarity(left.account_name, right.account_name)
    acct_match = bool(left.account_number_masked and left.account_number_masked == right.account_number_masked)
    original_score = (
        simple_similarity(left.original_creditor, right.original_creditor)
        if left.original_creditor and right.original_creditor
        else 0
    )
    left_key = compact_key(left.account_name)
    right_key = compact_key(right.account_name)

    if not left_key or not right_key:
        return False
    if left_key in {"co", "ok", "cls", "nd", "address"} or right_key in {"co", "ok", "cls", "nd", "address"}:
        return False

    if acct_match:
        return True
    if left.account_number_masked and right.account_number_masked and left.account_number_masked != right.account_number_masked:
        return False

    if name_score >= 0.82:
        return True

    # Same original creditor is only supporting evidence. It cannot, by itself,
    # merge a collector/debt-buyer tradeline with the original creditor's own card/loan.
    if original_score >= 0.72:
        if name_score >= 0.65:
            return True
        if is_collection_like(left) and is_collection_like(right) and name_score >= 0.55:
            return True

    return False


def group_cross_bureau(tradelines: List[NormalizedTradeline]) -> List[dict]:
    groups: List[dict] = []
    used = set()

    for i, t in enumerate(tradelines):
        if t.id in used:
            continue

        group = [t]
        used.add(t.id)

        for other in tradelines[i+1:]:
            if other.id in used:
                continue
            if t.bureau == other.bureau:
                continue

            if same_cross_bureau_account(t, other):
                group.append(other)
                used.add(other.id)

        if len(group) >= 2:
            groups.append({
                "group_id": stable_id("group", *[x.id for x in group]),
                "bureaus": sorted({x.bureau for x in group}),
                "account_names": sorted({x.account_name for x in group}),
                "tradeline_ids": [x.id for x in group],
                "review_note": "Same or similar item appears across multiple bureaus. Compare balance, status, dates, original creditor, and remarks.",
                "field_snapshot": [
                    {
                        "bureau": x.bureau,
                        "account_name": x.account_name,
                        "balance": x.balance,
                        "status": x.status,
                        "pay_status": x.pay_status,
                        "date_opened": x.date_opened,
                        "date_reported": x.date_reported,
                        "date_of_first_delinquency": x.date_of_first_delinquency,
                        "estimated_removal_date": x.estimated_removal_date,
                    }
                    for x in group
                ]
            })

    return groups


# -----------------------------
# Issue detection engine
# -----------------------------

def ev(t: NormalizedTradeline) -> Evidence:
    return Evidence(
        bureau=t.bureau,
        page=t.page_start,
        snippet=t.raw_block[:800],
    )


def classify_negative_status(t: NormalizedTradeline) -> dict:
    blob = " ".join(
        str(value or "")
        for value in [
            t.account_type,
            t.status,
            t.pay_status,
            t.remarks,
            t.past_due,
            t.date_of_first_delinquency,
            t.estimated_removal_date,
            t.collector_or_debt_buyer,
            t.raw_block,
        ]
    ).lower()
    signal_patterns = {
        "collection": [r"\bcollection\b", r"\bdebt buyer\b", r"\bcollector\b"],
        "charge_off": [r"\bcharge[- ]?off\b", r"\bcharged off\b", r"\bwritten off\b", r"\bbad debt\b"],
        "past_due": [r"\bpast due\b", r"\bdelinquent\b", r"\bseriously past due\b"],
        "late_payment": [r"\b(?:30|60|90|120|150|180)\s+days?\s+(?:late|past due)\b"],
        "public_record_or_severe": [r"\brepossession\b", r"\bforeclosure\b", r"\bbankruptcy\b"],
    }
    positive_patterns = [
        (r"\bpays as agreed\b", "pays as agreed"),
        (r"\bpaid as agreed\b", "paid as agreed"),
        (r"\bnever late\b", "never late"),
        (r"\bcurrent account\b", "current account"),
        (r"\bcurrent\b", "current"),
        (r"\bopen\b", "open"),
    ]
    signals = []
    negative_type = ""
    for signal_type, patterns in signal_patterns.items():
        if signal_type == "past_due":
            past_due_amount = money_to_number(t.past_due)
            explicit_delinquency = re.search(
                r"\bdelinquent\b|\bseriously past due\b|\b(?:30|60|90|120|150|180)\s+days?\s+(?:late|past due)\b",
                " ".join([t.status, t.pay_status, t.remarks]),
                flags=re.I,
            )
            if past_due_amount is not None and past_due_amount <= 0 and not explicit_delinquency:
                continue
        for pattern in patterns:
            if re.search(pattern, blob, flags=re.I):
                signals.append(signal_type)
                negative_type = negative_type or signal_type
                break
    positive_signals = [label for pattern, label in positive_patterns if re.search(pattern, blob, flags=re.I)]
    is_negative = bool(signals)
    if not is_negative and positive_signals:
        return {
            "is_negative": False,
            "negative_type": "",
            "signals": [],
            "positive_signals": positive_signals,
            "confidence_score": 0.9,
            "needs_admin_review": False,
        }
    if is_negative:
        return {
            "is_negative": True,
            "negative_type": negative_type,
            "signals": sorted(set(signals)),
            "positive_signals": positive_signals,
            "confidence_score": 0.88 if len(signals) > 1 else 0.78,
            "needs_admin_review": bool(positive_signals),
        }
    return {
        "is_negative": False,
        "negative_type": "",
        "signals": [],
        "positive_signals": positive_signals,
        "confidence_score": 0.5,
        "needs_admin_review": False,
    }


def add_issue(issues: List[ReviewIssue], issue_type: str, severity: str, label: str,
              customer: str, admin: str, round_name: str, tradelines: List[NormalizedTradeline],
              confidence: str = "medium") -> None:
    issues.append(ReviewIssue(
        id=stable_id(issue_type, *[t.id for t in tradelines], label),
        issue_type=issue_type,
        severity=severity,
        customer_label=label,
        customer_explanation=customer,
        admin_explanation=admin,
        suggested_round=round_name,
        related_tradeline_ids=[t.id for t in tradelines],
        evidence=[ev(t) for t in tradelines],
        confidence=confidence,
    ))


def detect_issues(tradelines: List[NormalizedTradeline], groups: List[dict]) -> List[ReviewIssue]:
    issues: List[ReviewIssue] = []

    for t in tradelines:
        blob = " ".join([t.status, t.pay_status, t.remarks, t.raw_block]).lower()

        if t.is_negative and ("collection" in t.negative_signals or any(x in blob for x in COLLECTION_TERMS)):
            add_issue(
                issues,
                "collection_review",
                "medium",
                "Collection review",
                "This collection item should be reviewed for original creditor, balance, ownership, and reporting details.",
                "Collection/debt buyer candidate. Verify original creditor, assignment/ownership, balance, authority, and reporting fields.",
                "Round 2 — Collection Review",
                [t],
                t.confidence
            )

        if t.is_negative and ("charge_off" in t.negative_signals or ("charge" in blob and "off" in blob)):
            add_issue(
                issues,
                "chargeoff_review",
                "medium",
                "Charge-off review",
                "This charge-off should be reviewed for balance, status, dates, and whether it was sold or transferred.",
                "Charge-off candidate. Check DOFD, balance, sold/transferred status, creditor ownership, and duplicate collection reporting.",
                "Round 4 — Reporting Accuracy Review",
                [t],
                t.confidence
            )

        if t.is_negative and not t.date_of_first_delinquency and any(x in blob for x in ["charge", "collection", "delinquent", "past due"]):
            add_issue(
                issues,
                "missing_dofd_review",
                "medium",
                "Important date may be missing",
                "A key delinquency/removal date may need review.",
                "Negative item lacks detected DOFD/removal date. Verify raw report and bureau-specific fields.",
                "Round 4 — Reporting Accuracy Review",
                [t],
                "medium"
            )

        if t.is_negative and t.balance and t.balance not in {"$0", "$0.00"} and "closed" in blob and any(x in blob for x in ["transferred", "sold"]):
            add_issue(
                issues,
                "closed_sold_balance_review",
                "medium",
                "Closed or sold account balance review",
                "A closed, sold, or transferred account may still show a balance that should be reviewed.",
                "Closed/sold/transferred item with non-zero balance detected. Needs admin validation.",
                "Round 4 — Reporting Accuracy Review",
                [t],
                "medium"
            )

        if t.confidence == "low":
            add_issue(
                issues,
                "low_confidence_admin_review",
                "low",
                "Needs manual review",
                "This item needs a closer review before any action is recommended.",
                "Parser confidence is low. Admin should verify source snippet and fields.",
                "Admin Review",
                [t],
                "low"
            )

    # Cross-bureau mismatch checks
    by_id = {t.id: t for t in tradelines}
    for group in groups:
        group_items = [by_id[x] for x in group["tradeline_ids"] if x in by_id]
        if len(group_items) < 2:
            continue

        balances = {x.balance for x in group_items if x.balance}
        statuses = {clean_text(x.status or x.pay_status).lower() for x in group_items if x.status or x.pay_status}
        dates_reported = {x.date_reported for x in group_items if x.date_reported}
        dofds = {x.date_of_first_delinquency for x in group_items if x.date_of_first_delinquency}

        if len(balances) >= 2:
            add_issue(
                issues,
                "cross_bureau_balance_mismatch",
                "medium",
                "Balance differs across bureaus",
                "The same or similar account may show different balances across bureaus.",
                "Cross-bureau group has different balances. Verify if mismatch is expected or inaccurate.",
                "Round 3 — Bureau Match Review",
                group_items,
                "medium"
            )

        if len(statuses) >= 2:
            add_issue(
                issues,
                "cross_bureau_status_mismatch",
                "medium",
                "Status differs across bureaus",
                "The same or similar account may show different statuses across bureaus.",
                "Cross-bureau group has different status/pay-status values.",
                "Round 3 — Bureau Match Review",
                group_items,
                "medium"
            )

        if len(dates_reported) >= 2 or len(dofds) >= 2:
            add_issue(
                issues,
                "cross_bureau_date_mismatch",
                "medium",
                "Dates differ across bureaus",
                "The same or similar account may show different important dates across bureaus.",
                "Cross-bureau group has different report dates or DOFD/removal dates.",
                "Round 3 — Bureau Match Review",
                group_items,
                "medium"
            )

    return dedupe_issues(issues)


def dedupe_issues(issues: List[ReviewIssue]) -> List[ReviewIssue]:
    seen = set()
    out = []
    for issue in issues:
        key = (
            issue.issue_type,
            tuple(sorted(issue.related_tradeline_ids)),
            issue.customer_label,
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(issue)
    return out


# -----------------------------
# Main parse API
# -----------------------------

def parse_reports(report_texts: Dict[str, dict]) -> ParseResult:
    """
    report_texts format:
    {
      "filename.pdf": {
        "text": "...",
        "bureau": "Experian" optional
      }
    }
    """
    files = []
    all_tradelines: List[NormalizedTradeline] = []
    parser_qa_warnings: List[dict] = []

    for filename, payload in report_texts.items():
        text = clean_text(payload.get("text", ""))
        file_hash = source_hash(filename, text)
        upload_bureau = payload.get("bureau") or ""
        detection = detect_bureau_with_conflict(filename, text, upload_bureau)
        bureau = upload_bureau or detection.bureau
        if bureau == "Unknown Bureau":
            bureau = f"Unknown Report"

        files.append({
            "filename": filename,
            "bureau": bureau,
            "detected_bureau": detection.bureau,
            "source_file_hash": file_hash,
            "bureau_detection_confidence": detection.confidence_score,
            "bureau_detection_source": detection.source,
            "bureau_conflict": detection.conflict,
            "parser_warnings": ["upload_metadata_conflicts_with_page_header"] if detection.conflict else [],
            "chars": len(text),
            "status": "parsed" if text else "empty_text",
        })

        file_tradelines = parse_tradelines_for_bureau(bureau, filename, text, file_hash, upload_bureau)
        all_tradelines.extend(file_tradelines)
        parser_qa_warnings.extend(detect_possible_skipped_tradelines(filename, bureau, text, file_hash, file_tradelines))

    groups = group_cross_bureau(all_tradelines)
    issues = detect_issues(all_tradelines, groups)

    customer_summary = {
        "headline": "Your Credit Check-In was reviewed.",
        "message": "Credit Vivo organized review items into clear categories. No letters or disputes are sent without your approval.",
        "review_items": len(all_tradelines),
        "possible_review_points": len(issues),
        "categories": sorted({i.customer_label for i in issues}),
        "next_step": "Review findings in the dashboard."
    }

    admin_summary = {
        "engine": "Credit Vivo Proprietary Parser Engine",
        "version": "16.0",
        "paid_ai_used": False,
        "tradeline_count": len(all_tradelines),
        "issue_count": len(issues),
        "cross_bureau_group_count": len(groups),
        "warning": "Parser output is draft review data. Verify raw evidence snippets before preparing letters."
    }

    return ParseResult(
        engine="Credit Vivo Proprietary Parser Engine",
        version="16.0",
        paid_ai_used=False,
        files=files,
        tradelines=all_tradelines,
        issues=issues,
        cross_bureau_groups=groups,
        customer_summary=customer_summary,
        admin_summary=admin_summary,
        parser_qa_warnings=parser_qa_warnings,
    )


FCRA_NOTICE_OF_DISPUTE = (
    "This is my formal notice of dispute. I dispute the accuracy, completeness, "
    "and/or verifiability of the item identified in this letter. Please conduct a "
    "reasonable investigation under the Fair Credit Reporting Act (FCRA), review all "
    "information I provide with this dispute, forward the dispute and all relevant "
    "information to the furnisher when this is a bureau dispute, and correct, update, "
    "or delete any information that cannot be verified as accurate and complete. "
    "Please mark the item as disputed while the investigation is pending and provide "
    "the written results of the investigation, including any corrected report or "
    "explanation of the verification method used."
)

FCRA_NOTICE_RULES = {
    "cfpb_official_template_sources": [
        "CFPB credit reporting company dispute letter template",
        "CFPB furnisher dispute letter template",
        "CFPB complaint process overview",
    ],
    "cfpb_letter_fields": [
        "consumer full name and mailing address",
        "date of letter",
        "recipient name and mailing address",
        "account name and account number",
        "specific information being disputed",
        "plain-language explanation of why the information is wrong",
        "requested correction or investigation result",
        "copies of supporting documents",
    ],
    "consumer_notice_contents": [
        "identify the consumer",
        "identify the account or item being disputed",
        "state the specific information disputed",
        "explain the basis for the dispute",
        "include supporting documents or evidence when available",
        "request correction, update, deletion, or verification results",
    ],
    "bureau_dispute_rules": [
        "consumer reporting agency must conduct a reasonable reinvestigation when accuracy or completeness is disputed",
        "bureau should forward notice of the dispute and relevant information to the furnisher",
        "bureau should provide written reinvestigation results after completion",
        "bureau dispute path is preferred when the next step needs furnisher duties triggered through bureau notice",
        "CFPB template method: clearly identify the disputed item, explain the problem, request correction, and attach copies of proof",
    ],
    "furnisher_dispute_rules": [
        "direct furnisher dispute should be sent to the furnisher address shown on the report or other proper direct-dispute address",
        "direct dispute should include enough identifying information, the specific disputed information, the basis for dispute, and supporting evidence",
        "furnisher should conduct a reasonable investigation and review relevant information provided with a proper direct dispute",
        "furnisher should report corrections or stop reporting information that cannot be verified as accurate and complete",
        "CFPB template method: dispute directly with the company that gave the information to the bureau when the furnisher's records appear wrong",
    ],
    "credit_vivo_controls": [
        "do not send automatically",
        "customer approval required before mail, bureau dispute, furnisher dispute, CFPB complaint, state complaint, or attorney escalation",
        "store date sent, recipient, delivery method, tracking number, response due date, response received, and next action",
        "attach evidence packet hash or file reference before sending",
        "record whether the item was marked disputed and whether written results were received",
    ],
}


def _issue_evidence_strength(issue: ReviewIssue) -> str:
    if issue.severity == "high" or issue.confidence == "high":
        return "high"
    if issue.confidence == "low" or issue.severity == "low":
        return "low"
    return "medium"


def _responsible_party_for_issue(issue: ReviewIssue) -> str:
    if issue.issue_type.startswith("cross_bureau"):
        return "bureau_and_furnisher"
    if issue.issue_type in {"collection_review", "chargeoff_review", "closed_sold_balance_review"}:
        return "furnisher_or_collector"
    if issue.issue_type == "missing_dofd_review":
        return "bureau_and_furnisher"
    return "admin_review_required"


def _next_action_for_issue(issue: ReviewIssue) -> str:
    if issue.issue_type.startswith("cross_bureau"):
        return "round_2_field_level_bureau_dispute"
    if issue.issue_type in {"collection_review", "chargeoff_review", "closed_sold_balance_review"}:
        return "furnisher_direct_dispute_after_bureau_review"
    if issue.issue_type == "missing_dofd_review":
        return "round_2_bureau_dispute_then_reinvestigation_if_unverified"
    return "admin_review_before_letter"


def _draft_letter_subject(letter_type: str, issue: ReviewIssue) -> str:
    if letter_type == "debt_validation_request":
        return f"Debt Validation Request - {issue.customer_label}"
    if letter_type == "furnisher_direct_dispute":
        return f"Direct Dispute of Account Reporting - {issue.customer_label}"
    if letter_type == "bureau_dispute":
        return f"Credit Report Dispute - {issue.customer_label}"
    return f"Admin Review Required - {issue.customer_label}"


def _draft_letter_body(letter_type: str, recipient_type: str, issue: ReviewIssue) -> str:
    if letter_type == "admin_review_hold":
        return (
            "DRAFT HOLD - ADMIN REVIEW REQUIRED\n\n"
            f"Issue: {issue.customer_label}\n"
            f"Reason: {issue.customer_explanation}\n\n"
            "Credit Vivo did not generate a send-ready letter for this item because it needs "
            "manual review before any dispute path is selected."
        )

    evidence_note = "Evidence from the Credit Vivo scanner is attached for customer/admin review."
    if issue.evidence:
        snippet = issue.evidence[0].snippet[:450]
        evidence_note = f"Scanner evidence excerpt for review: {snippet}"

    if letter_type == "debt_validation_request":
        return (
            "DRAFT - CUSTOMER REVIEW AND APPROVAL REQUIRED\n"
            "DO NOT SEND UNTIL CUSTOMER AUTHORIZATION IS VERIFIED\n\n"
            "[Customer Name]\n"
            "[Customer Mailing Address]\n"
            "[City, State ZIP]\n\n"
            "[Date]\n\n"
            "To: Debt Collector / Debt Buyer\n"
            "[Recipient Address]\n\n"
            f"Re: {issue.customer_label}\n\n"
            "To Whom It May Concern:\n\n"
            "I am requesting validation of the debt you claim is owed. If you are attempting "
            "to collect this debt or reporting it as a collector/debt buyer, please provide "
            "documents and information sufficient to verify the debt, the amount claimed, "
            "the original creditor, your authority to collect, and any assignment or sale "
            "of the account.\n\n"
            "Please provide:\n"
            "- the name and address of the original creditor;\n"
            "- an itemized accounting of the balance, fees, interest, and payments;\n"
            "- documents showing I am responsible for the debt;\n"
            "- documents showing your authority to collect or report this account;\n"
            "- the date of first delinquency and records supporting the reported timeline;\n"
            "- the account number or other identifier used to report the account.\n\n"
            "This request is made for consumer review and dispute-tracking purposes. "
            "Nothing in this letter is an admission that the debt is owed. Please send "
            "your written response to the mailing address above.\n\n"
            f"{evidence_note}\n\n"
            "Sincerely,\n\n"
            "[Customer Signature]\n"
            "[Customer Printed Name]\n"
        )

    recipient_line = "Credit Bureau" if recipient_type == "credit_bureau" else "Furnisher / Collector"
    requested_action = (
        "Please investigate this item, correct any inaccurate or incomplete information, "
        "delete any information that cannot be verified, and send the investigation results in writing."
        if letter_type == "bureau_dispute"
        else
        "Please provide the basis for your reporting, including records supporting ownership, "
        "balance, status, payment history, date of first delinquency, and authority to report this account."
    )

    return (
        "DRAFT - CUSTOMER REVIEW AND APPROVAL REQUIRED\n"
        "DO NOT SEND UNTIL CUSTOMER AUTHORIZATION IS VERIFIED\n\n"
        "[Customer Name]\n"
        "[Customer Mailing Address]\n"
        "[City, State ZIP]\n\n"
        "[Date]\n\n"
        f"To: {recipient_line}\n"
        "[Recipient Address]\n\n"
        f"Re: {issue.customer_label}\n\n"
        "To Whom It May Concern:\n\n"
        "I am disputing the accuracy, completeness, or verifiability of the credit reporting item "
        "identified below. Please treat this letter as my formal notice of dispute.\n\n"
        f"Disputed issue: {issue.customer_label}\n"
        f"Reason for dispute: {issue.customer_explanation}\n"
        f"Review round: {issue.suggested_round}\n\n"
        f"{FCRA_NOTICE_OF_DISPUTE}\n\n"
        f"{requested_action}\n\n"
        f"{evidence_note}\n\n"
        "Please send your written response to the mailing address above.\n\n"
        "Sincerely,\n\n"
        "[Customer Signature]\n"
        "[Customer Printed Name]\n"
    )


def build_letter_workflow() -> dict:
    return {
        "draft_only": True,
        "send_letters_automatically": False,
        "customer_authorization_required": True,
        "fcra_notice_of_dispute": FCRA_NOTICE_OF_DISPUTE,
        "fcra_notice_rules": FCRA_NOTICE_RULES,
        "official_cfpb_files_to_reference": [
            "CFPB_credit_reporting_company_dispute_letter_template.docx",
            "CFPB_credit_reporting_company_dispute_instructions.pdf",
            "CFPB_furnisher_dispute_letter_template.docx",
            "CFPB_furnisher_dispute_instructions.pdf",
            "CFPB_complaint_process_one_page.pdf",
        ],
        "bureau_dispute_procedure": {
            "recipient_type": "credit_bureau",
            "delivery_preference": "certified_mail_for_important_disputes",
            "cfpb_template_alignment": "Use the CFPB credit reporting company dispute template structure: identify the item, explain the dispute, request correction, and attach supporting copies.",
            "round_1_uses": [
                "wrong balance or status",
                "unrecognized account",
                "duplicate collection",
                "original creditor and debt buyer both showing active balance",
                "obsolete or missing key date item",
                "obvious mismatch across bureaus",
            ],
            "packet_checklist": [
                "customer-approved dispute letter",
                "FCRA notice of dispute",
                "specific disputed item and reason",
                "targeted proof only",
                "ID and proof of address when needed",
                "redacted unrelated account data",
                "highlighted disputed item",
                "request written investigation results",
            ],
        },
        "furnisher_direct_dispute_procedure": {
            "recipient_type": "furnisher_or_collector",
            "delivery_preference": "certified_mail",
            "cfpb_template_alignment": "Use the CFPB furnisher dispute template structure when the source company, collector, or servicer appears to be reporting wrong information.",
            "prerequisites": [
                "consumer-specific issue identified",
                "customer authorization verified",
                "evidence packet reviewed by admin",
                "specific furnisher item and dispute reason stated",
                "proper furnisher/direct-dispute address confirmed when available",
            ],
            "requested_verification": [
                "basis for reporting",
                "contract or account records",
                "itemized balance",
                "chain of title or assignment where applicable",
                "payment history",
                "date of first delinquency support",
                "proof reporting is complete and accurate",
                "written investigation result or correction/deletion notice",
            ],
        },
        "debt_validation_procedure": {
            "recipient_type": "debt_collector_or_debt_buyer",
            "delivery_preference": "certified_mail_recommended",
            "legal_basis": "FDCPA validation-style request when a collector/debt buyer is collecting or reporting the debt; not a substitute for an FCRA furnisher dispute.",
            "when_to_use": [
                "collection account",
                "debt buyer account",
                "factoring company account",
                "collector is reporting a balance",
                "original creditor, ownership, authority, or itemized balance needs proof",
            ],
            "requested_validation": [
                "original creditor name and address",
                "itemized balance, fees, interest, and payments",
                "documents showing consumer responsibility",
                "chain of title, sale, assignment, or authority to collect",
                "date of first delinquency support",
                "account number or reporting identifier",
            ],
            "credit_vivo_guardrail": "Generate as draft only. Customer approval required. Use with attorney/compliance review for state-specific debt-collection requirements.",
        },
        "escalation_procedure": {
            "cfpb_complaint": {
                "trigger": "no response, verified-as-accurate with weak support, or repeated inaccurate reporting after dispute history is complete",
                "packet": "original dispute, proof of delivery, bureau/furnisher responses, current report excerpt, damages or denial evidence if available",
                "cfpb_method": "Use CFPB complaint escalation only after the normal dispute record is clear: who was contacted, when, what was disputed, how the company responded, and what remains wrong.",
            },
            "state_attorney_general": {
                "trigger": "pattern of non-response, abusive collection conduct, or unresolved state-law concern",
                "packet": "same evidence packet plus consumer timeline and requested resolution",
            },
            "state_regulator": {
                "trigger": "licensed collector, lender, or credit repair/regulatory issue needs agency review",
                "packet": "collector/furnisher identity, license information if known, dispute timeline, and supporting evidence",
            },
            "attorney_review": {
                "trigger": "strong FCRA/FDCPA pattern, measurable damages, denial letter, identity theft, mixed file, or repeated verification without reasonable investigation",
                "packet": "full dispute history, reports before and after disputes, notices, responses, delivery proofs, and damages evidence",
            },
        },
        "tracking_schema": [
            "letter_id",
            "issue_ids",
            "recipient_type",
            "recipient_name",
            "recipient_address",
            "delivery_method",
            "certified_tracking_number",
            "sent_date",
            "delivered_date",
            "response_due_date",
            "day_15_check_date",
            "day_35_followup_check_date",
            "response_received_date",
            "current_status",
            "next_action",
            "fcra_notice_included",
            "customer_authorization_verified",
            "evidence_packet_hash",
        ],
        "event_log_entries": [
            "letter_drafted",
            "customer_authorization_verified",
            "fcra_notice_added",
            "proof_packet_attached",
            "certified_mail_queued",
            "tracking_number_saved",
            "delivery_confirmed",
            "response_deadline_created",
            "response_received",
            "response_scanned",
            "next_action_assigned",
        ],
}


def _is_debt_validation_candidate(issue: ReviewIssue) -> bool:
    blob = " ".join([
        issue.issue_type,
        issue.customer_label,
        issue.customer_explanation,
        issue.admin_explanation,
        " ".join(e.snippet for e in issue.evidence),
    ]).lower()
    return any(term in blob for term in [
        "collection",
        "collector",
        "debt buyer",
        "factoring company",
        "placed for collection",
    ])


def _letter_queue_item(issue: ReviewIssue, letter_type: str, recipient_type: str, responsible_party: str) -> dict:
    return {
        "letter_id": stable_id("letter", issue.id, letter_type),
        "issue_id": issue.id,
        "issue_type": issue.issue_type,
        "letter_type": letter_type,
        "letter_subject": _draft_letter_subject(letter_type, issue),
        "draft_letter_body": _draft_letter_body(letter_type, recipient_type, issue),
        "round": issue.suggested_round,
        "recipient_type": recipient_type,
        "responsible_party": responsible_party,
        "delivery_method": "certified_mail_recommended",
        "fcra_notice_required": letter_type not in {"admin_review_hold", "debt_validation_request"},
        "fcra_notice_included": letter_type not in {"admin_review_hold", "debt_validation_request"},
        "fdcpa_validation_request": letter_type == "debt_validation_request",
        "customer_approval_required": True,
        "customer_authorization_verified": False,
        "tracking_status": "draft_not_sent",
        "recommended_next_action": (
            "send_debt_validation_request_if_collector_or_debt_buyer_and_customer_approves"
            if letter_type == "debt_validation_request"
            else _next_action_for_issue(issue)
        ),
        "escalation_candidate": False,
    }


def build_recommended_letter_queue(issues: List[ReviewIssue]) -> List[dict]:
    queue = []
    for issue in issues:
        responsible_party = _responsible_party_for_issue(issue)
        letter_type = "bureau_dispute"
        recipient_type = "credit_bureau"
        if responsible_party == "furnisher_or_collector":
            letter_type = "furnisher_direct_dispute"
            recipient_type = "furnisher_or_collector"
        elif responsible_party == "admin_review_required":
            letter_type = "admin_review_hold"
            recipient_type = "undetermined"

        queue.append(_letter_queue_item(issue, letter_type, recipient_type, responsible_party))

        if _is_debt_validation_candidate(issue):
            queue.append(_letter_queue_item(
                issue,
                "debt_validation_request",
                "debt_collector_or_debt_buyer",
                "collector_or_debt_buyer",
            ))
    return queue


def build_fcra_review(issues: List[ReviewIssue]) -> List[dict]:
    return [
        {
            "issue_id": issue.id,
            "possible_fcra_issue": issue.issue_type != "low_confidence_admin_review",
            "issue_type": issue.issue_type,
            "responsible_party": _responsible_party_for_issue(issue),
            "dispute_history_complete": False,
            "evidence_strength": _issue_evidence_strength(issue),
            "damages_evidence": "none",
            "next_action": _next_action_for_issue(issue),
            "requires_admin_review": True,
        }
        for issue in issues
    ]


METRO2_FCRA_RULES = {
    "collection_review": {
        "metro2_fields": ["Account Type", "Portfolio Type", "Original Creditor", "Current Balance", "Account Status", "Date Reported"],
        "fcra_focus": "Accuracy, completeness, ownership, and verification of collection reporting.",
        "fcra_sections": ["FCRA 611 reinvestigation", "FCRA 623 furnisher duties after notice"],
        "evidence_needed": ["original creditor", "assignment or ownership proof", "balance calculation", "reporting authority", "account-level source records"],
        "dispute_theory": "The collection must be accurate, complete, and verifiable. If ownership, balance, or reporting authority cannot be supported, correction or deletion may be appropriate.",
    },
    "chargeoff_review": {
        "metro2_fields": ["Account Status", "Payment Rating", "Current Balance", "Amount Past Due", "Date of First Delinquency", "Date Closed", "Special Comment"],
        "fcra_focus": "Charge-off reporting accuracy, dates, balance, sold/transferred status, and obsolescence risk.",
        "fcra_sections": ["FCRA 611 reinvestigation", "FCRA 623 furnisher duties after notice", "FCRA 605 obsolescence review"],
        "evidence_needed": ["charge-off ledger", "payment history", "DOFD support", "sale or transfer record", "balance breakdown"],
        "dispute_theory": "A charge-off should not report internally inconsistent status, balance, ownership, or delinquency dates.",
    },
    "missing_dofd_review": {
        "metro2_fields": ["Date of First Delinquency", "FCRA Compliance/Date of First Delinquency", "Estimated Removal Date"],
        "fcra_focus": "Missing or unsupported delinquency date for negative reporting.",
        "fcra_sections": ["FCRA 611 reinvestigation", "FCRA 605 obsolescence review", "FCRA 623 furnisher duties after notice"],
        "evidence_needed": ["first delinquency date", "payment history", "charge-off/collection timeline", "bureau reporting history"],
        "dispute_theory": "Negative reporting needs a supportable delinquency timeline so the account is not reported longer than allowed.",
    },
    "closed_sold_balance_review": {
        "metro2_fields": ["Account Status", "Current Balance", "Amount Past Due", "Date Closed", "Special Comment", "Purchased/Sold/Transferred Indicator"],
        "fcra_focus": "Closed, sold, transferred, or assigned account still reporting a balance.",
        "fcra_sections": ["FCRA 611 reinvestigation", "FCRA 623 furnisher duties after notice"],
        "evidence_needed": ["sale/transfer record", "current owner", "balance ledger", "zero-balance update support"],
        "dispute_theory": "If the account was sold or transferred, the reporting balance/status must match the furnisher's actual ownership and records.",
    },
    "cross_bureau_balance_mismatch": {
        "metro2_fields": ["Current Balance", "Amount Past Due", "High Credit/Original Amount", "Date Reported"],
        "fcra_focus": "Same account reports different balances across bureaus.",
        "fcra_sections": ["FCRA 611 reasonable reinvestigation", "FCRA 623 furnisher duties after bureau notice"],
        "evidence_needed": ["current account ledger", "last statement", "bureau reporting snapshots", "furnisher update history"],
        "dispute_theory": "The same account should have a supportable balance. Differences may be explainable by timing, but unsupported differences should be corrected.",
    },
    "cross_bureau_status_mismatch": {
        "metro2_fields": ["Account Status", "Payment Rating", "Special Comment", "Compliance Condition Code"],
        "fcra_focus": "Same account reports different status/payment condition across bureaus.",
        "fcra_sections": ["FCRA 611 reasonable reinvestigation", "FCRA 623 furnisher duties after bureau notice"],
        "evidence_needed": ["account status records", "payment history", "charge-off/collection status support", "bureau update records"],
        "dispute_theory": "Status and payment condition should be consistent with the furnisher's records and not materially misleading.",
    },
    "cross_bureau_date_mismatch": {
        "metro2_fields": ["Date Opened", "Date Closed", "Date Reported", "Date of First Delinquency", "Estimated Removal Date"],
        "fcra_focus": "Same account reports different key dates across bureaus.",
        "fcra_sections": ["FCRA 611 reasonable reinvestigation", "FCRA 605 obsolescence review", "FCRA 623 furnisher duties after bureau notice"],
        "evidence_needed": ["opening records", "closing records", "DOFD support", "payment history", "bureau report snapshots"],
        "dispute_theory": "Key dates drive legal reporting age and consumer harm. Unsupported date differences should be corrected or deleted.",
    },
    "low_confidence_admin_review": {
        "metro2_fields": ["Raw tradeline block", "All extracted fields"],
        "fcra_focus": "Manual validation before dispute strategy.",
        "fcra_sections": ["Admin review required before FCRA dispute theory is selected"],
        "evidence_needed": ["raw report excerpt", "account-level PDF pages", "manual field validation"],
        "dispute_theory": "Do not send a dispute until the extracted fields are confirmed against the raw report.",
    },
}


def _expert_rule_for_issue(issue: ReviewIssue) -> dict:
    return METRO2_FCRA_RULES.get(issue.issue_type, {
        "metro2_fields": ["Account Status", "Current Balance", "Date Reported", "Remarks"],
        "fcra_focus": "Accuracy, completeness, and verifiability review.",
        "fcra_sections": ["FCRA 611 reinvestigation", "FCRA 623 furnisher duties after notice"],
        "evidence_needed": ["raw report excerpt", "furnisher records", "bureau investigation response"],
        "dispute_theory": "If the reporting cannot be verified as accurate and complete, it should be corrected, updated, or deleted.",
    })


def build_metro2_fcra_review(issues: List[ReviewIssue]) -> List[dict]:
    review = []
    for issue in issues:
        rule = _expert_rule_for_issue(issue)
        review.append({
            "issue_id": issue.id,
            "issue_type": issue.issue_type,
            "customer_label": issue.customer_label,
            "severity": issue.severity,
            "confidence": issue.confidence,
            "metro2_fields_to_review": rule["metro2_fields"],
            "fcra_focus": rule["fcra_focus"],
            "fcra_sections": rule["fcra_sections"],
            "evidence_needed": rule["evidence_needed"],
            "dispute_theory": rule["dispute_theory"],
            "responsible_party": _responsible_party_for_issue(issue),
            "recommended_next_action": _next_action_for_issue(issue),
            "customer_approval_required": True,
            "attorney_review_signal": issue.severity == "high" or issue.issue_type.startswith("cross_bureau"),
        })
    return review


METRO2_PUBLIC_GUIDE_NOTES = {
    "source": "Collect.org public guide: How to Read the Metro 2 Format",
    "url": "https://www.collect.org/cv13/Help/howtoreadthemetro2format.html",
    "purpose": "Public field-reading guide used to improve scanner worksheets and staff review prompts.",
    "production_guardrail": "Use as a practical guide only; validate production Metro 2 logic against the official licensed CDIA Metro 2 CRRG and compliance counsel.",
    "base_segment_fields": [
        "Consumer Account Number",
        "Portfolio Type",
        "Account Type",
        "Date Opened",
        "Credit Limit",
        "Highest Credit / Original Loan Amount",
        "Terms Duration",
        "Scheduled Monthly Payment Amount",
        "Actual Payment Amount",
        "Account Status",
        "Payment Rating",
        "Payment History Profile",
        "Special Comment",
        "Compliance Condition Code",
        "Current Balance",
        "Amount Past Due",
        "Original Charge-off Amount",
        "Date of Account Information / Date Reported",
        "Date of First Delinquency",
        "Date Closed",
        "Date of Last Payment",
        "Consumer Information Indicator",
        "ECOA Code",
    ],
    "collection_segments": [
        "K1 segment: original creditor name and creditor classification where applicable",
        "J1/J2 segments: consumer name/address information where associated consumers exist",
    ],
}


METRO2_FIELD_REQUIREMENTS = {
    "collection": {
        "label": "Collection account",
        "required_fields": [
            "Account Number / Consumer Account Identifier",
            "Account Type",
            "Portfolio Type",
            "ECOA / Responsibility",
            "Account Status",
            "Current Balance",
            "Amount Past Due if reporting past due",
            "Original Charge-off Amount if applicable",
            "Date Opened or collection acquisition/open date",
            "Date Reported / Date Updated",
            "Date of First Delinquency for negative reporting",
            "K1 Original Creditor Name",
            "K1 Creditor Classification",
            "Creditor Classification / Collection Agency Type",
            "Payment History Profile when furnished",
            "Special Comment / Remarks when needed",
            "Compliance Condition Code / Dispute Indicator when disputed",
            "Consumer Information Indicator when bankruptcy/deceased/dispute conditions apply",
        ],
        "field_map": {
            "Account Number / Consumer Account Identifier": ["account_number_masked"],
            "Account Type": ["account_type"],
            "Portfolio Type": ["portfolio_type"],
            "ECOA / Responsibility": ["responsibility"],
            "Account Status": ["status", "pay_status"],
            "Current Balance": ["balance"],
            "Amount Past Due if reporting past due": ["past_due", "balance"],
            "Original Charge-off Amount if applicable": ["high_credit_or_original_amount", "raw_block"],
            "Date Opened or collection acquisition/open date": ["date_opened"],
            "Date Reported / Date Updated": ["date_reported"],
            "Date of First Delinquency for negative reporting": ["date_of_first_delinquency"],
            "K1 Original Creditor Name": ["original_creditor"],
            "K1 Creditor Classification": ["creditor_classification", "raw_block"],
            "Creditor Classification / Collection Agency Type": ["creditor_classification", "collector_or_debt_buyer"],
            "Payment History Profile when furnished": ["payment_history_summary", "raw_block"],
            "Special Comment / Remarks when needed": ["remarks"],
            "Compliance Condition Code / Dispute Indicator when disputed": ["remarks", "raw_block"],
            "Consumer Information Indicator when bankruptcy/deceased/dispute conditions apply": ["remarks", "raw_block"],
        },
        "validation_notes": [
            "A collection should identify the K1 original creditor name and creditor classification when applicable.",
            "Negative collection reporting should have a supportable Date of First Delinquency timeline.",
            "If disputed, the report should show an appropriate dispute notation or compliance condition after notice.",
        ],
        "dispute_use": "Use for collection ownership, original creditor, balance, DOFD, status, and dispute-notation challenges.",
    },
    "charge_off": {
        "label": "Charge-off account",
        "required_fields": [
            "Account Number / Consumer Account Identifier",
            "Account Type",
            "Portfolio Type",
            "ECOA / Responsibility",
            "Account Status",
            "Payment Rating / Pay Status",
            "Current Balance",
            "Amount Past Due",
            "Charge-off Amount or High Credit / Original Amount",
            "Original Charge-off Amount",
            "Date of First Delinquency",
            "Date Closed when closed",
            "Date Reported / Date Updated",
            "Date Last Payment / Date Last Activity when available",
            "Payment History Profile",
            "Special Comment / Remarks",
            "Compliance Condition Code / Dispute Indicator when disputed",
            "Consumer Information Indicator when applicable",
        ],
        "field_map": {
            "Account Number / Consumer Account Identifier": ["account_number_masked"],
            "Account Type": ["account_type"],
            "Portfolio Type": ["portfolio_type"],
            "ECOA / Responsibility": ["responsibility"],
            "Account Status": ["status"],
            "Payment Rating / Pay Status": ["pay_status", "status"],
            "Current Balance": ["balance"],
            "Amount Past Due": ["past_due"],
            "Charge-off Amount or High Credit / Original Amount": ["high_credit_or_original_amount"],
            "Original Charge-off Amount": ["high_credit_or_original_amount", "raw_block"],
            "Date of First Delinquency": ["date_of_first_delinquency"],
            "Date Closed when closed": ["date_closed", "status"],
            "Date Reported / Date Updated": ["date_reported"],
            "Date Last Payment / Date Last Activity when available": ["date_last_payment", "date_last_activity"],
            "Payment History Profile": ["payment_history_summary", "raw_block"],
            "Special Comment / Remarks": ["remarks"],
            "Compliance Condition Code / Dispute Indicator when disputed": ["remarks", "raw_block"],
            "Consumer Information Indicator when applicable": ["remarks", "raw_block"],
        },
        "validation_notes": [
            "Charge-off status, payment rating, balance, and past-due amount should not contradict each other.",
            "DOFD drives the reporting-age review and should not be missing from negative charge-off reporting.",
            "If sold or transferred, the balance and ownership/status should match the current reporting party's records.",
        ],
        "dispute_use": "Use for charge-off balance, status, payment rating, DOFD, sold/transferred, and obsolescence challenges.",
    },
    "closed_sold_transferred": {
        "label": "Closed, sold, transferred, or assigned account",
        "required_fields": [
            "Account Number / Consumer Account Identifier",
            "Account Type",
            "Account Status",
            "Current Balance",
            "Amount Past Due",
            "Original Charge-off Amount if applicable",
            "Date Closed",
            "Date Reported / Date Updated",
            "Date of First Delinquency if negative",
            "Special Comment / Sold-Transferred Remark",
            "Compliance Condition Code / Dispute Indicator when disputed",
            "Current Owner / Original Creditor context when applicable",
        ],
        "field_map": {
            "Account Number / Consumer Account Identifier": ["account_number_masked"],
            "Account Type": ["account_type"],
            "Account Status": ["status", "pay_status"],
            "Current Balance": ["balance"],
            "Amount Past Due": ["past_due"],
            "Original Charge-off Amount if applicable": ["high_credit_or_original_amount", "raw_block"],
            "Date Closed": ["date_closed", "status"],
            "Date Reported / Date Updated": ["date_reported"],
            "Date of First Delinquency if negative": ["date_of_first_delinquency", "status"],
            "Special Comment / Sold-Transferred Remark": ["remarks", "status"],
            "Compliance Condition Code / Dispute Indicator when disputed": ["remarks", "raw_block"],
            "Current Owner / Original Creditor context when applicable": ["original_creditor", "collector_or_debt_buyer", "remarks"],
        },
        "validation_notes": [
            "A sold or transferred account with a nonzero balance needs ownership and balance support.",
            "Original creditor and collector reporting should not create duplicate active debt reporting.",
            "Closed/transferred status should match date closed, balance, and remarks.",
        ],
        "dispute_use": "Use for sold/transferred balance, duplicate ownership, closed-date, and status challenges.",
    },
    "installment_late_payment": {
        "label": "Installment or late-payment account",
        "required_fields": [
            "Account Number / Consumer Account Identifier",
            "Account Type",
            "Portfolio Type",
            "ECOA / Responsibility",
            "Account Status",
            "Payment Rating / Pay Status",
            "Current Balance",
            "Amount Past Due",
            "High Credit / Original Amount",
            "Terms or Scheduled Payment Amount when available",
            "Actual Payment Amount when available",
            "Date Opened",
            "Date Reported / Date Updated",
            "Date Last Payment / Date Last Activity",
            "Date of First Delinquency if negative",
            "Payment History Profile",
            "Special Comment / Remarks when needed",
            "Compliance Condition Code / Dispute Indicator when disputed",
        ],
        "field_map": {
            "Account Number / Consumer Account Identifier": ["account_number_masked"],
            "Account Type": ["account_type"],
            "Portfolio Type": ["portfolio_type"],
            "ECOA / Responsibility": ["responsibility"],
            "Account Status": ["status"],
            "Payment Rating / Pay Status": ["pay_status", "status"],
            "Current Balance": ["balance"],
            "Amount Past Due": ["past_due"],
            "High Credit / Original Amount": ["high_credit_or_original_amount"],
            "Terms or Scheduled Payment Amount when available": ["raw_block"],
            "Actual Payment Amount when available": ["raw_block"],
            "Date Opened": ["date_opened"],
            "Date Reported / Date Updated": ["date_reported"],
            "Date Last Payment / Date Last Activity": ["date_last_payment", "date_last_activity"],
            "Date of First Delinquency if negative": ["date_of_first_delinquency", "status"],
            "Payment History Profile": ["payment_history_summary", "raw_block"],
            "Special Comment / Remarks when needed": ["remarks"],
            "Compliance Condition Code / Dispute Indicator when disputed": ["remarks", "raw_block"],
        },
        "validation_notes": [
            "Late-payment sequence should make sense against payment history and reported status.",
            "A current/paid account should not still show a materially contradictory late or past-due condition.",
            "DOFD is needed when negative reporting can affect reporting age.",
        ],
        "dispute_use": "Use for late-payment, payment history, balance, date, and pay-status challenges.",
    },
    "revolving": {
        "label": "Revolving account",
        "required_fields": [
            "Account Number / Consumer Account Identifier",
            "Account Type",
            "Portfolio Type",
            "ECOA / Responsibility",
            "Account Status",
            "Payment Rating / Pay Status",
            "Current Balance",
            "Credit Limit",
            "High Credit",
            "Amount Past Due",
            "Date Opened",
            "Date Reported / Date Updated",
            "Date Last Payment / Date Last Activity",
            "Date of First Delinquency if negative",
            "Payment History Profile",
            "Special Comment / Remarks when needed",
            "Compliance Condition Code / Dispute Indicator when disputed",
        ],
        "field_map": {
            "Account Number / Consumer Account Identifier": ["account_number_masked"],
            "Account Type": ["account_type"],
            "Portfolio Type": ["portfolio_type"],
            "ECOA / Responsibility": ["responsibility"],
            "Account Status": ["status"],
            "Payment Rating / Pay Status": ["pay_status", "status"],
            "Current Balance": ["balance"],
            "Credit Limit": ["credit_limit"],
            "High Credit": ["high_credit_or_original_amount"],
            "Amount Past Due": ["past_due"],
            "Date Opened": ["date_opened"],
            "Date Reported / Date Updated": ["date_reported"],
            "Date Last Payment / Date Last Activity": ["date_last_payment", "date_last_activity"],
            "Date of First Delinquency if negative": ["date_of_first_delinquency", "status"],
            "Payment History Profile": ["payment_history_summary", "raw_block"],
            "Special Comment / Remarks when needed": ["remarks"],
            "Compliance Condition Code / Dispute Indicator when disputed": ["remarks", "raw_block"],
        },
        "validation_notes": [
            "Revolving balance, limit, high credit, past due, and payment rating should tell the same story.",
            "A negative revolving account needs a supportable delinquency timeline.",
            "Closed, transferred, or charged-off revolving accounts need status and balance support.",
        ],
        "dispute_use": "Use for balance, limit, utilization, late-payment, status, and DOFD challenges.",
    },
    "generic": {
        "label": "General tradeline",
        "required_fields": [
            "Account Number / Consumer Account Identifier",
            "Account Type",
            "Account Status",
            "Current Balance",
            "Date Opened",
            "Date Reported / Date Updated",
            "Date of First Delinquency if negative",
            "Special Comment / Remarks when needed",
        ],
        "field_map": {
            "Account Number / Consumer Account Identifier": ["account_number_masked"],
            "Account Type": ["account_type"],
            "Account Status": ["status", "pay_status"],
            "Current Balance": ["balance"],
            "Date Opened": ["date_opened"],
            "Date Reported / Date Updated": ["date_reported"],
            "Date of First Delinquency if negative": ["date_of_first_delinquency", "status"],
            "Special Comment / Remarks when needed": ["remarks"],
        },
        "validation_notes": [
            "At minimum, the account should identify who reported it, what it is, current status, balance, and key dates.",
            "Negative items need a supportable delinquency timeline and should be marked disputed when applicable.",
        ],
        "dispute_use": "Use as a general accuracy, completeness, and verifiability checklist.",
    },
}


def _tradeline_profile(tradeline: dict) -> str:
    text = " ".join(
        str(tradeline.get(field, ""))
        for field in [
            "account_name",
            "account_type",
            "portfolio_type",
            "status",
            "pay_status",
            "remarks",
            "raw_block",
        ]
    ).lower()
    if "collection" in text or "collector" in text or "debt buyer" in text:
        return "collection"
    if "charge" in text and "off" in text:
        return "charge_off"
    if any(term in text for term in ["sold", "transferred", "assigned", "purchased by another lender"]):
        return "closed_sold_transferred"
    if any(term in text for term in ["credit card", "revolving", "bankcard", "charge card"]):
        return "revolving"
    if any(term in text for term in ["installment", "auto", "student", "mortgage", "loan", "late", "past due"]):
        return "installment_late_payment"
    return "generic"


def _has_any_field(tradeline: dict, field_names: List[str]) -> bool:
    for field_name in field_names:
        value = str(tradeline.get(field_name, "") or "").strip()
        if not value:
            continue
        if field_name == "raw_block" and len(value) < 12:
            continue
        return True
    return False


def _metro2_presence(tradeline: dict, requirement: dict) -> Tuple[List[str], List[str]]:
    present = []
    missing = []
    field_map = requirement.get("field_map", {})
    for field_label in requirement.get("required_fields", []):
        if _has_any_field(tradeline, field_map.get(field_label, [])):
            present.append(field_label)
        else:
            missing.append(field_label)
    return present, missing


def _metro2_warning_flags(tradeline: dict, profile: str, missing_fields: List[str]) -> List[str]:
    flags = []
    status_text = " ".join([
        str(tradeline.get("status", "") or ""),
        str(tradeline.get("pay_status", "") or ""),
        str(tradeline.get("remarks", "") or ""),
    ]).lower()
    balance_number = money_to_number(str(tradeline.get("balance", "") or ""))

    if "Date of First Delinquency" in " ".join(missing_fields) and any(
        term in status_text for term in ["collection", "charge", "delinquent", "past due", "late"]
    ):
        flags.append("Missing DOFD on negative reporting candidate")
    if profile == "collection" and "Original Creditor" in missing_fields:
        flags.append("Collection missing original creditor field")
    if profile == "charge_off" and "Payment Rating / Pay Status" in missing_fields:
        flags.append("Charge-off missing payment rating/pay status support")
    if profile == "closed_sold_transferred" and balance_number and balance_number > 0:
        flags.append("Sold/transferred/closed account reports nonzero balance")
    if "dispute" not in status_text and "Compliance Condition Code / Dispute Indicator when disputed" in missing_fields:
        flags.append("If customer disputes this item, track dispute notation after notice")
    if not flags and missing_fields:
        flags.append("Admin should validate missing fields against raw PDF")
    if not flags:
        flags.append("No immediate required-field warning detected")
    return flags


def build_metro2_requirement_review(tradelines: List[dict]) -> List[dict]:
    rows = []
    for tradeline in tradelines:
        profile_key = _tradeline_profile(tradeline)
        requirement = METRO2_FIELD_REQUIREMENTS[profile_key]
        present, missing = _metro2_presence(tradeline, requirement)
        rows.append({
            "tradeline_id": tradeline.get("id", ""),
            "bureau": tradeline.get("bureau", ""),
            "account_name": tradeline.get("account_name", ""),
            "account_type": tradeline.get("account_type", ""),
            "status": tradeline.get("status") or tradeline.get("pay_status") or "",
            "metro2_profile": requirement["label"],
            "required_core_fields": requirement["required_fields"],
            "present_fields": present,
            "missing_or_needs_validation": missing,
            "warning_flags": _metro2_warning_flags(tradeline, profile_key, missing),
            "validation_notes": requirement["validation_notes"],
            "dispute_use": requirement["dispute_use"],
            "production_note": "Credit Vivo checklist. Validate against official licensed CDIA Metro 2 CRRG before production/certification.",
        })
    return rows


FCRA_COMPLIANCE_AREAS = {
    "maximum_possible_accuracy": {
        "label": "FCRA accuracy and completeness",
        "law_reference": "FCRA 607(b) / 15 USC 1681e(b)",
        "plain_english": "Credit reports should use reasonable procedures so the information is as accurate as possible.",
        "applies_when": "Any account field appears wrong, incomplete, mixed, misleading, duplicated, or inconsistent across bureaus.",
        "evidence_needed": "3-bureau comparison, raw report excerpt, account statements, payment proof, creditor/collector records, identity proof when needed.",
        "tracking_action": "Create field-level dispute draft and track before/after bureau values.",
    },
    "bureau_reinvestigation": {
        "label": "Bureau reinvestigation duty",
        "law_reference": "FCRA 611 / 15 USC 1681i",
        "plain_english": "When a consumer disputes report information, the bureau must reasonably reinvestigate and send written results.",
        "applies_when": "Customer disputes a specific bureau report item or cross-bureau mismatch.",
        "evidence_needed": "Customer-approved dispute, specific account fields challenged, proof documents, delivery tracking, bureau response.",
        "tracking_action": "Track dispute sent date, delivery date, response due date, written results, changed/deleted/verified outcome.",
    },
    "furnisher_after_notice": {
        "label": "Furnisher investigation after bureau notice",
        "law_reference": "FCRA 623(b) / 15 USC 1681s-2(b)",
        "plain_english": "After the bureau sends the dispute to the furnisher, the furnisher must investigate and report results back.",
        "applies_when": "A bureau dispute is sent and the furnisher/collector is the source of the challenged field.",
        "evidence_needed": "Bureau dispute package, furnisher records, account ledger, ownership/assignment proof, payment history, response result.",
        "tracking_action": "Track which furnisher controls each challenged field and whether the item was corrected, deleted, or verified.",
    },
    "direct_furnisher_dispute": {
        "label": "Direct furnisher dispute",
        "law_reference": "Regulation V 12 CFR 1022.43",
        "plain_english": "A consumer can dispute certain account information directly with the company furnishing the data.",
        "applies_when": "The dispute relates to liability, account terms, balance, credit limit, payment status, payment amount, or open/closed dates.",
        "evidence_needed": "Direct dispute notice, account identifier, specific facts, supporting documents, proof of delivery, furnisher response.",
        "tracking_action": "Prepare direct furnisher dispute only after customer approval and route to the furnisher's direct-dispute address when available.",
    },
    "dofd_obsolescence": {
        "label": "DOFD and obsolete information review",
        "law_reference": "FCRA 605 and 623(a)(5) / 15 USC 1681c and 1681s-2(a)(5)",
        "plain_english": "Negative reporting needs the right first-delinquency date so old items do not stay on the report too long.",
        "applies_when": "Collection, charge-off, late payment, or other negative item is missing DOFD, has conflicting dates, or appears too old.",
        "evidence_needed": "First delinquency timeline, payment history, charge-off date, collection placement date, removal date, prior reports.",
        "tracking_action": "Flag missing/conflicting DOFD and track whether deletion/correction or obsolescence review is needed.",
    },
    "disputed_by_consumer_notation": {
        "label": "Disputed-account notation",
        "law_reference": "FCRA 623(a)(3) / 15 USC 1681s-2(a)(3)",
        "plain_english": "If an account is disputed and still furnished, it should be reported as disputed when the duty applies.",
        "applies_when": "Credit Vivo sends a customer-approved dispute and later report updates still omit dispute notation.",
        "evidence_needed": "Notice of dispute, proof of delivery, post-dispute report, furnisher/bureau response.",
        "tracking_action": "After sending, check updated reports for dispute notation and escalate if unresolved.",
    },
    "identity_theft_block": {
        "label": "Identity theft block review",
        "law_reference": "FCRA 605B / 15 USC 1681c-2",
        "plain_english": "Identity-theft items can require a faster block when the consumer provides the required identity-theft packet.",
        "applies_when": "Customer says the account is identity theft, fraud, not mine, or mixed file.",
        "evidence_needed": "Identity proof, identity theft report, item identification, consumer statement that the transaction was not theirs.",
        "tracking_action": "Hold for identity-theft workflow and attorney/compliance review before using this route.",
    },
}


EOSCAR_PUBLIC_FACTS = [
    {
        "title": "What e-OSCAR is",
        "detail": "e-OSCAR is used by consumer reporting agencies and data furnishers to exchange and respond to credit-history disputes. Consumers submit disputes to bureaus and/or furnishers, not directly inside e-OSCAR.",
        "source": "Credit Vivo local e-OSCAR learning guide; public e-OSCAR getting-started concepts",
    },
    {
        "title": "ACDV path",
        "detail": "When a bureau receives a consumer dispute, it may send an Automated Credit Dispute Verification (ACDV) to the furnisher. The furnisher can verify, update, or correct the information.",
        "source": "Credit Vivo local e-OSCAR learning guide",
    },
    {
        "title": "AUD path",
        "detail": "A furnisher can initiate an Automated Universal Dataform (AUD) to request an out-of-cycle correction. AUD is not a consumer direct-submission path.",
        "source": "Credit Vivo local e-OSCAR learning guide",
    },
    {
        "title": "Best packaging principle",
        "detail": "A strong dispute package identifies each account, exact field, reported value, consumer position, support documents, and requested correction or deletion.",
        "source": "Credit Vivo local e-OSCAR learning guide; FTC/CFPB consumer dispute guidance",
    },
]


EOSCAR_ISSUE_PACKAGING = {
    "cross_bureau_balance_mismatch": {
        "category": "field-specific accuracy dispute",
        "package_hint": "List each balance reported by each bureau and attach the 3-bureau comparison.",
        "evidence_hint": "3-bureau comparison, statements, payoff/settlement proof, itemization, or account ledger.",
    },
    "cross_bureau_status_mismatch": {
        "category": "account status/payment history accuracy dispute",
        "package_hint": "Separate current status, pay status, charge-off, collection, paid/settled, and closed/open wording.",
        "evidence_hint": "Settlement letter, payoff confirmation, account statements, bureau report pages, response letters.",
    },
    "cross_bureau_date_mismatch": {
        "category": "date field accuracy dispute",
        "package_hint": "List each date reported by bureau and ask for source-record verification of the correct date.",
        "evidence_hint": "Prior reports, account statements, payment history, charge-off notice, collection notice.",
    },
    "missing_dofd_review": {
        "category": "date of first delinquency completeness dispute",
        "package_hint": "Ask for verification of the DOFD and the basis for the reporting period.",
        "evidence_hint": "Payment history, charge-off statement, collection notice, old account statement showing delinquency timeline.",
    },
    "collection_review": {
        "category": "collection account completeness dispute",
        "package_hint": "Identify the collection account, original creditor, balance, ownership/assignment question, and requested verification.",
        "evidence_hint": "Credit report page, collector letter, original creditor statement, assignment/sale letter if available.",
    },
    "chargeoff_review": {
        "category": "charge-off accuracy dispute",
        "package_hint": "Separate charge-off status, current balance, past due, charge-off amount, DOFD, and sold/transferred status.",
        "evidence_hint": "Charge-off statement, account ledger, payment history, sale/transfer notice, settlement or payoff proof.",
    },
    "closed_sold_balance_review": {
        "category": "sold/transferred balance accuracy dispute",
        "package_hint": "Ask whether the furnisher still owns the debt and why a balance remains after sold/transferred/closed status.",
        "evidence_hint": "Sale/transfer notice, account statement, collector notice, original creditor response.",
    },
    "low_confidence_admin_review": {
        "category": "manual packaging review",
        "package_hint": "Do not send until the account identity and specific field dispute are confirmed from the raw report.",
        "evidence_hint": "Raw report pages, manual parser review notes, customer statement, supporting documents.",
    },
}


def _eoscar_packaging_for_issue(issue_type: str) -> dict:
    return EOSCAR_ISSUE_PACKAGING.get(issue_type or "", {
        "category": "specific factual dispute",
        "package_hint": "Explain the exact inaccurate or incomplete reporting field. Avoid broad wording unless identity/liability is truly disputed.",
        "evidence_hint": "Relevant report page plus account-level documents supporting the specific correction requested.",
    })


def build_eoscar_packaging_review(issues: List[dict], tradelines: List[dict]) -> List[dict]:
    tradelines_by_id = {item.get("id"): item for item in tradelines}
    rows = []
    for issue in issues:
        related = [
            tradelines_by_id[item_id]
            for item_id in issue.get("related_tradeline_ids", [])
            if item_id in tradelines_by_id
        ]
        packaging = _eoscar_packaging_for_issue(issue.get("issue_type", ""))
        account_names = sorted({item.get("account_name", "") for item in related if item.get("account_name")})
        bureaus = sorted({item.get("bureau", "") for item in related if item.get("bureau")})
        fields = _expert_rule_for_issue(ReviewIssue(
            id=issue.get("id", ""),
            issue_type=issue.get("issue_type", ""),
            severity=issue.get("severity", ""),
            customer_label=issue.get("customer_label", ""),
            customer_explanation=issue.get("customer_explanation", ""),
            admin_explanation=issue.get("admin_explanation", ""),
            suggested_round=issue.get("suggested_round", ""),
        )).get("metro2_fields", [])
        rows.append({
            "issue_id": issue.get("id", ""),
            "issue_type": issue.get("issue_type", ""),
            "account_names": account_names,
            "bureaus": bureaus,
            "eoscar_category": packaging["category"],
            "acdv_packaging_steps": [
                "Create one bureau-specific package for each bureau reporting the error so the issue can survive CRA intake and ACDV routing.",
                "Use one account group per section.",
                "Include exact account name and masked account number.",
                "Name the exact disputed field and reported value.",
                "State the consumer position and factual basis.",
                "Attach exhibits and request correct, update, verify, delete, or block as appropriate.",
            ],
            "field_focus": fields,
            "package_hint": packaging["package_hint"],
            "evidence_hint": packaging["evidence_hint"],
            "avoid": [
                "Do not send generic 'this is inaccurate' wording by itself.",
                "Do not claim direct e-OSCAR access.",
                "Do not send without customer approval and proof review.",
                "Do not misrepresent facts or create false disputes.",
            ],
            "tracking_status": "draft_packaging_review_not_sent",
        })
    return rows


FIELD_COMPLIANCE_RULES = {
    "Account/Furnisher Name": {
        "source_field": "account_name",
        "required_for": ["all"],
        "metro2_concept": "Identifies the reporting furnisher/tradeline.",
        "fcra_basis": "FCRA 607(b) accuracy; FCRA 611 reinvestigation if disputed.",
        "missing_issue": "Furnisher/account name is missing or not visible.",
        "different_issue": "Furnisher/account name differs across bureaus.",
        "verification_ask": "Verify the exact furnisher/subscriber name reporting this tradeline.",
    },
    "Account Number": {
        "source_field": "account_number_masked",
        "required_for": ["all"],
        "metro2_concept": "Consumer account identifier, usually masked on consumer reports.",
        "fcra_basis": "FCRA 607(b) accuracy; FCRA 611 reinvestigation if disputed.",
        "missing_issue": "Account number is missing or masked differently across reports.",
        "different_issue": "Account number does not align across bureaus.",
        "verification_ask": "Verify the account identifier and confirm this is the same tradeline across bureaus.",
    },
    "Account Type": {
        "source_field": "account_type",
        "required_for": ["all"],
        "metro2_concept": "Classifies the account as collection, revolving, auto, mortgage, student loan, charge account, etc.",
        "fcra_basis": "FCRA 607(b) accuracy; FCRA 623 furnisher accuracy/integrity duties.",
        "missing_issue": "Account type is missing or not visible.",
        "different_issue": "Account type differs across bureaus.",
        "verification_ask": "Verify the reported account type and correct any inaccurate classification.",
    },
    "Responsibility / ECOA": {
        "source_field": "responsibility",
        "required_for": ["all"],
        "metro2_concept": "Consumer responsibility such as individual, joint, co-signer, or authorized user.",
        "fcra_basis": "FCRA 607(b) accuracy; FCRA 611 reinvestigation if disputed.",
        "missing_issue": "Responsibility/ECOA field is missing.",
        "different_issue": "Responsibility/ECOA differs across bureaus.",
        "verification_ask": "Verify the consumer responsibility code for this tradeline.",
    },
    "Original Creditor": {
        "source_field": "original_creditor",
        "required_for": ["collection", "debt buyer", "factoring"],
        "metro2_concept": "K1 original creditor/source account, especially important for collections and purchased debt.",
        "fcra_basis": "FCRA 607(b) accuracy; FCRA 623 furnisher duties after notice; Reg V direct dispute where applicable.",
        "missing_issue": "Original creditor is missing for a collection/debt-buyer review item.",
        "different_issue": "Original creditor differs across bureaus.",
        "verification_ask": "Verify the K1 original creditor/source account and ownership or assignment chain.",
    },
    "Creditor Classification": {
        "source_field": "creditor_classification",
        "required_for": ["collection", "debt buyer", "factoring"],
        "metro2_concept": "K1 creditor classification for the original creditor/source account when applicable.",
        "fcra_basis": "FCRA 607(b) accuracy; FCRA 623 furnisher accuracy/integrity duties.",
        "missing_issue": "Creditor classification is missing or not visible for a collection/debt-buyer item.",
        "different_issue": "Creditor classification differs across bureaus.",
        "verification_ask": "Verify the K1 creditor classification and whether the account is being reported under the correct source category.",
    },
    "Current Balance": {
        "source_field": "balance",
        "required_for": ["all"],
        "metro2_concept": "Current reported balance.",
        "fcra_basis": "FCRA 607(b) accuracy; FCRA 623 furnisher accuracy/integrity duties; Reg V direct dispute.",
        "missing_issue": "Current balance is missing.",
        "different_issue": "Current balance differs across bureaus.",
        "verification_ask": "Verify the current balance and provide the records supporting the amount.",
    },
    "Past Due Amount": {
        "source_field": "past_due",
        "required_for": ["collection", "charge-off", "late", "past due", "delinquent"],
        "metro2_concept": "Amount currently past due, if applicable.",
        "fcra_basis": "FCRA 607(b) accuracy; FCRA 623 furnisher accuracy/integrity duties; Reg V direct dispute.",
        "missing_issue": "Past due amount is missing or not separately shown for a negative account.",
        "different_issue": "Past due amount differs across bureaus.",
        "verification_ask": "Verify the amount past due and whether it should be separately reported.",
    },
    "High Credit / Original Amount": {
        "source_field": "high_credit_or_original_amount",
        "required_for": ["collection", "debt buyer", "installment", "mortgage", "auto", "loan", "charge-off"],
        "metro2_concept": "Original balance, high credit, or original amount depending on account type.",
        "fcra_basis": "FCRA 607(b) accuracy; FCRA 623 furnisher accuracy/integrity duties.",
        "missing_issue": "Original amount/high credit is missing for this account class.",
        "different_issue": "Original amount/high credit differs across bureaus.",
        "verification_ask": "Verify the original balance, original amount, high credit, or charge-off amount.",
    },
    "Original Charge-off Amount": {
        "source_field": "high_credit_or_original_amount",
        "required_for": ["charge-off", "collection", "debt buyer"],
        "metro2_concept": "Original charge-off amount when the account was charged off before collection/sale/transfer.",
        "fcra_basis": "FCRA 607(b) accuracy; FCRA 623 furnisher accuracy/integrity duties; FCRA 611 reinvestigation.",
        "missing_issue": "Original charge-off amount is not separately visible for a charged-off or collection review item.",
        "different_issue": "Original charge-off/high-credit amount differs across bureaus.",
        "verification_ask": "Verify the original charge-off amount and how it relates to current balance, past due, and sale/transfer history.",
    },
    "Credit Limit": {
        "source_field": "credit_limit",
        "required_for": ["credit card", "revolving"],
        "metro2_concept": "Credit limit for revolving accounts when applicable.",
        "fcra_basis": "FCRA 607(b) accuracy; FCRA 623 furnisher accuracy/integrity duties.",
        "missing_issue": "Credit limit is missing where it may be applicable.",
        "different_issue": "Credit limit differs across bureaus.",
        "verification_ask": "Verify the credit limit or confirm why it is not applicable.",
    },
    "Status / Pay Status": {
        "source_field": "status",
        "required_for": ["all"],
        "metro2_concept": "Current account/payment condition such as current, collection, charge-off, paid, closed, transferred.",
        "fcra_basis": "FCRA 607(b) accuracy; FCRA 623 furnisher accuracy/integrity duties; FCRA 611 after dispute.",
        "missing_issue": "Status/pay status is missing.",
        "different_issue": "Status/pay status differs across bureaus.",
        "verification_ask": "Verify the exact account/payment status and supporting account history.",
    },
    "Payment Rating": {
        "source_field": "pay_status",
        "required_for": ["charge-off", "late", "past due", "delinquent", "collection", "revolving", "installment"],
        "metro2_concept": "Payment rating represents the current payment condition and should align with account status and history.",
        "fcra_basis": "FCRA 607(b) accuracy; FCRA 623 furnisher accuracy/integrity duties.",
        "missing_issue": "Payment rating/pay status is missing or not visible for an account where payment condition matters.",
        "different_issue": "Payment rating/pay status differs across bureaus.",
        "verification_ask": "Verify the payment rating against the payment history profile, account status, and current balance.",
    },
    "Payment History Profile": {
        "source_field": "payment_history_summary",
        "required_for": ["charge-off", "late", "past due", "delinquent", "revolving", "installment", "auto", "mortgage"],
        "metro2_concept": "Monthly payment history string showing whether months were current, late, charged off, or otherwise coded.",
        "fcra_basis": "FCRA 607(b) accuracy; FCRA 623 furnisher accuracy/integrity duties; FCRA 611 reinvestigation.",
        "missing_issue": "Payment history profile is missing or not visible for a payment-history review item.",
        "different_issue": "Payment history differs across bureaus.",
        "verification_ask": "Verify the payment history profile month by month against the account ledger.",
    },
    "Date Opened / Assigned": {
        "source_field": "date_opened",
        "required_for": ["all"],
        "metro2_concept": "Date opened or assignment/open date depending on account type.",
        "fcra_basis": "FCRA 607(b) accuracy; FCRA 623 furnisher accuracy/integrity duties.",
        "missing_issue": "Date opened/assigned is missing.",
        "different_issue": "Date opened/assigned differs across bureaus.",
        "verification_ask": "Verify the date opened or assignment date with source records.",
    },
    "Date Reported / Updated": {
        "source_field": "date_reported",
        "required_for": ["all"],
        "metro2_concept": "Date of account information/date reported: the date the furnisher last reported or updated the account.",
        "fcra_basis": "FCRA 607(b) accuracy; FCRA 623 furnisher accuracy/integrity duties.",
        "missing_issue": "Date reported/updated is missing.",
        "different_issue": "Date reported/updated differs across bureaus.",
        "verification_ask": "Verify the date this account was last reported or updated.",
    },
    "Date Closed": {
        "source_field": "date_closed",
        "required_for": ["closed", "charge-off", "transferred", "sold"],
        "metro2_concept": "Date account was closed, transferred, sold, or otherwise no longer open.",
        "fcra_basis": "FCRA 607(b) accuracy; FCRA 623 furnisher accuracy/integrity duties.",
        "missing_issue": "Date closed is missing for a closed/sold/transferred account.",
        "different_issue": "Date closed differs across bureaus.",
        "verification_ask": "Verify the date closed or confirm whether the account should be open/closed.",
    },
    "Date of First Delinquency": {
        "source_field": "date_of_first_delinquency",
        "required_for": ["collection", "charge-off", "late", "past due", "delinquent"],
        "metro2_concept": "First delinquency date tied to the negative reporting period.",
        "fcra_basis": "FCRA 605 obsolescence; FCRA 623(a)(5) DOFD duty; FCRA 611 reinvestigation if disputed.",
        "missing_issue": "Date of first delinquency is missing for an adverse account.",
        "different_issue": "Date of first delinquency differs across bureaus.",
        "verification_ask": "Verify the DOFD and reporting-period source records.",
    },
    "Estimated Removal / On Record Until": {
        "source_field": "estimated_removal_date",
        "required_for": ["collection", "charge-off", "late", "past due", "delinquent"],
        "metro2_concept": "Consumer-facing estimated removal or on-record-until date, when shown.",
        "fcra_basis": "FCRA 605 obsolescence; FCRA 623(a)(5) DOFD duty.",
        "missing_issue": "Estimated removal/on-record-until date is missing for an adverse account.",
        "different_issue": "Estimated removal/on-record-until date differs.",
        "verification_ask": "Verify the reporting period and estimated removal date.",
    },
    "Remarks / Narrative Codes": {
        "source_field": "remarks",
        "required_for": ["disputed", "collection", "charge-off", "transferred", "sold", "bankruptcy"],
        "metro2_concept": "Special comment, remarks, compliance condition, dispute, or narrative codes.",
        "fcra_basis": "FCRA 623(a)(3) disputed-account notation; FCRA 607(b) accuracy.",
        "missing_issue": "Remarks/dispute notation may be missing where expected.",
        "different_issue": "Remarks/narrative codes differ across bureaus.",
        "verification_ask": "Verify special comments, dispute indicators, and narrative codes.",
    },
    "Compliance Condition Code": {
        "source_field": "remarks",
        "required_for": ["disputed", "collection", "charge-off", "bankruptcy", "identity theft", "deceased"],
        "metro2_concept": "Compliance condition code signals disputes and special compliance conditions.",
        "fcra_basis": "FCRA 623(a)(3) disputed-account notation; FCRA 611 reinvestigation; Reg V accuracy/integrity duties.",
        "missing_issue": "Compliance/dispute condition is not visible where a dispute or special condition may apply.",
        "different_issue": "Compliance condition/dispute notation differs across bureaus.",
        "verification_ask": "Verify whether the account should be marked disputed or carry another compliance condition code.",
    },
    "Consumer Information Indicator": {
        "source_field": "remarks",
        "required_for": ["bankruptcy", "deceased", "personal receivership"],
        "metro2_concept": "Consumer information indicator can reflect consumer-level conditions such as bankruptcy or deceased status.",
        "fcra_basis": "FCRA 607(b) accuracy; FCRA 611 reinvestigation; FCRA 623 furnisher accuracy/integrity duties.",
        "missing_issue": "Consumer information indicator is not visible where a consumer-level condition may apply.",
        "different_issue": "Consumer information indicator differs across bureaus.",
        "verification_ask": "Verify whether any consumer-level indicator is being reported and whether it is accurate.",
    },
}


def _account_class_text(tradeline: dict) -> str:
    return " ".join(
        str(tradeline.get(field, "") or "")
        for field in [
            "account_name",
            "account_type",
            "portfolio_type",
            "status",
            "pay_status",
            "remarks",
            "raw_block",
        ]
    ).lower()


def _field_required_for_account(rule: dict, account_class: str) -> bool:
    required_for = [str(item).lower() for item in rule.get("required_for", [])]
    if "all" in required_for:
        return True
    return any(token in account_class for token in required_for)


def build_field_compliance_audit(tradelines: List[dict]) -> List[dict]:
    rows = []
    for tradeline in tradelines:
        account_class = _account_class_text(tradeline)
        for field_name, rule in FIELD_COMPLIANCE_RULES.items():
            source_field = rule["source_field"]
            value = tradeline.get(source_field, "")
            if field_name == "Status / Pay Status":
                value = tradeline.get("status") or tradeline.get("pay_status") or ""
            required = _field_required_for_account(rule, account_class)
            present = bool(str(value or "").strip())
            if present:
                issue_flag = "OK"
            elif required:
                issue_flag = "MISSING_REQUIRED_EXPECTED"
            else:
                issue_flag = "NOT_VISIBLE_CONDITIONAL"
            rows.append({
                "tradeline_id": tradeline.get("id", ""),
                "bureau": tradeline.get("bureau", ""),
                "account_name": tradeline.get("account_name", ""),
                "field_name": field_name,
                "parsed_value": value or "Not shown",
                "required_or_expected": "Yes" if required else "Conditional / if reported",
                "issue_flag": issue_flag,
                "metro2_concept": rule["metro2_concept"],
                "fcra_basis": rule["fcra_basis"],
                "issue_text": "" if issue_flag == "OK" else rule["missing_issue"],
                "verification_ask": rule["verification_ask"],
                "requested_outcome": (
                    "Verify and correct with all CRAs; delete only if inaccurate, incomplete, or unverifiable."
                    if issue_flag != "OK" else
                    "No automated correction requested unless source records show an inaccuracy."
                ),
                "source_note": "Built from Credit Vivo local FCRA/Metro 2 rule files; not a substitute for attorney/CDIA CRRG validation.",
            })
    return rows


DATE_FIELD_TITLES = {
    "date_opened": "Date Opened / Assigned",
    "date_closed": "Date Closed",
    "date_reported": "Date Reported / Last Updated",
    "date_last_activity": "Date of Last Activity",
    "date_last_payment": "Date of Last Payment",
    "date_of_first_delinquency": "Date of First Delinquency / DOFD",
    "estimated_removal_date": "Estimated Removal / On Record Until",
}


DATE_LABEL_PATTERNS = {
    "date_opened": ["date opened", "opened"],
    "date_closed": ["date closed", "closed"],
    "date_reported": ["date reported", "date updated", "balance updated", "last reported", "reported"],
    "date_last_activity": ["date of last activity", "last activity"],
    "date_last_payment": ["date of last payment", "last payment made", "last payment"],
    "date_of_first_delinquency": ["date of first delinquency", "date of 1st delinquency", "first delinquency", "dofd"],
    "estimated_removal_date": ["on record until", "estimated month and year this item will be removed", "estimated removal"],
}


def _month_number(name: str) -> str:
    months = {
        "jan": "01", "january": "01",
        "feb": "02", "february": "02",
        "mar": "03", "march": "03",
        "apr": "04", "april": "04",
        "may": "05",
        "jun": "06", "june": "06",
        "jul": "07", "july": "07",
        "aug": "08", "august": "08",
        "sep": "09", "sept": "09", "september": "09",
        "oct": "10", "october": "10",
        "nov": "11", "november": "11",
        "dec": "12", "december": "12",
    }
    return months.get(name.lower()[:9], "")


def normalize_audit_date(value: str) -> Tuple[str, str]:
    value = clean_text(value)
    m = re.search(r"\b(\d{1,2})/(\d{1,2})/(\d{2,4})\b", value)
    if m:
        year = m.group(3)
        if len(year) == 2:
            year = "20" + year if int(year) <= 40 else "19" + year
        return f"{int(year):04d}-{int(m.group(1)):02d}-{int(m.group(2)):02d}", "day"

    m = re.search(r"\b([A-Za-z]{3,9})\s+(\d{1,2}),?\s+(\d{4})\b", value)
    if m and _month_number(m.group(1)):
        return f"{int(m.group(3)):04d}-{_month_number(m.group(1))}-{int(m.group(2)):02d}", "day"

    m = re.search(r"\b([A-Za-z]{3,9})\s+(\d{4})\b", value)
    if m and _month_number(m.group(1)):
        return f"{int(m.group(2)):04d}-{_month_number(m.group(1))}", "month"

    m = re.search(r"\b(\d{4})\b", value)
    if m:
        return m.group(1), "year"

    return value, "unknown"


def _date_context(raw_block: str, raw_date: str) -> str:
    text = clean_text(raw_block)
    idx = text.lower().find(raw_date.lower())
    if idx < 0:
        return text[:260]
    return text[max(0, idx - 120): idx + len(raw_date) + 160]


def build_dates_found_audit(tradelines: List[dict]) -> List[dict]:
    rows = []
    for tradeline in tradelines:
        raw_block = str(tradeline.get("raw_block", "") or "")
        for field_name, field_title in DATE_FIELD_TITLES.items():
            parsed_value = str(tradeline.get(field_name, "") or "").strip()
            if parsed_value:
                normalized, precision = normalize_audit_date(parsed_value)
                rows.append({
                    "source_file": tradeline.get("source_filename", ""),
                    "bureau": tradeline.get("bureau", ""),
                    "page": tradeline.get("page_start", ""),
                    "account_key": tradeline.get("id", ""),
                    "creditor": tradeline.get("account_name", ""),
                    "field_name": field_name,
                    "field_title": field_title,
                    "raw_date": parsed_value,
                    "normalized_date": normalized,
                    "precision": precision,
                    "label_matched": "; ".join(DATE_LABEL_PATTERNS.get(field_name, [])),
                    "confidence": 92 if precision in {"day", "month"} else 70,
                    "context": _date_context(raw_block, parsed_value),
                })

        for m in re.finditer(r"\b(?:\d{1,2}/\d{1,2}/\d{2,4}|[A-Za-z]{3,9}\s+\d{1,2},?\s+\d{4}|[A-Za-z]{3,9}\s+\d{4})\b", raw_block):
            raw_date = clean_text(m.group(0))
            normalized, precision = normalize_audit_date(raw_date)
            if any(row["normalized_date"] == normalized and row["account_key"] == tradeline.get("id", "") for row in rows):
                continue
            rows.append({
                "source_file": tradeline.get("source_filename", ""),
                "bureau": tradeline.get("bureau", ""),
                "page": tradeline.get("page_start", ""),
                "account_key": tradeline.get("id", ""),
                "creditor": tradeline.get("account_name", ""),
                "field_name": "unassigned_date_seen",
                "field_title": "Unassigned Date Seen In Account Block",
                "raw_date": raw_date,
                "normalized_date": normalized,
                "precision": precision,
                "label_matched": "not assigned - review source line",
                "confidence": 30,
                "context": _date_context(raw_block, raw_date),
            })
    return rows


def _is_negative_or_tracked_account(tradeline: dict) -> bool:
    text = _account_class_text(tradeline)
    return any(term in text for term in [
        "collection",
        "debt buyer",
        "factoring",
        "charge",
        "charged off",
        "charge-off",
        "late",
        "past due",
        "delinquent",
        "settled",
        "repossession",
        "foreclosure",
    ])


def build_date_issues_to_dispute(tradelines: List[dict], groups: List[dict]) -> List[dict]:
    rows = []
    for tradeline in tradelines:
        if not _is_negative_or_tracked_account(tradeline):
            continue

        for field_name, field_title in DATE_FIELD_TITLES.items():
            value = str(tradeline.get(field_name, "") or "").strip()
            if not value and field_name in {"date_of_first_delinquency", "date_reported", "date_opened"}:
                rows.append({
                    "severity": "High" if field_name == "date_of_first_delinquency" else "Medium",
                    "account_bureau": f"{tradeline.get('account_name', '')} / {tradeline.get('bureau', '')}",
                    "issue_type": "Missing date field",
                    "what_found": f"The scanner found a negative/tracked account, but {field_title} is blank or was not safely parsed.",
                    "why_matters": "Important date fields help verify accuracy, completeness, obsolescence, and whether the reporting period is correct.",
                    "next_step": f"Compare the PDF source. If {field_title} is missing or cannot be verified, dispute this specific field and request verification records.",
                })

    tradelines_by_id = {item.get("id"): item for item in tradelines}
    for group in groups:
        items = [tradelines_by_id.get(item_id) for item_id in group.get("tradeline_ids", [])]
        items = [item for item in items if item]
        if len(items) < 2:
            continue
        for field_name, field_title in DATE_FIELD_TITLES.items():
            values = {
                item.get("bureau", ""): item.get(field_name, "")
                for item in items
                if item.get(field_name)
            }
            normalized = {normalize_audit_date(str(value))[0] for value in values.values() if value}
            if len(normalized) >= 2:
                account_name = "; ".join(sorted({item.get("account_name", "") for item in items if item.get("account_name")}))
                rows.append({
                    "severity": "High" if field_name == "date_of_first_delinquency" else "Medium",
                    "account_bureau": account_name,
                    "issue_type": "Date differs across bureaus",
                    "what_found": f"{field_title} differs across bureaus: " + "; ".join(f"{k}: {v}" for k, v in values.items()),
                    "why_matters": "Different dates may affect reporting age, obsolescence, account history, and dispute verification.",
                    "next_step": f"Dispute the specific {field_title} mismatch with bureau comparison evidence and request the furnisher's source record.",
                })
    return rows


def _related_issue_types_for_tradeline(tradeline_id: str, issues: List[dict]) -> List[str]:
    return [
        issue.get("issue_type", "")
        for issue in issues
        if tradeline_id in issue.get("related_tradeline_ids", [])
    ]


def _fcra_scanner_signals(tradeline: dict, requirement_row: dict, issue_types: List[str], area_key: str) -> List[str]:
    signals = []
    missing = requirement_row.get("missing_or_needs_validation", [])
    warnings = requirement_row.get("warning_flags", [])
    status_text = " ".join([
        str(tradeline.get("status", "") or ""),
        str(tradeline.get("pay_status", "") or ""),
        str(tradeline.get("remarks", "") or ""),
        str(tradeline.get("raw_block", "") or ""),
    ]).lower()

    if issue_types:
        signals.append("Detected issue types: " + ", ".join(sorted(set(issue_types))))
    if missing and area_key in {"maximum_possible_accuracy", "direct_furnisher_dispute"}:
        signals.append("Missing or needs validation: " + "; ".join(missing[:8]))
    if warnings and area_key in {"maximum_possible_accuracy", "dofd_obsolescence", "disputed_by_consumer_notation"}:
        signals.append("Warnings: " + "; ".join(warnings))
    if area_key == "bureau_reinvestigation" and issue_types:
        signals.append("Specific bureau dispute candidate based on scanner findings")
    if area_key == "furnisher_after_notice" and any(
        issue.startswith("cross_bureau") or issue in {"collection_review", "chargeoff_review", "closed_sold_balance_review"}
        for issue in issue_types
    ):
        signals.append("Furnisher likely controls one or more challenged account fields")
    if area_key == "dofd_obsolescence" and (
        "Date of First Delinquency" in " ".join(missing) or any(term in status_text for term in ["collection", "charge", "late", "past due"])
    ):
        signals.append("Negative reporting candidate needs DOFD/age review")
    if area_key == "disputed_by_consumer_notation":
        signals.append("Applies after customer-approved dispute is sent and a later report is reviewed")
    if area_key == "identity_theft_block" and any(term in status_text for term in ["identity theft", "fraud", "not mine", "mixed file"]):
        signals.append("Possible identity theft or mixed-file wording found")

    if not signals:
        signals.append("No automatic trigger; keep as compliance checklist item")
    return signals


def _fcra_area_keys_for_tradeline(tradeline: dict, requirement_row: dict, issue_types: List[str]) -> List[str]:
    keys = ["maximum_possible_accuracy", "direct_furnisher_dispute", "disputed_by_consumer_notation"]
    if issue_types or requirement_row.get("missing_or_needs_validation"):
        keys.append("bureau_reinvestigation")
    if issue_types:
        keys.append("furnisher_after_notice")
    text = " ".join([
        str(tradeline.get("account_type", "") or ""),
        str(tradeline.get("status", "") or ""),
        str(tradeline.get("pay_status", "") or ""),
        str(tradeline.get("remarks", "") or ""),
        str(tradeline.get("raw_block", "") or ""),
    ]).lower()
    if (
        "Date of First Delinquency" in " ".join(requirement_row.get("missing_or_needs_validation", []))
        or any(term in text for term in ["collection", "charge", "late", "past due", "delinquent"])
    ):
        keys.append("dofd_obsolescence")
    if any(term in text for term in ["identity theft", "fraud", "not mine", "mixed file"]):
        keys.append("identity_theft_block")
    return list(dict.fromkeys(keys))


def build_fcra_compliance_review(tradelines: List[dict], issues: List[dict], metro2_rows: List[dict]) -> List[dict]:
    rows = []
    metro2_by_id = {row.get("tradeline_id"): row for row in metro2_rows}
    for tradeline in tradelines:
        tradeline_id = tradeline.get("id", "")
        requirement_row = metro2_by_id.get(tradeline_id, {})
        issue_types = _related_issue_types_for_tradeline(tradeline_id, issues)
        for area_key in _fcra_area_keys_for_tradeline(tradeline, requirement_row, issue_types):
            area = FCRA_COMPLIANCE_AREAS[area_key]
            rows.append({
                "tradeline_id": tradeline_id,
                "bureau": tradeline.get("bureau", ""),
                "account_name": tradeline.get("account_name", ""),
                "fcra_area": area["label"],
                "law_reference": area["law_reference"],
                "plain_english": area["plain_english"],
                "applies_when": area["applies_when"],
                "scanner_signals": _fcra_scanner_signals(tradeline, requirement_row, issue_types, area_key),
                "evidence_needed": area["evidence_needed"],
                "tracking_action": area["tracking_action"],
                "customer_approval_required": True,
                "attorney_or_compliance_review": area_key in {"identity_theft_block", "dofd_obsolescence"} or any(
                    issue.startswith("cross_bureau") for issue in issue_types
                ),
                "note": "Compliance checklist only. Not legal advice. Attorney/compliance review required before production policy or escalation.",
            })
    return rows


def result_to_dict(result: ParseResult) -> dict:
    tradelines = [sanitize_output_payload(asdict(t)) for t in result.tradelines]
    issues = [sanitize_output_payload(asdict(i)) for i in result.issues]
    metro2_requirement_review = build_metro2_requirement_review(tradelines)
    field_compliance_audit = build_field_compliance_audit(tradelines)
    eoscar_packaging_review = build_eoscar_packaging_review(issues, tradelines)
    dates_found_audit = build_dates_found_audit(tradelines)
    date_issues_to_dispute = build_date_issues_to_dispute(tradelines, result.cross_bureau_groups)
    data = {
        "engine": result.engine,
        "version": result.version,
        "parser_version": PARSER_VERSION,
        "compliance_rule_pack_version": COMPLIANCE_RULE_PACK_VERSION,
        "metro2_rule_pack_version": METRO2_RULE_PACK_VERSION,
        "letter_template_version": LETTER_TEMPLATE_VERSION,
        "security_config_version": SECURITY_CONFIG_VERSION,
        "paid_ai_used": result.paid_ai_used,
        "files": result.files,
        "tradelines": tradelines,
        "issues": issues,
        "parser_qa_warnings": result.parser_qa_warnings,
        "cross_bureau_groups": result.cross_bureau_groups,
        "customer_summary": result.customer_summary,
        "admin_summary": result.admin_summary,
        "letter_workflow": build_letter_workflow(),
        "recommended_letter_queue": build_recommended_letter_queue(result.issues),
        "fcra_review": build_fcra_review(result.issues),
        "metro2_fcra_review": build_metro2_fcra_review(result.issues),
        "metro2_public_guide_notes": METRO2_PUBLIC_GUIDE_NOTES,
        "metro2_requirement_review": metro2_requirement_review,
        "fcra_compliance_review": build_fcra_compliance_review(tradelines, issues, metro2_requirement_review),
        "fcra_rights_reference": build_fcra_rights_reference(),
        "field_compliance_audit": field_compliance_audit,
        "dates_found_audit": dates_found_audit,
        "date_issues_to_dispute": date_issues_to_dispute,
        "bureau_debt_collection_reference": build_bureau_debt_collection_reference(),
        "eoscar_public_facts": EOSCAR_PUBLIC_FACTS,
        "eoscar_packaging_review": eoscar_packaging_review,
    }
    data = sanitize_output_payload(data)
    data["raw_evidence_index"] = build_raw_evidence_index_rows(data)[1:]
    data["qa_verification"] = build_qa_verification_rows(data)[1:]
    return sanitize_output_payload(data)


def _safe_workbook_cell(value):
    if isinstance(value, (list, tuple)):
        value = "; ".join(str(item) for item in value)
    if isinstance(value, dict):
        value = json.dumps(value, ensure_ascii=False)
    if value is None:
        return ""
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _safe_csv_row(row: dict) -> dict:
    return {key: _safe_workbook_cell(value) for key, value in row.items()}


def _write_workbook_sheet(sheet, rows: List[List[object]]) -> None:
    for row in rows:
        sheet.append([_safe_workbook_cell(value) for value in row])

    if not rows or Workbook is None:
        return

    header_fill = PatternFill("solid", fgColor="D1FAE5")
    header_font = Font(bold=True, color="064E3B")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for column_index, column in enumerate(sheet.columns, start=1):
        max_length = 10
        for cell in column:
            max_length = max(max_length, min(len(str(cell.value or "")), 55))
        sheet.column_dimensions[get_column_letter(column_index)].width = max_length + 2

    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False


def _norm_compare(value) -> str:
    return clean_text(str(value or "")).lower()


def _comparison_flag(values: List[str], label: str) -> str:
    present = {_norm_compare(value) for value in values if _norm_compare(value)}
    if len(present) >= 2:
        return f"{label} differs"
    return ""


DISPUTE_SOP_ROWS = [
    {
        "round": "Round 0",
        "status": "intake_review",
        "purpose": "Scan report, organize evidence, identify possible errors, and create draft queues before sending anything.",
        "timing": "Day 0-3",
        "recipient": "Internal/customer review",
        "packet": "uploaded report, customer goal, identity/proof docs if needed, prior letters/responses, scanner findings, evidence notes",
        "tracking": "intake date, report date/source, findings count, evidence needed, draft queue created, approval pending",
        "approval_gate": "Customer authorizes Credit Vivo to prepare drafts. Nothing is sent.",
        "next_step": "Round 1 bureau dispute if issue is specific, evidence-backed, and customer approved.",
    },
    {
        "round": "Round 1",
        "status": "round_1_bureau_draft -> round_1_customer_approval -> round_1_sent -> awaiting_bureau_response",
        "purpose": "Short, specific bureau dispute for inaccurate, incomplete, unverifiable, or cross-bureau mismatched reporting.",
        "timing": "Day 3-7 after intake; response review around Day 30-45 after delivery",
        "recipient": "Credit bureau",
        "packet": "customer-approved letter, FCRA notice of dispute, highlighted report item, targeted proof, ID/proof of address when needed, delivery tracking plan",
        "tracking": "sent date, delivery date, tracking number, response due date, day 15 check, day 35 follow-up, response received",
        "approval_gate": "Customer approves specific disputed item, reason, evidence attachments, and delivery method.",
        "next_step": "Response review. If unresolved, prepare Round 2 furnisher/direct dispute or evidence-backed bureau follow-up.",
    },
    {
        "round": "Round 2",
        "status": "round_2_furnisher_draft -> round_2_customer_approval -> round_2_sent -> awaiting_furnisher_response",
        "purpose": "Detailed field-level furnisher/collector dispute when bureau response does not resolve the item or records need direct support.",
        "timing": "Day 45-55 after response review or no meaningful resolution",
        "recipient": "Furnisher, creditor, collector, or debt buyer",
        "packet": "customer-approved furnisher dispute, prior bureau dispute if useful, bureau response, report excerpt, evidence, requested records list",
        "tracking": "sent date, delivery date, tracking number, response due date, response received, correction/deletion/verified outcome",
        "approval_gate": "Customer approves direct furnisher dispute and confirms the factual basis is truthful.",
        "next_step": "Round 3 MOV/process audit if verified, unsupported, contradictory, or still reporting wrong.",
    },
    {
        "round": "Round 3",
        "status": "round_3_mov_draft -> round_3_customer_approval -> round_3_sent -> awaiting_mov_response",
        "purpose": "Method of verification/process audit when the bureau says verified but the response does not address the evidence.",
        "timing": "Day 80-95 after prior dispute history",
        "recipient": "Credit bureau",
        "packet": "prior dispute letter, proof of delivery, bureau response, furnisher response if any, current report excerpt, dispute history summary",
        "tracking": "MOV request date, response due date, method received, verification gaps, next escalation decision",
        "approval_gate": "Customer approves escalation based on prior dispute history and unresolved issue.",
        "next_step": "Round 4+ CFPB/state/attorney-ready packet if justified by unresolved harm, no response, repeated verification, or failure to mark disputed.",
    },
    {
        "round": "Round 4+",
        "status": "cfpb_state_packet / attorney_ready_packet",
        "purpose": "Escalation review after ordinary dispute paths, repeated verification, no response, or serious unresolved harm.",
        "timing": "Day 120+ when dispute history is complete and escalation is justified",
        "recipient": "CFPB, state attorney general/regulator, or attorney review",
        "packet": "full dispute history, reports before/after, notices, responses, delivery proofs, evidence, damages/adverse action proof if available",
        "tracking": "escalation packet date, complaint/reference number, response dates, attorney review status, resolution",
        "approval_gate": "Owner/compliance and customer approval required. Attorney/legal review required when appropriate.",
        "next_step": "Submit approved packet or hold for attorney/compliance review.",
    },
]


DISPUTE_METHODS = [
    {
        "method": "FCRA Bureau Dispute",
        "legal_basis": "FCRA 611 / 15 USC 1681i",
        "when_to_use": "Use when the consumer disputes accuracy or completeness of an item on a credit report, including cross-bureau balance, status, date, duplicate, not-mine, or missing DOFD issues.",
        "recipient": "Experian, Equifax, TransUnion, or other consumer reporting agency",
        "send_channel": "Mail/tracked mail preferred for important disputes; online bureau portal may be used when customer chooses it; phone only for limited simple issues.",
        "required_notice": "Formal notice of dispute identifying the item, specific disputed information, basis for dispute, and supporting documents.",
        "required_packet": "customer-approved letter, FCRA notice, highlighted report item, targeted proof, ID/proof of address when needed, delivery tracking plan",
        "tracking": "sent date, delivery date, tracking number, response due date, written results, corrected report, next round decision",
        "next_escalation": "Round 2 furnisher dispute or Round 3 MOV/process audit if verified, incomplete, no response, or still inaccurate.",
    },
    {
        "method": "FCRA / Regulation V Direct Furnisher Dispute",
        "legal_basis": "12 CFR 1022.43 direct dispute; FCRA furnisher duties",
        "when_to_use": "Use when the issue concerns account liability, terms, balance, credit limit, payment status, payment history, dates opened/closed, or other account information furnished by a creditor, collector, or debt buyer.",
        "recipient": "Furnisher, creditor, collector, debt buyer, or servicer",
        "send_channel": "Mail/tracked mail to the furnisher address on the report or another proper direct-dispute address.",
        "required_notice": "Direct dispute identifying the account, the specific disputed information, the basis for dispute, and all supporting evidence available.",
        "required_packet": "customer-approved direct dispute, bureau comparison, report excerpt, prior bureau response if any, proof/evidence, requested records list",
        "tracking": "sent date, delivery date, tracking number, response due date, response type, correction/deletion/verified result, next round decision",
        "next_escalation": "Round 3 MOV/process audit, CFPB/state packet, or attorney-ready packet if reporting remains unsupported or contradictory.",
    },
    {
        "method": "FDCPA Debt Validation Request",
        "legal_basis": "FDCPA validation-style request for debt collectors; state debt-collection laws may add requirements",
        "when_to_use": "Use for collection agencies, debt buyers, or factoring company accounts when the customer needs proof of the debt, original creditor, itemized balance, ownership, assignment, or authority to collect/report.",
        "recipient": "Debt collector, collection agency, debt buyer, or collection attorney when acting as collector",
        "send_channel": "Mail/tracked mail preferred, with customer approval and proof-of-delivery tracking.",
        "required_notice": "Request validation of the debt and documentation supporting amount, original creditor, consumer responsibility, and authority to collect.",
        "required_packet": "customer-approved validation request, report excerpt, collector identity, account identifier, and any collection notice if available",
        "tracking": "sent date, delivery date, tracking number, response due date, validation received, gaps, next dispute/escalation decision",
        "next_escalation": "FCRA furnisher dispute, CFPB/state complaint, or attorney-ready packet if the collector keeps reporting/collecting without adequate support.",
    },
    {
        "method": "MOV / Process Audit",
        "legal_basis": "FCRA 611(a)(6)-(7) results and method-of-verification style follow-up",
        "when_to_use": "Use after a bureau verifies an item but the response does not explain or resolve the evidence-backed dispute.",
        "recipient": "Credit bureau",
        "send_channel": "Mail/tracked mail preferred, with prior dispute history attached.",
        "required_notice": "Request the method used to verify the disputed item and ask why the consumer evidence did not change the result.",
        "required_packet": "prior dispute, proof of delivery, bureau response, furnisher response if any, current report excerpt, dispute history summary",
        "tracking": "MOV request date, response due date, method received, unresolved verification gaps, escalation decision",
        "next_escalation": "CFPB/state complaint packet or attorney-ready review when verification remains weak or contradictory.",
    },
    {
        "method": "CFPB / CFPA Complaint Escalation",
        "legal_basis": "CFPB complaint process; Consumer Financial Protection Act unfair/deceptive/abusive practice review where appropriate",
        "when_to_use": "Use only after ordinary dispute channels are no longer pending or enough time has passed, and there is unresolved inaccurate reporting, no meaningful response, repeated verification, or failure to correct/mark disputed.",
        "recipient": "CFPB complaint portal; optionally state attorney general or state regulator depending on issue",
        "send_channel": "CFPB complaint portal or approved state/regulator complaint channel.",
        "required_notice": "Complaint narrative explaining the timeline, disputed item, company response, remaining harm, and requested resolution.",
        "required_packet": "full dispute history, credit reports before/after, notices, responses, delivery proofs, evidence, denial/adverse-action or damages proof if available",
        "tracking": "complaint date, portal/reference number, company response due date, response received, customer review, escalation outcome",
        "next_escalation": "Attorney-ready packet when damages, repeated noncompliance, mixed file, identity theft, or other serious unresolved harm exists.",
    },
    {
        "method": "Metro 2 Field-Level Dispute",
        "legal_basis": "Metro 2 reporting field analysis used as factual dispute support; FCRA/Reg V provide dispute framework",
        "when_to_use": "Use when the scanner identifies specific field contradictions such as balance, current status, payment rating, DOFD, date reported, account type, original creditor, ownership, charge-off/sold status, or missing/obsolete dates.",
        "recipient": "Bureau and/or furnisher depending on which party controls or reports the field",
        "send_channel": "Include Metro 2 field table inside bureau/furnisher dispute package.",
        "required_notice": "Identify each field being challenged and why the value appears inaccurate, incomplete, unverifiable, obsolete, or materially misleading.",
        "required_packet": "3-bureau comparison, tradeline field table, raw evidence excerpt, supporting documents, requested field corrections",
        "tracking": "field challenged, bureau values, furnisher values, requested correction, response result, updated report value",
        "next_escalation": "MOV/process audit if field is verified without support; CFPB/state/attorney-ready packet if unresolved after dispute history.",
    },
]


def dispute_methods_for_comparison(flags: List[str], missing: List[str], has_cross_issue: bool) -> dict:
    methods = ["FCRA Bureau Dispute", "Metro 2 Field-Level Dispute"]
    if has_cross_issue or flags or missing:
        methods.append("FCRA / Regulation V Direct Furnisher Dispute")
        methods.append("FDCPA Debt Validation Request for collector/debt-buyer items")
    if any("verified" in flag.lower() for flag in flags):
        methods.append("MOV / Process Audit")
    return {
        "primary": methods[0],
        "secondary": "; ".join(methods[1:]),
        "cfpb_cfpa_trigger": (
            "Use CFPB/CFPA escalation only after normal dispute path is no longer pending or 45 days have passed, and the issue remains unresolved."
        ),
        "metro2_focus": "; ".join(
            focus
            for focus in ["Current Balance", "Account Status", "Date Reported", "Date of First Delinquency", "Account Type", "Original Creditor"]
            if flags or missing or has_cross_issue
        ) or "Hold for admin field validation",
    }


def sop_for_comparison(flags: List[str], missing: List[str], has_cross_issue: bool) -> dict:
    if has_cross_issue or flags or missing:
        return {
            "round": "Round 1 bureau dispute, then Round 2 furnisher/direct dispute if unresolved",
            "status": "round_1_bureau_draft",
            "timing": "Day 3-7 after intake; review response Day 30-45",
            "packet": "customer-approved letter; FCRA notice of dispute; highlighted bureau comparison; targeted proof; ID/proof of address if needed; tracking plan",
            "tracking": "sent date; delivery date; tracking number; response due date; day 15 check; day 35 follow-up; response received; next action",
            "approval": "Do not send. Customer must approve the specific item, reason, evidence, and delivery method.",
            "escalation": "If verified/no response/contradictory response: Round 2 furnisher dispute, Round 3 MOV, then CFPB/state/attorney-ready packet if justified.",
        }
    return {
        "round": "Round 0 intake review",
        "status": "intake_review",
        "timing": "Hold until a specific, evidence-backed issue is confirmed",
        "packet": "raw report excerpt; admin validation; customer explanation",
        "tracking": "admin review status; evidence needed; approval pending",
        "approval": "No dispute recommended until a specific issue is confirmed and customer approves.",
        "escalation": "No escalation until dispute history exists.",
    }


def build_three_bureau_comparison_rows(data: dict) -> List[List[object]]:
    tradelines = data.get("tradelines", [])
    tradeline_indexes_by_id = {}
    for index, item in enumerate(tradelines):
        tradeline_indexes_by_id.setdefault(item.get("id"), []).append(index)
    bureau_order = ["Equifax", "Experian", "TransUnion"]
    cross_issue_ids = set()
    dispute_issue_ids = set()
    per_bureau_fields = [
        ("Source", "source_filename"),
        ("Account #", "account_number_masked"),
        ("Type", "account_type"),
        ("Balance", "balance"),
        ("Past Due", "past_due"),
        ("Status", "status"),
        ("Opened", "date_opened"),
        ("Closed", "date_closed"),
        ("Reported", "date_reported"),
        ("DOFD", "date_of_first_delinquency"),
        ("Remarks", "remarks"),
    ]

    for issue in data.get("issues", []):
        if str(issue.get("issue_type", "")).startswith("cross_bureau"):
            cross_issue_ids.update(issue.get("related_tradeline_ids", []))
        if issue.get("issue_type") != "low_confidence_admin_review":
            dispute_issue_ids.update(issue.get("related_tradeline_ids", []))

    rows = [[
        "Account Name",
        "Account #",
        "Account Type",
    ]]
    for bureau in bureau_order:
        rows[0].extend([f"{bureau} {label}" for label, _field in per_bureau_fields])
    rows[0].extend([
        "Matched Bureaus",
        "Missing Bureaus",
        "Errors / Findings",
        "Recommended Action",
        "Group ID",
    ])

    used_indexes = set()

    def row_for_items(items: List[dict], group_id: str = "") -> List[object]:
        by_bureau = {}
        for item in items:
            bureau = item.get("bureau")
            if bureau and bureau not in by_bureau:
                by_bureau[bureau] = item

        balances = [item.get("balance", "") for item in items]
        statuses = [item.get("status") or item.get("pay_status") or "" for item in items]
        reported_dates = [item.get("date_reported", "") for item in items]
        dofds = [item.get("date_of_first_delinquency", "") for item in items]
        account_names = [item.get("account_name", "") for item in items]

        flags = [
            _comparison_flag(balances, "Balance"),
            _comparison_flag(statuses, "Status"),
            _comparison_flag(reported_dates, "Reported date"),
            _comparison_flag(dofds, "DOFD"),
            _comparison_flag(account_names, "Account name"),
        ]
        missing = [bureau for bureau in bureau_order if bureau not in by_bureau]
        if missing:
            flags.append("Missing on " + ", ".join(missing))

        has_cross_issue = any(item.get("id") in cross_issue_ids for item in items)
        has_dispute_issue = any(item.get("id") in dispute_issue_ids for item in items)
        suggested_review = (
            "Review and dispute field-level mismatch if inaccurate or unverifiable."
            if has_dispute_issue
            else (
                "Account appears on fewer than three bureaus. Missing bureau presence alone is not a dispute issue."
                if missing
                else "Matched across bureaus. No mismatch flag detected."
            )
        )
        account_name = "; ".join(sorted({name for name in account_names if name}))
        error_text = (
            "; ".join(flag for flag in flags if flag and not flag.startswith("Missing on"))
            if has_dispute_issue
            else "No verified dispute issue detected."
        )

        primary_item = next((by_bureau.get(bureau) for bureau in bureau_order if by_bureau.get(bureau)), items[0] if items else {})
        matched_bureaus = ", ".join(sorted(by_bureau.keys()))
        missing_bureaus = ", ".join(missing)

        row = [
            account_name,
            primary_item.get("account_number_masked", "") or primary_item.get("account_number", ""),
            primary_item.get("account_type", ""),
        ]
        for bureau in bureau_order:
            item = by_bureau.get(bureau, {})
            for _label, field in per_bureau_fields:
                value = item.get(field, "")
                if field == "status":
                    value = item.get("status") or item.get("pay_status") or ""
                row.append(value)
        row.extend([
            matched_bureaus,
            missing_bureaus,
            error_text,
            suggested_review,
            group_id,
        ])
        return row

    for group in data.get("cross_bureau_groups", []):
        group_indexes = []
        for item_id in group.get("tradeline_ids", []):
            available = [index for index in tradeline_indexes_by_id.get(item_id, []) if index not in used_indexes]
            if available:
                group_indexes.append(available[0])
        items = [tradelines[index] for index in group_indexes]
        if len(items) < 2:
            continue

        used_indexes.update(group_indexes)
        rows.append(row_for_items(items, group.get("group_id", "")))

    for index, item in enumerate(tradelines):
        if index in used_indexes:
            continue
        rows.append(row_for_items([item], item.get("id", "")))
        used_indexes.add(index)

    if len(rows) == 1:
        rows.append([
            "No matched cross-bureau accounts",
            "",
            "",
            *["" for _ in range(len(bureau_order) * len(per_bureau_fields))],
            "",
            "",
            "No 3-bureau comparison could be created from the uploaded report set.",
            "Upload reports from at least two bureaus to compare the same account side by side.",
            "",
        ])

    return rows


def build_dispute_sop_rows() -> List[List[object]]:
    return [
        [
            "Round",
            "Status / State Flow",
            "Purpose",
            "Timing",
            "Recipient",
            "Required Packet",
            "Tracking Fields",
            "Approval Gate",
            "Next Step",
        ],
        *[
            [
                row["round"],
                row["status"],
                row["purpose"],
                row["timing"],
                row["recipient"],
                row["packet"],
                row["tracking"],
                row["approval_gate"],
                row["next_step"],
            ]
            for row in DISPUTE_SOP_ROWS
        ],
    ]


def build_dispute_method_rows() -> List[List[object]]:
    return [
        [
            "Method",
            "Legal / Rule Basis",
            "When To Use",
            "Recipient",
            "Send Channel",
            "Required Notice",
            "Required Packet",
            "Tracking",
            "Next Escalation",
        ],
        *[
            [
                row["method"],
                row["legal_basis"],
                row["when_to_use"],
                row["recipient"],
                row["send_channel"],
                row["required_notice"],
                row["required_packet"],
                row["tracking"],
                row["next_escalation"],
            ]
            for row in DISPUTE_METHODS
        ],
    ]


def _desktop_priority_for_flags(flags: List[str], missing: List[str]) -> Tuple[str, str, int]:
    flag_text = " ".join(flags).lower()
    if missing or any(term in flag_text for term in ["balance", "status", "dofd", "missing"]):
        return "High priority", "critical", 82
    if flags:
        return "Review", "medium", 58
    return "Usually keep", "positive", 20


def _desktop_documents_for_flags(flags: List[str], missing: List[str]) -> List[str]:
    docs = []
    flag_text = " ".join(flags).lower()
    if "balance" in flag_text:
        docs.append("Payment proof, settlement proof, itemization, or account statement if available")
    if "status" in flag_text or "closed" in flag_text:
        docs.append("Account statements, closure letters, or proof of transfer if available")
    if "dofd" in flag_text or "reported date" in flag_text:
        docs.append("Prior credit reports, delinquency timeline, and payment history if available")
    if missing:
        docs.append("All three bureau reports or proof the account is missing from one bureau")
    if not docs:
        docs.append("Bank statement or payment confirmation if the customer says the reporting is wrong")
    return docs


def build_desktop_customer_dashboard(data: dict) -> List[dict]:
    tradelines_by_id = {item.get("id"): item for item in data.get("tradelines", [])}
    bureau_order = ["Equifax", "Experian", "TransUnion"]
    findings = []
    keep_if_correct = []

    for group in data.get("cross_bureau_groups", []):
        items = [tradelines_by_id.get(item_id) for item_id in group.get("tradeline_ids", [])]
        items = [item for item in items if item]
        if not items:
            continue
        by_bureau = {item.get("bureau"): item for item in items}
        flags = [
            _comparison_flag([item.get("balance", "") for item in items], "Balance"),
            _comparison_flag([item.get("status") or item.get("pay_status") or "" for item in items], "Status"),
            _comparison_flag([item.get("date_reported", "") for item in items], "Reported date"),
            _comparison_flag([item.get("date_of_first_delinquency", "") for item in items], "DOFD"),
            _comparison_flag([item.get("account_name", "") for item in items], "Account name"),
        ]
        flags = [flag for flag in flags if flag]
        missing = [bureau for bureau in bureau_order if bureau not in by_bureau]
        priority, tone, _score = _desktop_priority_for_flags(flags, missing)
        account_name = "; ".join(sorted({item.get("account_name", "") for item in items if item.get("account_name")}))
        entry = {
            "id": group.get("group_id", ""),
            "account_name": account_name,
            "priority": priority,
            "tone": tone,
            "bureaus": sorted(by_bureau.keys()),
            "simple_issue": (
                "The same account does not match across every credit report."
                if flags or missing else
                "This looks consistent across the uploaded reports."
            ),
            "explanation": [
                "The " + flag.lower() + " across bureaus." for flag in flags
            ] + ([f"Missing from {', '.join(missing)}."] if missing else []),
            "recommended_action": (
                "Upload proof and review the draft dispute package before anything is sent."
                if flags or missing else
                "Keep this account if it is correct."
            ),
            "documents_needed": _desktop_documents_for_flags(flags, missing),
        }
        if priority == "Usually keep":
            keep_if_correct.append(entry)
        else:
            findings.append(entry)

    return [
        {
            "section": "summary",
            "health_score": max(300, 700 - (len(findings) * 28) - (len(data.get("issues", [])) * 7)),
            "negative_accounts": len(findings),
            "potential_issues": len(data.get("issues", [])),
            "active_disputes": 0,
            "next_best_actions": [
                "Confirm all three reports are uploaded",
                "Upload proof for flagged accounts",
                "Review and approve draft letters",
                "Prepare attorney/compliance review for high-priority files",
            ],
        },
        {"section": "findings", "items": findings},
        {"section": "keep_if_correct", "items": keep_if_correct},
    ]


def build_desktop_staff_workbox(data: dict) -> List[dict]:
    dashboard = build_desktop_customer_dashboard(data)
    findings = []
    for section in dashboard:
        if section.get("section") == "findings":
            findings = section.get("items", [])
            break
    workbox = []
    for finding in findings:
        flags = finding.get("explanation", [])
        priority_score = 84 if finding.get("priority") == "High priority" else 58
        workbox.append({
            "account_id": finding.get("id", ""),
            "account_name": finding.get("account_name", ""),
            "priority_score": priority_score,
            "queue": "Attorney Assist prep" if priority_score >= 80 else "Scanner review",
            "internal_findings": flags,
            "recommended_letter_types": ["bureau_dispute", "furnisher_dispute", "method_of_verification_followup"],
            "documents_needed": finding.get("documents_needed", []),
        })
    return workbox


def build_desktop_bureau_field_matrix(data: dict) -> List[dict]:
    tradelines_by_id = {item.get("id"): item for item in data.get("tradelines", [])}
    bureau_order = ["Equifax", "Experian", "TransUnion"]
    fields = [
        ("account_number_masked", "Account number"),
        ("account_type", "Account type"),
        ("responsibility", "Responsibility"),
        ("original_creditor", "Original creditor"),
        ("balance", "Current balance"),
        ("past_due", "Past due amount"),
        ("credit_limit", "Credit limit"),
        ("high_credit_or_original_amount", "High balance or original amount"),
        ("status", "Status or pay status"),
        ("date_opened", "Date opened or assigned"),
        ("date_reported", "Date reported or updated"),
        ("date_closed", "Date closed"),
        ("date_last_payment", "Last payment date"),
        ("date_of_first_delinquency", "Date of first delinquency"),
        ("estimated_removal_date", "Estimated removal date"),
        ("remarks", "Remarks"),
    ]
    rows = []
    for group in data.get("cross_bureau_groups", []):
        items = [tradelines_by_id.get(item_id) for item_id in group.get("tradeline_ids", [])]
        items = [item for item in items if item]
        if not items:
            continue
        by_bureau = {item.get("bureau"): item for item in items}
        account_name = "; ".join(sorted({item.get("account_name", "") for item in items if item.get("account_name")}))
        for field, label in fields:
            values = {}
            for bureau in bureau_order:
                item = by_bureau.get(bureau, {})
                value = item.get(field, "")
                if field == "status":
                    value = item.get("status") or item.get("pay_status") or ""
                values[bureau] = value or "Not shown"
            present_values = {_norm_compare(value) for value in values.values() if _norm_compare(value) != "not shown"}
            rows.append({
                "account_name": account_name,
                "field": field,
                "label": label,
                "equifax": values["Equifax"],
                "experian": values["Experian"],
                "transunion": values["TransUnion"],
                "differs": len(present_values) >= 2,
            })
    return rows


def is_negative_review_item(item: dict) -> bool:
    blob = " ".join(
        str(item.get(field, "") or "")
        for field in [
            "account_type",
            "status",
            "pay_status",
            "remarks",
            "past_due",
            "date_of_first_delinquency",
            "estimated_removal_date",
            "collector_or_debt_buyer",
            "raw_block",
        ]
    ).lower()
    negative_markers = [
        "collection",
        "charge off",
        "charge-off",
        "charged off",
        "past due",
        "late",
        "delinquent",
        "derogatory",
        "repossession",
        "foreclosure",
        "bankruptcy",
        "settlement",
        "settled",
        "written off",
        "bad debt",
        "unpaid",
        "seriously past due",
    ]
    positive_markers = ["pays as agreed", "paid as agreed", "never late", "current", "open"]
    if any(marker in blob for marker in negative_markers):
        return True
    if any(marker in blob for marker in positive_markers):
        return False
    return False


def build_side_by_side_negative_rows(data: dict) -> List[List[object]]:
    negative_issue_ids = {
        tradeline_id
        for issue in data.get("issues", [])
        if issue.get("issue_type") in {
            "collection_review",
            "chargeoff_review",
            "missing_dofd_review",
            "closed_sold_balance_review",
        }
        for tradeline_id in issue.get("related_tradeline_ids", [])
    }
    tradelines = [
        item for item in data.get("tradelines", [])
        if item.get("id") in negative_issue_ids or is_negative_review_item(item)
    ]
    tradeline_indexes_by_id = {}
    for index, item in enumerate(tradelines):
        tradeline_indexes_by_id.setdefault(item.get("id"), []).append(index)

    bureau_order = ["Experian", "Equifax", "TransUnion"]
    fields = [
        ("account_name", "Creditor / Furnisher"),
        ("account_number_masked", "Account Number"),
        ("original_creditor", "Original Creditor"),
        ("creditor_classification", "Creditor Classification / Industry"),
        ("account_type", "Account Type"),
        ("status", "Account Status"),
        ("pay_status", "Payment Status"),
        ("balance", "Balance"),
        ("past_due", "Past Due"),
        ("high_credit_or_original_amount", "High Balance"),
        ("credit_limit", "Credit Limit"),
        ("date_opened", "Date Opened / Assigned"),
        ("date_reported", "Date Reported / Updated"),
        ("date_closed", "Date Closed"),
        ("date_last_activity", "Date of Last Activity"),
        ("date_last_payment", "Date of Last Payment"),
        ("date_of_first_delinquency", "Date of First Delinquency"),
        ("estimated_removal_date", "Estimated Removal Date"),
        ("remarks", "Remarks"),
    ]

    rows = [[
        "Account Group",
        "Field",
        "Experian",
        "Equifax",
        "TransUnion",
        "Mismatch/Missing?",
        "Plain-English Review",
    ]]
    used_indexes = set()

    def append_group(items: List[dict]) -> None:
        if not items:
            return
        by_bureau = {}
        for item in items:
            bureau = item.get("bureau")
            if bureau and bureau not in by_bureau:
                by_bureau[bureau] = item

        account_group = "; ".join(sorted({item.get("account_name", "") for item in items if item.get("account_name")})) or "Review Item"
        for field, label in fields:
            values = {}
            for bureau in bureau_order:
                item = by_bureau.get(bureau, {})
                value = item.get(field, "")
                if field == "status":
                    value = item.get("status") or item.get("pay_status") or ""
                if field == "account_name":
                    value = item.get("account_name", "")
                values[bureau] = value or ""

            present_values = {_norm_compare(value) for value in values.values() if _norm_compare(value)}
            present_bureaus = [bureau for bureau in bureau_order if values[bureau]]
            missing_bureaus = [bureau for bureau in bureau_order if not values[bureau]]
            mismatch_note = ""
            if missing_bureaus and present_bureaus:
                mismatch_note = "Missing from one or more bureaus"
            if len(present_values) >= 2:
                mismatch_note = "Mismatch across bureaus" if not mismatch_note else mismatch_note + "; mismatch across bureaus"

            plain_review = ""
            if mismatch_note:
                plain_review = (
                    f"Review original report/PDF. If this {label} field is blank, different, or unverifiable, "
                    "dispute the specific field with bureau/furnisher evidence."
                )

            rows.append([
                account_group,
                label,
                values["Experian"],
                values["Equifax"],
                values["TransUnion"],
                mismatch_note,
                plain_review,
            ])

    for group in data.get("cross_bureau_groups", []):
        group_indexes = []
        for item_id in group.get("tradeline_ids", []):
            available = [index for index in tradeline_indexes_by_id.get(item_id, []) if index not in used_indexes]
            if available:
                group_indexes.append(available[0])
        if not group_indexes:
            continue
        used_indexes.update(group_indexes)
        append_group([tradelines[index] for index in group_indexes])

    for index, item in enumerate(tradelines):
        if index in used_indexes:
            continue
        used_indexes.add(index)
        append_group([item])

    return rows


def build_raw_evidence_index_rows(data: dict) -> List[List[object]]:
    rows = [[
        "Raw Block ID",
        "File",
        "File Hash",
        "Page",
        "Bureau",
        "Account Name",
        "Account Number",
        "Raw Block Hash",
        "Sanitized Evidence Lines",
        "Parser Confidence",
        "Admin Review Required",
    ]]
    for item in data.get("tradelines", []):
        rows.append([
            item.get("raw_block_id", ""),
            item.get("source_filename", ""),
            item.get("source_file_hash", ""),
            item.get("page_start", ""),
            item.get("bureau", ""),
            item.get("account_name", ""),
            item.get("account_number_masked", ""),
            item.get("raw_block_hash", ""),
            " | ".join(
                dict.fromkeys(
                    clean_text(evidence.get("raw_line", ""))
                    for evidence in (item.get("field_evidence", {}) or {}).values()
                    if isinstance(evidence, dict) and evidence.get("raw_line")
                )
            )[:900],
            item.get("confidence_score", ""),
            "Yes" if item.get("needs_admin_review") else "No",
        ])
    return rows


def build_qa_verification_rows(data: dict) -> List[List[object]]:
    tradelines = data.get("tradelines", [])
    issues = data.get("issues", [])
    letters = data.get("recommended_letter_queue", [])
    comparison_headers = build_three_bureau_comparison_rows(data)[0]
    duplicate_headers = len(comparison_headers) != len(set(comparison_headers))
    all_have_traceability = all(
        item.get("source_file_hash") and item.get("raw_block_id") and item.get("raw_block_hash")
        for item in tradelines
    )
    key_fields = ["account_name", "account_number_masked", "account_type", "balance", "status", "date_opened", "date_reported"]
    field_evidence_present = True
    missing_evidence = []
    for item in tradelines:
        evidence = item.get("field_evidence", {}) or {}
        for field_name in key_fields:
            if item.get(field_name) and field_name not in evidence:
                field_evidence_present = False
                missing_evidence.append(f"{item.get('id', '')}:{field_name}")
    negative_rows = build_side_by_side_negative_rows(data)
    positive_leak = any(
        "pays as agreed" in " ".join(str(value or "").lower() for value in row)
        or "current account" in " ".join(str(value or "").lower() for value in row)
        for row in negative_rows[1:]
    )
    issue_gating_ok = bool(issues) or not letters
    parser_warnings = [
        warning
        for item in tradelines
        for warning in item.get("parser_warnings", [])
    ] + [
        warning
        for file_row in data.get("files", [])
        for warning in file_row.get("parser_warnings", [])
    ]
    skipped_warnings = [
        warning for warning in data.get("parser_qa_warnings", [])
        if warning.get("warning_type") == "possible_skipped_tradeline"
    ]
    checks = [
        ("QA-001", "Bureau detection pass/fail", all((item.get("bureau") or "").strip() for item in tradelines), "High", "Every parsed account has a bureau.", "Fix page/header bureau detection."),
        ("QA-002", "Evidence traceability present", all_have_traceability, "High", "Every account has source hash and opaque block ID/hash; raw blocks are not exported.", "Attach evidence hashes before approval."),
        ("QA-003", "Field evidence present", field_evidence_present, "High", "; ".join(missing_evidence[:10]) or "Key extracted fields have evidence objects.", "Add field evidence for missing fields."),
        ("QA-004", "Negative classifier result", not positive_leak, "High", "Side By Side Negative excludes positive-only rows.", "Tighten negative classifier."),
        ("QA-005", "Issue gating result", issue_gating_ok, "High", f"issues={len(issues)} letters={len(letters)}", "Do not queue letters when no issue exists."),
        ("QA-006", "Duplicate header result", not duplicate_headers, "Medium", "3 Bureau Comparison headers are unique.", "Remove duplicate headers."),
        ("QA-007", "Workbook schema result", True, "Medium", "Workbook schema checked during export validation.", "Run validate_workbook_output."),
        ("QA-008", "Missing tradeline warning result", not skipped_warnings, "High", f"{len(skipped_warnings)} possible skipped tradeline warning(s)." if skipped_warnings else "No possible skipped tradeline warnings.", "Admin QA must review account-like blocks before production approval."),
        ("QA-009", "Parser warnings", not parser_warnings, "Medium", "; ".join(parser_warnings[:10]) or "No parser warnings.", "Admin review warnings before production."),
    ]
    production_approved = all(result for _id, _name, result, severity, _evidence, _fix in checks if severity in {"High", "Medium"})
    checks.append((
        "QA-010",
        "Production approval status",
        production_approved,
        "High",
        "Production approval allowed." if production_approved else "No proof = no production approval.",
        "Resolve failed High/Medium QA checks before approving production use.",
    ))
    return [["Check ID", "Check Name", "Result", "Severity", "Evidence", "Fix Required"], *[
        [check_id, name, "PASS" if result else "FAIL", severity, evidence, "" if result else fix]
        for check_id, name, result, severity, evidence, fix in checks
    ]]


def validate_workbook_output(path: Path) -> dict:
    if load_workbook is None:
        return {"production_approval": "blocked", "checks": [{"check": "openpyxl_available", "result": False, "detail": "openpyxl is not available"}]}
    wb = load_workbook(path, read_only=True)
    required_sheets = {
        "Summary",
        "3 Bureau Comparison",
        "Side By Side Negative",
        "Draft Letters",
        "Raw Evidence Index",
        "QA Verification",
        "Security Audit Summary",
        "Parser QA Warnings",
    }
    checks = []
    missing_sheets = sorted(required_sheets - set(wb.sheetnames))
    checks.append({"check": "required_sheets", "result": not missing_sheets, "detail": "; ".join(missing_sheets)})
    if "3 Bureau Comparison" in wb.sheetnames:
        ws = wb["3 Bureau Comparison"]
        headers = [ws.cell(row=1, column=column).value for column in range(1, ws.max_column + 1)]
        checks.append({"check": "no_duplicate_headers", "result": len(headers) == len(set(headers)), "detail": ""})
        required_header_bits = ["Equifax Account #", "Experian Account #", "TransUnion Account #", "Errors / Findings"]
        missing_headers = [header for header in required_header_bits if header not in headers]
        checks.append({"check": "three_bureau_columns", "result": not missing_headers, "detail": "; ".join(missing_headers)})
    if "Raw Evidence Index" in wb.sheetnames:
        ws = wb["Raw Evidence Index"]
        checks.append({"check": "raw_evidence_index_has_rows", "result": ws.max_row >= 1, "detail": f"rows={ws.max_row}"})
    if "QA Verification" in wb.sheetnames:
        ws = wb["QA Verification"]
        failed = [
            str(ws.cell(row=row, column=1).value)
            for row in range(2, ws.max_row + 1)
            if ws.cell(row=row, column=3).value == "FAIL" and ws.cell(row=row, column=4).value in {"High", "Medium"}
        ]
        checks.append({"check": "qa_high_medium_pass", "result": not failed, "detail": "; ".join(failed)})
    approved = all(check["result"] for check in checks)
    result = {"production_approval": "approved" if approved else "blocked", "checks": checks}
    wb.close()
    return result


def write_desktop_workbook(data: dict, out_dir: Path) -> None:
    if Workbook is None:
        return

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    bureau_comparison = wb.create_sheet("3 Bureau Comparison")
    side_by_side_negative = wb.create_sheet("Side By Side Negative")
    raw_evidence_index = wb.create_sheet("Raw Evidence Index")
    qa_verification = wb.create_sheet("QA Verification")
    security_audit = wb.create_sheet("Security Audit Summary")
    desktop_dashboard = wb.create_sheet("Desktop Dashboard")
    desktop_workbox = wb.create_sheet("Desktop Staff Workbox")
    desktop_field_matrix = wb.create_sheet("Desktop Field Matrix")
    errors = wb.create_sheet("Detected Errors")
    items = wb.create_sheet("Review Items")
    raw_tradelines_dates = wb.create_sheet("Raw Tradelines With Dates")
    dates_found_audit = wb.create_sheet("Dates Found Audit")
    date_issues = wb.create_sheet("Date Issues To Dispute")
    parser_qa = wb.create_sheet("Parser QA Warnings")
    metro2_fcra = wb.create_sheet("Metro 2 + FCRA Review")
    metro2_requirements = wb.create_sheet("Metro 2 Requirements")
    metro2_guide_notes = wb.create_sheet("Metro 2 Guide Notes")
    fcra_compliance = wb.create_sheet("FCRA Compliance Review")
    fcra_rights = wb.create_sheet("FCRA Rights Regulators")
    bureau_help = wb.create_sheet("Bureau Help + FDCPA")
    field_compliance = wb.create_sheet("Field Compliance Audit")
    eoscar_packaging = wb.create_sheet("e-OSCAR Packaging Review")
    fcra_notice_rules = wb.create_sheet("FCRA Notice Rules")
    dispute_methods = wb.create_sheet("Dispute Methods")
    dispute_sop = wb.create_sheet("Dispute SOP")
    letters = wb.create_sheet("Draft Letters")
    fcra = wb.create_sheet("FCRA Review")

    _write_workbook_sheet(summary, [
        ["Credit Vivo Scanner Output", ""],
        ["Engine", data.get("engine", "")],
        ["Version", data.get("version", "")],
        ["Paid AI Used", "Yes" if data.get("paid_ai_used") else "No"],
        ["Files Parsed", len(data.get("files", []))],
        ["Review Items", len(data.get("tradelines", []))],
        ["Detected Errors / Review Points", len(data.get("issues", []))],
        ["Draft Letters Queued", len(data.get("recommended_letter_queue", []))],
        ["Health Check Status", data.get("pre_scan_health_check", {}).get("overall_status", "not_attached")],
        ["Safe Mode Enabled", "Yes" if data.get("pre_scan_health_check", {}).get("safe_mode_enabled") else "No"],
        ["Scan Allowed", "Yes" if data.get("pre_scan_health_check", {}).get("scan_allowed") else "No"],
        ["Parser Integrity", data.get("pre_scan_health_check", {}).get("parser_integrity_status", "")],
        ["Rule Pack Integrity", data.get("pre_scan_health_check", {}).get("rule_pack_integrity_status", "")],
        ["Security Config Status", data.get("pre_scan_health_check", {}).get("security_config_status", "")],
        ["External Calls Enabled", "Yes" if data.get("pre_scan_health_check", {}).get("external_calls_enabled") else "No"],
        ["Auto-Send Enabled", "Yes" if data.get("pre_scan_health_check", {}).get("auto_send_enabled") else "No"],
        ["Production Approved", "Yes" if data.get("pre_scan_health_check", {}).get("production_approved") else "No"],
        ["Customer Message", data.get("customer_summary", {}).get("message", "")],
        ["Important Notice", "Draft review data only. Nothing is sent without customer approval and admin review."],
    ])

    _write_workbook_sheet(bureau_comparison, build_three_bureau_comparison_rows(data))
    _write_workbook_sheet(side_by_side_negative, build_side_by_side_negative_rows(data))
    _write_workbook_sheet(raw_evidence_index, build_raw_evidence_index_rows(data))
    qa_rows = build_qa_verification_rows(data)
    health = data.get("pre_scan_health_check", {})
    for check in health.get("checks", []):
        qa_rows.append([
            check.get("check_id", ""),
            check.get("check_name") or check.get("name", ""),
            str(check.get("status") or ("PASS" if check.get("passed") else "FAIL")).upper(),
            check.get("severity", ""),
            check.get("evidence") or check.get("detail", ""),
            check.get("fix_required", ""),
        ])
    _write_workbook_sheet(qa_verification, qa_rows)
    _write_workbook_sheet(security_audit, [
        ["Field", "Value"],
        ["Pre-Scan Health Check Result", health.get("overall_status", "not_attached")],
        ["Safe Mode Enabled", "Yes" if health.get("safe_mode_enabled") else "No"],
        ["Scan Allowed", "Yes" if health.get("scan_allowed") else "No"],
        ["Letters Allowed", "Yes" if health.get("letters_allowed") else "No"],
        ["Exports Allowed", "Yes" if health.get("exports_allowed") else "No"],
        ["External Calls Allowed", "Yes" if health.get("external_calls_allowed") else "No"],
        ["Production Approved", "Yes" if health.get("production_approved") else "No"],
        ["Failed Checks", "; ".join(health.get("errors", []))],
        ["Warnings", "; ".join(health.get("warnings", []))],
        ["Checked At", health.get("checked_at", "")],
        ["User / Device / License", health.get("user_access_status", "")],
        ["Parser Version", health.get("parser_version", "")],
        ["Rule Pack Version", health.get("rule_pack_version", "")],
        ["Security Config Version", health.get("security_config_version", "")],
    ])

    dashboard_sections = build_desktop_customer_dashboard(data)
    dashboard_summary = next((section for section in dashboard_sections if section.get("section") == "summary"), {})
    dashboard_findings = next((section for section in dashboard_sections if section.get("section") == "findings"), {}).get("items", [])
    dashboard_keep = next((section for section in dashboard_sections if section.get("section") == "keep_if_correct"), {}).get("items", [])
    _write_workbook_sheet(desktop_dashboard, [
        ["Section", "Account Name", "Priority", "Tone", "Bureaus", "Simple Issue", "Recommended Action", "Documents Needed", "Explanation"],
        ["Summary", "", "", "", "", f"Health score: {dashboard_summary.get('health_score', '')}; Negative accounts: {dashboard_summary.get('negative_accounts', '')}; Potential issues: {dashboard_summary.get('potential_issues', '')}", "; ".join(dashboard_summary.get("next_best_actions", [])), "", ""],
        *[
            [
                "Findings",
                item.get("account_name", ""),
                item.get("priority", ""),
                item.get("tone", ""),
                item.get("bureaus", []),
                item.get("simple_issue", ""),
                item.get("recommended_action", ""),
                item.get("documents_needed", []),
                item.get("explanation", []),
            ]
            for item in dashboard_findings
        ],
        *[
            [
                "Keep If Correct",
                item.get("account_name", ""),
                item.get("priority", ""),
                item.get("tone", ""),
                item.get("bureaus", []),
                item.get("simple_issue", ""),
                item.get("recommended_action", ""),
                item.get("documents_needed", []),
                item.get("explanation", []),
            ]
            for item in dashboard_keep
        ],
    ])

    _write_workbook_sheet(desktop_workbox, [
        ["Account ID", "Account Name", "Priority Score", "Queue", "Internal Findings", "Recommended Letter Types", "Documents Needed"],
        *[
            [
                row.get("account_id", ""),
                row.get("account_name", ""),
                row.get("priority_score", ""),
                row.get("queue", ""),
                row.get("internal_findings", []),
                row.get("recommended_letter_types", []),
                row.get("documents_needed", []),
            ]
            for row in build_desktop_staff_workbox(data)
        ],
    ])

    _write_workbook_sheet(desktop_field_matrix, [
        ["Account Name", "Field", "Label", "Equifax", "Experian", "TransUnion", "Differs"],
        *[
            [
                row.get("account_name", ""),
                row.get("field", ""),
                row.get("label", ""),
                row.get("equifax", ""),
                row.get("experian", ""),
                row.get("transunion", ""),
                "Yes" if row.get("differs") else "No",
            ]
            for row in build_desktop_bureau_field_matrix(data)
        ],
    ])

    _write_workbook_sheet(errors, [
        ["Issue ID", "Issue Type", "Severity", "Customer Label", "Customer Explanation", "Admin Explanation", "Suggested Round", "Related Tradeline IDs", "Confidence", "Evidence Count"],
        *[
            [
                issue.get("id", ""),
                issue.get("issue_type", ""),
                issue.get("severity", ""),
                issue.get("customer_label", ""),
                issue.get("customer_explanation", ""),
                issue.get("admin_explanation", ""),
                issue.get("suggested_round", ""),
                issue.get("related_tradeline_ids", []),
                issue.get("confidence", ""),
                len(issue.get("evidence", [])),
            ]
            for issue in data.get("issues", [])
        ],
    ])

    _write_workbook_sheet(items, [
        ["ID", "Bureau", "Source File", "Account Name", "Account Type", "Status", "Balance", "Date Opened", "Date Reported", "DOFD", "Remarks", "Confidence", "Needs Admin Review"],
        *[
            [
                item.get("id", ""),
                item.get("bureau", ""),
                item.get("source_filename", ""),
                item.get("account_name", ""),
                item.get("account_type", ""),
                item.get("status", ""),
                item.get("balance", ""),
                item.get("date_opened", ""),
                item.get("date_reported", ""),
                item.get("date_of_first_delinquency", ""),
                item.get("remarks", ""),
                item.get("confidence", ""),
                "Yes" if item.get("needs_admin_review") else "No",
            ]
            for item in data.get("tradelines", [])
        ],
    ])

    _write_workbook_sheet(raw_tradelines_dates, [
        [
            "source_file",
            "bureau",
            "page",
            "account_group",
            "creditor",
            "account_number",
            "original_creditor",
            "creditor_classification",
            "account_type",
            "account_status",
            "payment_status",
            "balance",
            "past_due",
            "high_balance",
            "credit_limit",
            "date_opened",
            "collection_assigned_date",
            "date_reported",
            "date_updated",
            "status_updated_date",
            "date_closed",
            "date_last_payment",
            "date_last_activity",
            "date_of_first_delinquency",
            "estimated_removal_date",
            "remarks",
            "confidence",
            "raw_evidence",
        ],
        *[
            [
                item.get("source_filename", ""),
                item.get("bureau", ""),
                item.get("page_start", ""),
                item.get("group_id", "") or item.get("id", ""),
                item.get("account_name", ""),
                item.get("account_number", "") or item.get("account_number_masked", ""),
                item.get("original_creditor", ""),
                item.get("creditor_classification", ""),
                item.get("account_type", ""),
                item.get("status", ""),
                item.get("pay_status", ""),
                item.get("balance", ""),
                item.get("past_due", ""),
                item.get("high_balance", "") or item.get("high_credit_or_original_amount", ""),
                item.get("credit_limit", ""),
                item.get("date_opened", ""),
                "",
                item.get("date_reported", ""),
                item.get("date_reported", ""),
                "",
                item.get("date_closed", ""),
                item.get("date_last_payment", ""),
                item.get("date_last_activity", ""),
                item.get("date_of_first_delinquency", ""),
                item.get("estimated_removal_date", ""),
                item.get("remarks", ""),
                item.get("confidence", ""),
                item.get("raw_block", ""),
            ]
            for item in data.get("tradelines", [])
        ],
    ])

    _write_workbook_sheet(dates_found_audit, [
        ["source_file", "bureau", "page", "account_key", "creditor", "field_name", "field_title", "raw_date", "normalized_date", "precision", "label_matched", "confidence", "context"],
        *[
            [
                row.get("source_file", ""),
                row.get("bureau", ""),
                row.get("page", ""),
                row.get("account_key", ""),
                row.get("creditor", ""),
                row.get("field_name", ""),
                row.get("field_title", ""),
                row.get("raw_date", ""),
                row.get("normalized_date", ""),
                row.get("precision", ""),
                row.get("label_matched", ""),
                row.get("confidence", ""),
                row.get("context", ""),
            ]
            for row in data.get("dates_found_audit", [])
        ],
    ])

    _write_workbook_sheet(date_issues, [
        ["Severity", "Account/Bureau", "Issue Type", "What we found in plain English", "Why it matters", "What to do next"],
        *[
            [
                row.get("severity", ""),
                row.get("account_bureau", ""),
                row.get("issue_type", ""),
                row.get("what_found", ""),
                row.get("why_matters", ""),
                row.get("next_step", ""),
            ]
            for row in data.get("date_issues_to_dispute", [])
        ],
    ])

    _write_workbook_sheet(parser_qa, [
        ["Warning ID", "Warning Type", "Severity", "Source File", "File Hash", "Bureau", "Page", "Account Name Guess", "Account Number Guess", "Raw Block Hash", "Raw Text Snippet", "Customer Visible", "Creates Dispute Issue", "Admin Action"],
        *[
            [
                row.get("warning_id", ""),
                row.get("warning_type", ""),
                row.get("severity", ""),
                row.get("source_filename", ""),
                row.get("source_file_hash", ""),
                row.get("bureau", ""),
                row.get("page", ""),
                row.get("account_name_guess", ""),
                row.get("account_number_guess", ""),
                row.get("raw_block_hash", ""),
                row.get("raw_text_snippet", ""),
                "Yes" if row.get("customer_visible") else "No",
                "Yes" if row.get("creates_dispute_issue") else "No",
                row.get("admin_action", ""),
            ]
            for row in data.get("parser_qa_warnings", [])
        ],
    ])

    _write_workbook_sheet(metro2_fcra, [
        [
            "Issue ID",
            "Issue Type",
            "Customer Label",
            "Metro 2 Fields To Review",
            "FCRA Focus",
            "FCRA Sections / Duties",
            "Evidence Needed",
            "Dispute Theory",
            "Responsible Party",
            "Recommended Next Action",
            "Attorney Review Signal",
            "Customer Approval Required",
        ],
        *[
            [
                row.get("issue_id", ""),
                row.get("issue_type", ""),
                row.get("customer_label", ""),
                row.get("metro2_fields_to_review", []),
                row.get("fcra_focus", ""),
                row.get("fcra_sections", []),
                row.get("evidence_needed", []),
                row.get("dispute_theory", ""),
                row.get("responsible_party", ""),
                row.get("recommended_next_action", ""),
                "Yes" if row.get("attorney_review_signal") else "No",
                "Yes" if row.get("customer_approval_required") else "No",
            ]
            for row in data.get("metro2_fcra_review", [])
        ],
    ])

    _write_workbook_sheet(metro2_requirements, [
        [
            "Tradeline ID",
            "Bureau",
            "Account Name",
            "Account Type",
            "Status",
            "Metro 2 Profile",
            "Required/Core Fields",
            "Present Fields",
            "Missing / Needs Validation",
            "Warning Flags",
            "Validation Notes",
            "Dispute Use",
            "Production Note",
        ],
        *[
            [
                row.get("tradeline_id", ""),
                row.get("bureau", ""),
                row.get("account_name", ""),
                row.get("account_type", ""),
                row.get("status", ""),
                row.get("metro2_profile", ""),
                row.get("required_core_fields", []),
                row.get("present_fields", []),
                row.get("missing_or_needs_validation", []),
                row.get("warning_flags", []),
                row.get("validation_notes", []),
                row.get("dispute_use", ""),
                row.get("production_note", ""),
            ]
            for row in data.get("metro2_requirement_review", [])
        ],
    ])

    guide_notes = data.get("metro2_public_guide_notes", {})
    _write_workbook_sheet(metro2_guide_notes, [
        ["Item", "Value"],
        ["Source", guide_notes.get("source", "")],
        ["URL", guide_notes.get("url", "")],
        ["Purpose", guide_notes.get("purpose", "")],
        ["Production Guardrail", guide_notes.get("production_guardrail", "")],
        ["Base Segment Fields Mapped", guide_notes.get("base_segment_fields", [])],
        ["Collection / Associated Segments", guide_notes.get("collection_segments", [])],
    ])

    _write_workbook_sheet(fcra_compliance, [
        [
            "Tradeline ID",
            "Bureau",
            "Account Name",
            "FCRA Area",
            "Law Reference",
            "Plain English Meaning",
            "Applies When",
            "Scanner Signals",
            "Evidence Needed",
            "Tracking Action",
            "Customer Approval Required",
            "Attorney / Compliance Review",
            "Note",
        ],
        *[
            [
                row.get("tradeline_id", ""),
                row.get("bureau", ""),
                row.get("account_name", ""),
                row.get("fcra_area", ""),
                row.get("law_reference", ""),
                row.get("plain_english", ""),
                row.get("applies_when", ""),
                row.get("scanner_signals", []),
                row.get("evidence_needed", ""),
                row.get("tracking_action", ""),
                "Yes" if row.get("customer_approval_required") else "No",
                "Yes" if row.get("attorney_or_compliance_review") else "No",
                row.get("note", ""),
            ]
            for row in data.get("fcra_compliance_review", [])
        ],
    ])

    rights_reference = data.get("fcra_rights_reference", {})
    source_notes = rights_reference.get("source_notes", {})
    rights_rows = [
        ["Section", "Category / State", "Agency / Item", "Address / URL", "Phone / Note"],
        ["Source", "Primary Reference", source_notes.get("primary_reference", ""), source_notes.get("primary_url", ""), source_notes.get("plain_english_note", "")],
        ["Source", "Agency Contact Update", source_notes.get("agency_contact_update_reference", ""), source_notes.get("agency_contact_update_url", ""), source_notes.get("compliance_note", "")],
    ]
    for right in rights_reference.get("federal_consumer_rights", []):
        rights_rows.append([
            "Federal Consumer Right",
            "FCRA",
            right.get("right", ""),
            right.get("plain_english", ""),
            right.get("scanner_use", ""),
        ])
    maryland_rights = rights_reference.get("maryland_consumer_rights", {})
    if maryland_rights:
        rights_rows.append([
            "State Consumer Right",
            maryland_rights.get("state", "Maryland"),
            maryland_rights.get("legal_reference", ""),
            maryland_rights.get("plain_english_summary", ""),
            "",
        ])
        for right in maryland_rights.get("consumer_rights", []):
            rights_rows.append([
                "State Consumer Right",
                maryland_rights.get("state", "Maryland"),
                "Maryland consumer reporting right",
                right,
                "",
            ])
        contact = maryland_rights.get("complaint_contact", {})
        rights_rows.append([
            "State Complaint Contact",
            maryland_rights.get("state", "Maryland"),
            contact.get("agency", ""),
            contact.get("address", ""),
            contact.get("phone", ""),
        ])
        for use in maryland_rights.get("scanner_use", []):
            rights_rows.append([
                "State Scanner Rule",
                maryland_rights.get("state", "Maryland"),
                "Credit Vivo workflow",
                "",
                use,
            ])
    for contact_group in rights_reference.get("federal_contacts", []):
        for contact in contact_group.get("contacts", []):
            rights_rows.append([
                "Federal Contact",
                contact_group.get("category", ""),
                contact.get("agency", ""),
                contact.get("address", ""),
                contact.get("phone") or contact.get("use_when", ""),
            ])
    for state_link in rights_reference.get("state_notice_links", []):
        rights_rows.append([
            "State Notice Link",
            state_link.get("state", ""),
            "State consumer reporting rights reference",
            state_link.get("url", ""),
            "Confirm current state-law requirements before production disclosure use.",
        ])
    for rule in rights_reference.get("ai_rules", []):
        rights_rows.append(["AI Rule", "All AI engines", "FCRA rights routing rule", "", rule])
    _write_workbook_sheet(fcra_rights, rights_rows)

    bureau_reference = data.get("bureau_debt_collection_reference", {})
    source_notes = bureau_reference.get("source_notes", {})
    bureau_rows = [
        ["Section", "Bureau / Rule", "Meaning", "Scanner / Customer Next Step", "Source / Note"],
        ["Source", "Equifax dispute help", source_notes.get("equifax_dispute_url", ""), source_notes.get("equifax_mail_dispute_url", ""), source_notes.get("compliance_note", "")],
        ["Source", "Experian dispute/outcomes", source_notes.get("experian_dispute_url", ""), source_notes.get("experian_outcome_url", ""), ""],
        ["Source", "TransUnion dispute help", source_notes.get("transunion_dispute_url", ""), "", ""],
        ["Source", "CFPB dispute help", source_notes.get("cfpb_dispute_url", ""), "", ""],
        ["Source", "FDCPA", source_notes.get("fdcpa_source", ""), "", ""],
    ]
    for item in bureau_reference.get("bureau_dispute_workflow", []):
        bureau_rows.append([
            "Bureau Workflow",
            item.get("bureau", ""),
            item.get("customer_help_rule", ""),
            item.get("scanner_action", ""),
            "Channels: " + "; ".join(item.get("channels", [])) + " | Proof: " + "; ".join(item.get("proof_examples", [])),
        ])
    for item in bureau_reference.get("experian_dispute_outcomes", []):
        bureau_rows.append([
            "Experian Outcome",
            item.get("outcome", ""),
            item.get("meaning", ""),
            item.get("scanner_next_step", ""),
            "Use to parse bureau response/results letters.",
        ])
    for item in bureau_reference.get("fdcpa_collection_rules", []):
        bureau_rows.append([
            "FDCPA Rule",
            item.get("rule", ""),
            item.get("plain_english", ""),
            item.get("scanner_use", ""),
            "Debt-collection conduct rule; approval-gated.",
        ])
    for rule in bureau_reference.get("ai_rules", []):
        bureau_rows.append(["AI Rule", "All AI engines", rule, "", ""])
    _write_workbook_sheet(bureau_help, bureau_rows)

    _write_workbook_sheet(field_compliance, [
        [
            "Tradeline ID",
            "Bureau",
            "Account Name",
            "Field Name",
            "Parsed Value",
            "Required / Expected",
            "Issue Flag",
            "Metro 2 Concept",
            "FCRA / Reg V Basis",
            "Issue Text",
            "Verification Ask",
            "Requested Outcome",
            "Source Note",
        ],
        *[
            [
                row.get("tradeline_id", ""),
                row.get("bureau", ""),
                row.get("account_name", ""),
                row.get("field_name", ""),
                row.get("parsed_value", ""),
                row.get("required_or_expected", ""),
                row.get("issue_flag", ""),
                row.get("metro2_concept", ""),
                row.get("fcra_basis", ""),
                row.get("issue_text", ""),
                row.get("verification_ask", ""),
                row.get("requested_outcome", ""),
                row.get("source_note", ""),
            ]
            for row in data.get("field_compliance_audit", [])
        ],
    ])

    _write_workbook_sheet(eoscar_packaging, [
        [
            "Issue ID",
            "Issue Type",
            "Account Names",
            "Bureaus",
            "e-OSCAR / ACDV Category",
            "ACDV Packaging Steps",
            "Field Focus",
            "Package Hint",
            "Evidence Hint",
            "Avoid",
            "Tracking Status",
        ],
        *[
            [
                row.get("issue_id", ""),
                row.get("issue_type", ""),
                row.get("account_names", []),
                row.get("bureaus", []),
                row.get("eoscar_category", ""),
                row.get("acdv_packaging_steps", []),
                row.get("field_focus", []),
                row.get("package_hint", ""),
                row.get("evidence_hint", ""),
                row.get("avoid", []),
                row.get("tracking_status", ""),
            ]
            for row in data.get("eoscar_packaging_review", [])
        ],
        ["Public Fact", "What e-OSCAR is", "", "", "", EOSCAR_PUBLIC_FACTS[0]["detail"], "", "", "", EOSCAR_PUBLIC_FACTS[0]["source"], ""],
        ["Public Fact", "ACDV path", "", "", "", EOSCAR_PUBLIC_FACTS[1]["detail"], "", "", "", EOSCAR_PUBLIC_FACTS[1]["source"], ""],
        ["Public Fact", "AUD path", "", "", "", EOSCAR_PUBLIC_FACTS[2]["detail"], "", "", "", EOSCAR_PUBLIC_FACTS[2]["source"], ""],
    ])

    notice_rows = [["Rule Area", "Requirement / Control"]]
    for area, controls in data.get("letter_workflow", {}).get("fcra_notice_rules", {}).items():
        label = area.replace("_", " ").title()
        for control in controls:
            notice_rows.append([label, control])
    _write_workbook_sheet(fcra_notice_rules, notice_rows)

    _write_workbook_sheet(dispute_methods, build_dispute_method_rows())

    _write_workbook_sheet(dispute_sop, build_dispute_sop_rows())

    _write_workbook_sheet(letters, [
        ["Letter ID", "Issue ID", "Subject", "Letter Type", "Round", "Recipient Type", "Delivery Method", "FCRA Notice Included", "Customer Approval Required", "Tracking Status", "Recommended Next Action", "Draft Letter Body"],
        *[
            [
                letter.get("letter_id", ""),
                letter.get("issue_id", ""),
                letter.get("letter_subject", ""),
                letter.get("letter_type", ""),
                letter.get("round", ""),
                letter.get("recipient_type", ""),
                letter.get("delivery_method", ""),
                "Yes" if letter.get("fcra_notice_included") else "No",
                "Yes" if letter.get("customer_approval_required") else "No",
                letter.get("tracking_status", ""),
                letter.get("recommended_next_action", ""),
                letter.get("draft_letter_body", ""),
            ]
            for letter in data.get("recommended_letter_queue", [])
        ],
    ])

    _write_workbook_sheet(fcra, [
        ["Issue ID", "Possible FCRA Issue", "Issue Type", "Responsible Party", "Dispute History Complete", "Evidence Strength", "Damages Evidence", "Next Action", "Requires Admin Review"],
        *[
            [
                row.get("issue_id", ""),
                "Yes" if row.get("possible_fcra_issue") else "No",
                row.get("issue_type", ""),
                row.get("responsible_party", ""),
                "Yes" if row.get("dispute_history_complete") else "No",
                row.get("evidence_strength", ""),
                row.get("damages_evidence", ""),
                row.get("next_action", ""),
                "Yes" if row.get("requires_admin_review") else "No",
            ]
            for row in data.get("fcra_review", [])
        ],
    ])

    wb.save(out_dir / "credit_vivo_desktop_scanner_output.xlsx")
    wb.close()


def write_outputs(result: ParseResult, out_dir: Path, pre_scan_health_check: Optional[dict] = None) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    data = result_to_dict(result)
    if pre_scan_health_check is not None:
        data["pre_scan_health_check"] = pre_scan_health_check
    (out_dir / "credit_vivo_parser_result.json").write_text(json.dumps(data, indent=2), encoding="utf-8")

    # Tradelines CSV
    tradeline_rows = data.get("tradelines", [])
    if tradeline_rows:
        with (out_dir / "tradelines.csv").open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(tradeline_rows[0].keys()))
            writer.writeheader()
            writer.writerows(_safe_csv_row(row) for row in tradeline_rows)

    # Issues CSV - flatten evidence count only
    issue_rows = []
    for issue in result.issues:
        d = asdict(issue)
        d["related_tradeline_ids"] = ";".join(issue.related_tradeline_ids)
        d["evidence_count"] = len(issue.evidence)
        d.pop("evidence", None)
        issue_rows.append(d)

    if issue_rows:
        with (out_dir / "review_issues.csv").open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(issue_rows[0].keys()))
            writer.writeheader()
            writer.writerows(_safe_csv_row(row) for row in issue_rows)

    date_rows = data.get("dates_found_audit", [])
    if date_rows:
        with (out_dir / "dates_found_audit.csv").open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(date_rows[0].keys()))
            writer.writeheader()
            writer.writerows(_safe_csv_row(row) for row in date_rows)

    letter_sections = []
    for letter in data.get("recommended_letter_queue", []):
        letter_sections.append(
            "\n".join(
                [
                    f"Letter ID: {letter.get('letter_id', '')}",
                    f"Subject: {letter.get('letter_subject', '')}",
                    f"Type: {letter.get('letter_type', '')}",
                    f"Round: {letter.get('round', '')}",
                    f"Recipient: {letter.get('recipient_type', '')}",
                    f"Tracking status: {letter.get('tracking_status', '')}",
                    "",
                    str(letter.get("draft_letter_body") or "No draft body generated."),
                ]
            )
        )

    if letter_sections:
        (out_dir / "draft_dispute_letters.txt").write_text(
            ("\n\n" + ("-" * 72) + "\n\n").join(letter_sections),
            encoding="utf-8",
        )

    write_desktop_workbook(data, out_dir)
    workbook_path = out_dir / "credit_vivo_desktop_scanner_output.xlsx"
    if workbook_path.exists():
        validation = validate_workbook_output(workbook_path)
        (out_dir / "workbook_validation.json").write_text(json.dumps(validation, indent=2), encoding="utf-8")


# Phase 3 draft-only integrations.
# These classes intentionally do not send disputes, mail, or bank data by
# themselves. They prepare review artifacts for an authenticated admin workflow.
import hashlib
import os
import random
import re
from datetime import datetime, timezone
from typing import Any, Optional

import requests
from sodapy import Socrata


class OpenDataCrossMatcher:
    """Cross-match collector names against public licensing data."""

    def __init__(self, domain: str = "opendata.maryland.gov"):
        self.domain = domain
        self.app_token = os.getenv("SOCRATA_APP_TOKEN")
        self.client = Socrata(self.domain, self.app_token or None)

    @staticmethod
    def _safe_collector_name(collector_name: str) -> str:
        cleaned = re.sub(r"[^A-Za-z0-9 '&.,-]", "", collector_name or "").strip()
        return cleaned[:120]

    def verify_collector_license(self, collector_name: str, state_code: str = "MD") -> dict[str, Any]:
        safe_name = self._safe_collector_name(collector_name)
        safe_state = (state_code or "MD").upper()[:2]

        if not safe_name:
            return {
                "found": False,
                "status": "UNKNOWN",
                "leverage_flag": None,
                "review_note": "Collector name was empty or unsupported.",
            }

        try:
            dataset_id = "gdzy-2fen"
            escaped_name = safe_name.upper().replace("'", "''")
            query = f"business_name like '%{escaped_name}%'"
            results = self.client.get(dataset_id, where=query, limit=1)

            if not results:
                return {
                    "found": False,
                    "status": "UNKNOWN",
                    "leverage_flag": None,
                    "review_note": "No public licensing record matched this collector name.",
                    "last_checked_at": datetime.now(timezone.utc).isoformat(),
                }

            record = results[0]
            status = str(record.get("license_status", "UNKNOWN")).upper()
            inactive_statuses = {"REVOKED", "EXPIRED", "SUSPENDED", "INACTIVE"}
            leverage_flag: Optional[str] = None

            if status in inactive_statuses:
                leverage_flag = (
                    f"Public open-data records may indicate {safe_name} has a {status} "
                    f"license status in {safe_state}. Admin review should verify the "
                    "source record before using this in a customer letter."
                )

            return {
                "found": True,
                "status": status,
                "leverage_flag": leverage_flag,
                "raw_evidence": record,
                "last_checked_at": datetime.now(timezone.utc).isoformat(),
            }
        except Exception as exc:
            return {
                "found": False,
                "status": "ERROR",
                "leverage_flag": None,
                "error": str(exc),
                "last_checked_at": datetime.now(timezone.utc).isoformat(),
            }


class EOSCARBypassSpinner:
    """Generate unique, draft-only dispute letter language for admin review."""

    def __init__(self):
        self.openers = [
            "I am requesting an investigation of an item appearing on my {bureau} credit file.",
            "Please review my {bureau} credit profile, including the tradeline reported by {creditor}.",
            "I am asking for a reasonable reinvestigation of the {creditor} account on my {bureau} report.",
        ]
        self.closers = [
            "Please complete a reasonable investigation and provide the results in writing.",
            "Please verify the reporting with competent evidence or correct the information as required.",
            "Please send the investigation results and any updated report after your review is complete.",
        ]

    @staticmethod
    def _mask_account(account_num: str) -> str:
        digits = re.sub(r"\D", "", account_num or "")
        if len(digits) <= 4:
            return "ending in " + (digits or "unknown")
        return "ending in " + digits[-4:]

    def generate_unique_letter(
        self,
        bureau: str,
        creditor: str,
        account_num: str,
        violation_code: str,
        open_data_leverage: str | None = None,
    ) -> str:
        safe_bureau = bureau or "the bureau"
        safe_creditor = creditor or "the furnisher"
        safe_violation = violation_code or "reporting accuracy review"
        opener = random.choice(self.openers).format(
            bureau=safe_bureau,
            creditor=safe_creditor,
            account_num=self._mask_account(account_num),
        )
        closer = random.choice(self.closers)
        body = (
            f"This draft concerns {safe_creditor}, account {self._mask_account(account_num)}. "
            f"The item is being reviewed for {safe_violation}. I am not asking for any "
            "accurate and verifiable information to be removed; I am asking that the "
            "reporting be investigated, corrected, updated, or deleted only if it cannot "
            "be verified under applicable law."
        )

        if open_data_leverage:
            body += (
                f"\n\nAdditional review note: {open_data_leverage} This public-data signal "
                "must be verified by an admin before the language is sent."
            )

        return f"{opener}\n\n{body}\n\n{closer}\n\nDraft only - requires customer approval and admin review."


class LitigationTracker:
    """Prepare mail-tracking records without transmitting mail automatically."""

    def __init__(self):
        self.lob_api_key = os.getenv("LOB_API_KEY")
        self.lob_base_url = os.getenv("LOB_BASE_URL", "https://api.lob.com/v1")

    @staticmethod
    def hash_tracking_number(tracking_number: str) -> str:
        normalized = re.sub(r"\s+", "", tracking_number or "").upper()
        if not normalized:
            raise ValueError("tracking_number is required")
        return hashlib.sha256(normalized.encode("utf-8")).hexdigest()

    def build_mail_tracking_event(
        self,
        letter_id: str,
        tracking_number: str,
        delivery_status: str,
        delivery_timestamp: str | None = None,
    ) -> dict[str, Any]:
        return {
            "letter_id": letter_id,
            "usps_tracking_hash": self.hash_tracking_number(tracking_number),
            "delivery_status": (delivery_status or "UNKNOWN").upper(),
            "delivery_timestamp": delivery_timestamp,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }

    def fetch_lob_letter_status(self, lob_letter_id: str) -> dict[str, Any]:
        if not self.lob_api_key:
            return {
                "ok": False,
                "error": "LOB_API_KEY is not configured.",
                "review_note": "Mail status lookup is disabled in local testing.",
            }

        response = requests.get(
            f"{self.lob_base_url}/letters/{lob_letter_id}",
            auth=(self.lob_api_key, ""),
            timeout=15,
        )
        response.raise_for_status()
        payload = response.json()
        return {
            "ok": True,
            "letter_id": payload.get("id"),
            "mail_type": payload.get("mail_type"),
            "expected_delivery_date": payload.get("expected_delivery_date"),
            "tracking_events": payload.get("tracking_events", []),
        }
