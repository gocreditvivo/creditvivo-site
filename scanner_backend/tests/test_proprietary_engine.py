from credit_vivo_proprietary_engine import build_three_bureau_comparison_rows, is_bad_account_name, parse_reports, result_to_dict, validate_workbook_output, write_outputs
from openpyxl import load_workbook

SAMPLE = """
--- PAGE 1 ---
Experian Credit Report

MIDLAND CREDIT MANAGEMENT
Account Number: 1234567890
Account Type: Collection
Original Creditor: CAPITAL ONE BANK
Balance: $1,234
Status: Collection
Date Opened: 01/10/2021
Date Reported: 04/01/2026
Remarks: Account placed for collection

--- PAGE 2 ---
CAPITAL ONE
Account Number: 999988881111
Account Type: Credit Card
Balance: $0
Status: Charge-off transferred or sold
Date Opened: 05/01/2019
Date of First Delinquency: 10/01/2020
"""

SAMPLE_EQUIFAX = """
--- PAGE 1 ---
Equifax Credit Report

MIDLAND CREDIT MANAGEMENT
Account Number: 1234567890
Account Type: Collection
Original Creditor: CAPITAL ONE BANK
Balance: $1,999
Status: Collection Account
Date Opened: 01/10/2021
Date Reported: 05/01/2026
Date of First Delinquency: 02/01/2021
Remarks: Account placed for collection
"""

SAMPLE_TRANSUNION = """
--- PAGE 1 ---
TransUnion Credit Report

MIDLAND CREDIT MANAGEMENT
Account Number: 1234567890
Account Type: Collection
Original Creditor: CAPITAL ONE BANK
Balance: $1,234
Status: Paid collection
Date Opened: 01/10/2021
Date Reported: 04/01/2026
Date of First Delinquency: 01/01/2021
Remarks: Account placed for collection
"""

PAGE_LEVEL_POSITIVE_SAMPLE = """
--- PAGE 1 ---
Equifax Credit Report

CREDIT ONE BANK
Account Number: *1664
Account Type: Credit Card
Balance: $59
Status: Pays As Agreed
Date Opened: 03/09/2026
Date Reported: 06/19/2026

--- PAGE 2 ---
Experian Credit Report

CREDIT ONE BANK
Account Number: *4796
Account Type: Credit Card
Balance: $125
Status: Open
Date Opened: 03/09/2026
Date Reported: 06/11/2026
"""

POSSIBLE_SKIPPED_TRADELINE_SAMPLE = """
--- PAGE 1 ---
Equifax Credit Report

Account Number: *9999
Account Type: Credit Card
Balance: $44
Status: Open
Date Opened: 01/01/2024
Date Reported: 06/01/2026
"""

BOILERPLATE_SAMPLE = """
--- PAGE 1 ---
Experian Credit Report

For more information, go to www.consumerfinance.gov/learnmore for additional information.
Supplying information to a consumer reporting agency violates the FCRA if you know or have reasonable cause to believe the information is inaccurate.
This section represents the last reported status of the account.
Account types is good for your credit.
The company must give you the name, address, and phone number of the furnisher.
"""

REALISTIC_EQUIFAX = """
--- PAGE 1 ---
Equifax Credit Report

CREDIT ONE BANK
PO Box 98873, Las Vegas, NV 89193-8873 | (877) 825-3242 Date Reported: 03/11/2026 | Balance: $488
Account Number: *6902 | Owner: Individual Account Credit Limit: $500 | High Credit: $696
Loan/Account Type: Credit Card | Status: Pays As Agreed

Date Opened: 12/27/2022 Date of 1st Delinquency: Terms Frequency: Monthly
Date of Last Activity: 03/11/2026 Date Major Delinquency 1st Reported: Months Reviewed: 38
Scheduled Payment Amount: $30 Amount Past Due: $60
Actual Payment Amount: Charge Off Amount: Balloon Payment Amount:
Date of Last Payment: 02/11/2026 Date Closed: Balloon Payment Date:
"""

REALISTIC_EXPERIAN = """
--- PAGE 1 ---
Experian Credit Report

Account information disputed by consumer (Meets requirement
of the Fair Credit Reporting Act).
CREDIT ONE BANK
POTENTIALLY NEGATIVE
Account Info
Account Name CREDIT ONE BANK
Account Number 444796XXXXXXXXXX
Account Type Credit card
Responsibility Individual
Date Opened 12/27/2022
Status Open.
Status Updated Sep 2023
Balance $488
Balance Updated 03/11/2026
Recent Payment -
Monthly Payment $30
Credit Limit $500
Highest Balance $696
Terms -
Payment History
30 days past due as of Aug 2023
By Apr 2030, this account is scheduled to go to a positive status.
"""

REALISTIC_TRANSUNION = """
--- PAGE 1 ---
TransUnion Credit Report

CREDIT ONE BANK
444796268171****
Account Information
Address PO BOX 98872 LAS VEGAS, NV 89193-8872
Monthly Payment $30
Date Opened 12/27/2022
Responsibility Individual Account
Account Type Revolving Account
Loan Type CREDIT CARD
Balance $488
Date Updated 03/11/2026
Payment Received $0
Last Payment Made 02/11/2026
Pay Status Current Account
Terms $30 per month; paid Monthly
High Balance (Hist.)
High balance of $696 from 10/2023 to 03/2026
Credit Limit (Hist.)
Credit limit of $500 from 10/2023 to 03/2026
"""

EQUIFAX_HELP_SECTION_THAT_IS_NOT_AN_ACCOUNT = """
--- PAGE 1 ---
Equifax Credit Report

Your credit report can include fraud and active duty alerts, security freezes or locks,
and opt-outs of receiving prescreened offers of credit or insurance. You can also choose
to add a consumer statement to your credit report to provide an explanation for why you
missed a payment or why you believe something is being reported incorrectly. For more
information on these consumer added notices visit equifax.com/personal/help/.

Account Information
Your credit report can include all types of credit accounts, such as revolving accounts,
mortgage accounts, and any other installment loans or open lines of credit. When reviewing
your account information, see if an account is open, closed, paid on time, or past due.
Closed accounts should have no money due.

MIDLAND CREDIT MANAGEMENT - Closed
320 E BIG BEAVER RD STE 300, TROY, MI 48083-1271 | (877) 822-0381
Date Reported: 06/25/2026 | Balance: $1,473
Account Number: *8933 | Owner: Individual
Loan/Account Type: Debt Buyer Account | Status: Collection
Date Opened: 02/17/2022
Date of 1st Delinquency: 07/07/2021
Amount Past Due: $1,473
"""

EQUIFAX_DUPLICATE_ACCOUNT_WINDOWS = """
--- PAGE 1 ---
Equifax Credit Report

MIDLAND CREDIT MANAGEMENT - Closed
320 E BIG BEAVER RD STE 300, TROY, MI 48083-1271 | (877) 822-0381
Date Reported: 06/25/2026 | Balance: $1,473
Account Number: *8933 | Owner: Individual Account Credit Limit: | High Credit: $1,127
Loan/Account Type: Debt Buyer Account | Status: Collection
Date Opened: 02/17/2022 Date of 1st Delinquency: 07/07/2021
Amount Past Due: $1,473

Loan/Account Type: Debt Buyer Account | Status:
Date Opened: 02/17/2022 Date of 1st Delinquency: 07/07/2021
Amount Past Due: $1,473

Prepared for: TIM K DO Date: June 29, 2026 Confirmation # 6180577433
MIDLAND CREDIT MANAGEMENT - Closed
Date Reported: 06/25/2026 | Balance: $1,473
Account Number: *8933
"""

def test_parse_sample_report(tmp_path):
    result = parse_reports({
        "experian.pdf": {"text": SAMPLE, "bureau": "Experian"},
        "equifax.pdf": {"text": SAMPLE_EQUIFAX, "bureau": "Equifax"},
        "transunion.pdf": {"text": SAMPLE_TRANSUNION, "bureau": "TransUnion"},
    })
    data = result_to_dict(result)
    assert data["paid_ai_used"] is False
    assert data["tradelines"]
    assert data["files"][0]["source_file_hash"]
    assert all(item["source_file_hash"] for item in data["tradelines"])
    assert all(item["raw_block_id"] for item in data["tradelines"])
    assert all(item["raw_block_hash"] for item in data["tradelines"])
    for item in data["tradelines"]:
        evidence = item["field_evidence"]
        assert evidence
        for field_name in ["account_name", "account_number_masked", "account_type", "balance"]:
            if item.get(field_name):
                assert field_name in evidence
                assert evidence[field_name]["source_file_hash"] == item["source_file_hash"]
                assert evidence[field_name]["page"] == item["page_start"]
                assert evidence[field_name]["bureau"] == item["bureau"]
                assert evidence[field_name]["raw_block_id"] == item["raw_block_id"]
                assert evidence[field_name]["raw_block_hash"] == item["raw_block_hash"]
                assert evidence[field_name]["raw_line"]
                assert evidence[field_name]["extraction_rule_id"]
    assert data["issues"]
    labels = {x["customer_label"] for x in data["issues"]}
    assert "Collection review" in labels or "Charge-off review" in labels
    assert data["letter_workflow"]["send_letters_automatically"] is False
    assert "FCRA" in data["letter_workflow"]["fcra_notice_of_dispute"]
    assert "formal notice of dispute" in data["letter_workflow"]["fcra_notice_of_dispute"]
    assert "written results" in data["letter_workflow"]["fcra_notice_of_dispute"]
    assert data["letter_workflow"]["fcra_notice_rules"]["consumer_notice_contents"]
    assert data["letter_workflow"]["fcra_notice_rules"]["bureau_dispute_rules"]
    assert data["letter_workflow"]["fcra_notice_rules"]["furnisher_dispute_rules"]
    assert data["letter_workflow"]["fcra_notice_rules"]["credit_vivo_controls"]
    assert data["recommended_letter_queue"]
    assert data["recommended_letter_queue"][0]["tracking_status"] == "draft_not_sent"
    assert data["recommended_letter_queue"][0]["customer_approval_required"] is True
    assert data["recommended_letter_queue"][0]["letter_subject"]
    assert "DRAFT" in data["recommended_letter_queue"][0]["draft_letter_body"]
    assert "CUSTOMER REVIEW AND APPROVAL REQUIRED" in data["recommended_letter_queue"][0]["draft_letter_body"]
    assert "FCRA" in data["recommended_letter_queue"][0]["draft_letter_body"]
    letter_types = {row["letter_type"] for row in data["recommended_letter_queue"]}
    assert "debt_validation_request" in letter_types
    debt_validation = next(row for row in data["recommended_letter_queue"] if row["letter_type"] == "debt_validation_request")
    assert debt_validation["fdcpa_validation_request"] is True
    assert "Debt Validation Request" in debt_validation["letter_subject"]
    assert "original creditor" in debt_validation["draft_letter_body"].lower()
    assert "itemized" in debt_validation["draft_letter_body"].lower()
    assert "authority to collect" in debt_validation["draft_letter_body"].lower()
    assert data["fcra_review"]
    assert data["fcra_review"][0]["dispute_history_complete"] is False
    assert data["metro2_fcra_review"]
    metro2_fields = {
        field
        for row in data["metro2_fcra_review"]
        for field in row["metro2_fields_to_review"]
    }
    assert "Current Balance" in metro2_fields
    assert "Date of First Delinquency" in metro2_fields
    assert any("FCRA 611" in section for row in data["metro2_fcra_review"] for section in row["fcra_sections"])
    assert data["metro2_requirement_review"]
    requirement_profiles = {row["metro2_profile"] for row in data["metro2_requirement_review"]}
    assert "Collection account" in requirement_profiles
    assert "Charge-off account" in requirement_profiles or "Closed, sold, transferred, or assigned account" in requirement_profiles
    requirement_text = " ".join(
        " ".join(row["required_core_fields"]) + " " +
        " ".join(row["missing_or_needs_validation"]) + " " +
        " ".join(row["warning_flags"])
        for row in data["metro2_requirement_review"]
    )
    assert "Date of First Delinquency" in requirement_text
    assert "Original Creditor" in requirement_text
    assert "Validate against official licensed CDIA Metro 2 CRRG" in " ".join(
        row["production_note"] for row in data["metro2_requirement_review"]
    )
    assert data["fcra_compliance_review"]
    fcra_areas = {row["fcra_area"] for row in data["fcra_compliance_review"]}
    assert "FCRA accuracy and completeness" in fcra_areas
    assert "Bureau reinvestigation duty" in fcra_areas
    assert "Direct furnisher dispute" in fcra_areas
    assert "DOFD and obsolete information review" in fcra_areas
    assert "Disputed-account notation" in fcra_areas
    fcra_text = " ".join(
        row["law_reference"] + " " +
        row["plain_english"] + " " +
        " ".join(row["scanner_signals"]) + " " +
        row["tracking_action"]
        for row in data["fcra_compliance_review"]
    )
    assert "FCRA 611" in fcra_text
    assert "FCRA 623" in fcra_text
    assert "Regulation V 12 CFR 1022.43" in fcra_text
    assert "dispute notation" in fcra_text
    assert data["field_compliance_audit"]
    field_audit_text = " ".join(
        row["field_name"] + " " +
        row["required_or_expected"] + " " +
        row["issue_flag"] + " " +
        row["metro2_concept"] + " " +
        row["fcra_basis"] + " " +
        row["verification_ask"]
        for row in data["field_compliance_audit"]
    )
    assert "Date of First Delinquency" in field_audit_text
    assert "Original Creditor" in field_audit_text
    assert "FCRA 623(a)(5)" in field_audit_text
    assert "Reg V direct dispute" in field_audit_text or "Regulation V" in field_audit_text
    assert data["eoscar_packaging_review"]
    eoscar_text = " ".join(
        row["eoscar_category"] + " " +
        " ".join(row["acdv_packaging_steps"]) + " " +
        " ".join(row["field_focus"]) + " " +
        row["package_hint"] + " " +
        row["evidence_hint"]
        for row in data["eoscar_packaging_review"]
    )
    assert "ACDV" in eoscar_text
    assert "Date of First Delinquency" in eoscar_text
    assert "specific field" in eoscar_text or "field-specific" in eoscar_text
    assert data["eoscar_public_facts"]
    assert any("AUD" in fact["title"] or "AUD" in fact["detail"] for fact in data["eoscar_public_facts"])
    assert data["dates_found_audit"]
    date_audit_text = " ".join(
        row["field_title"] + " " +
        row["raw_date"] + " " +
        row["normalized_date"] + " " +
        row["context"]
        for row in data["dates_found_audit"]
    )
    assert "Date Opened" in date_audit_text
    assert "Date of First Delinquency" in date_audit_text
    assert data["date_issues_to_dispute"]
    date_issue_text = " ".join(
        row["severity"] + " " +
        row["issue_type"] + " " +
        row["what_found"] + " " +
        row["next_step"]
        for row in data["date_issues_to_dispute"]
    )
    assert "Date differs across bureaus" in date_issue_text or "Missing date field" in date_issue_text
    cross_labels = {x["customer_label"] for x in data["issues"] if x["issue_type"].startswith("cross_bureau")}
    assert "Balance differs across bureaus" in cross_labels
    assert "Status differs across bureaus" in cross_labels
    assert "Dates differ across bureaus" in cross_labels

    write_outputs(result, tmp_path)
    assert (tmp_path / "credit_vivo_parser_result.json").exists()
    assert (tmp_path / "tradelines.csv").exists()
    assert (tmp_path / "review_issues.csv").exists()
    assert (tmp_path / "dates_found_audit.csv").exists()
    workbook_path = tmp_path / "credit_vivo_desktop_scanner_output.xlsx"
    assert workbook_path.exists()
    validation_path = tmp_path / "workbook_validation.json"
    assert validation_path.exists()
    validation = validate_workbook_output(workbook_path)
    assert validation["production_approval"] in {"approved", "blocked"}
    assert any(check["check"] == "required_sheets" and check["result"] for check in validation["checks"])
    assert any(check["check"] == "no_duplicate_headers" and check["result"] for check in validation["checks"])
    workbook = load_workbook(workbook_path, read_only=True)
    assert workbook.sheetnames == [
        "Summary",
        "3 Bureau Comparison",
        "Side By Side Negative",
        "Raw Evidence Index",
        "QA Verification",
        "Security Audit Summary",
        "Desktop Dashboard",
        "Desktop Staff Workbox",
        "Desktop Field Matrix",
        "Detected Errors",
        "Review Items",
        "Raw Tradelines With Dates",
        "Dates Found Audit",
        "Date Issues To Dispute",
        "Parser QA Warnings",
        "Metro 2 + FCRA Review",
        "Metro 2 Requirements",
        "Metro 2 Guide Notes",
        "FCRA Compliance Review",
        "FCRA Rights Regulators",
        "Bureau Help + FDCPA",
        "Field Compliance Audit",
        "e-OSCAR Packaging Review",
        "FCRA Notice Rules",
        "Dispute Methods",
        "Dispute SOP",
        "Draft Letters",
        "FCRA Review",
    ]
    comparison = workbook["3 Bureau Comparison"]
    headers = [comparison.cell(row=1, column=column).value for column in range(1, comparison.max_column + 1)]
    assert headers[:3] == [
        "Account Name",
        "Account #",
        "Account Type",
    ]
    assert len(headers) == len(set(headers))
    assert headers[-5:] == [
        "Matched Bureaus",
        "Missing Bureaus",
        "Errors / Findings",
        "Recommended Action",
        "Group ID",
    ]
    assert "Equifax Account #" in headers
    assert "Equifax Balance" in headers
    assert "Experian Account #" in headers
    assert "Experian Balance" in headers
    assert "TransUnion Account #" in headers
    assert "TransUnion Balance" in headers
    assert "Equifax Raw Evidence" not in headers
    assert "Experian Raw Evidence" not in headers
    assert "Bureau Dispute Draft" not in headers
    assert "Furnisher Dispute Draft" not in headers
    assert "Debt Validation Draft" not in headers

    equifax_account_column = headers.index("Equifax Account #") + 1
    experian_account_column = headers.index("Experian Account #") + 1
    transunion_account_column = headers.index("TransUnion Account #") + 1
    equifax_balance_column = headers.index("Equifax Balance") + 1
    errors_column = headers.index("Errors / Findings") + 1
    comparison_flags = " ".join(
        str(comparison.cell(row=row, column=errors_column).value or "")
        for row in range(2, comparison.max_row + 1)
    )
    assert "Balance differs" in comparison_flags
    assert "Status differs" in comparison_flags
    assert "DOFD differs" in comparison_flags
    assert any(
        comparison.cell(row=row, column=equifax_account_column).value
        for row in range(2, comparison.max_row + 1)
    )
    assert any(
        comparison.cell(row=row, column=experian_account_column).value
        for row in range(2, comparison.max_row + 1)
    )
    assert any(
        comparison.cell(row=row, column=transunion_account_column).value
        for row in range(2, comparison.max_row + 1)
    )
    assert any(
        comparison.cell(row=row, column=equifax_balance_column).value
        for row in range(2, comparison.max_row + 1)
    )
    side_by_side = workbook["Side By Side Negative"]
    side_headers = [side_by_side.cell(row=1, column=column).value for column in range(1, side_by_side.max_column + 1)]
    assert side_headers == [
        "Account Group",
        "Field",
        "Experian",
        "Equifax",
        "TransUnion",
        "Mismatch/Missing?",
        "Plain-English Review",
    ]
    side_text = " ".join(
        str(side_by_side.cell(row=row, column=column).value or "")
        for row in range(2, side_by_side.max_row + 1)
        for column in range(1, side_by_side.max_column + 1)
    )
    assert "MIDLAND CREDIT MANAGEMENT" in side_text
    assert "Balance" in side_text
    assert "Missing from one or more bureaus" in side_text or "Mismatch across bureaus" in side_text
    raw_evidence = workbook["Raw Evidence Index"]
    raw_evidence_headers = [raw_evidence.cell(row=1, column=column).value for column in range(1, raw_evidence.max_column + 1)]
    assert raw_evidence_headers == [
        "Raw Block ID",
        "File",
        "File Hash",
        "Page",
        "Bureau",
        "Account Name",
        "Account Number",
        "Raw Block Hash",
        "Raw Text Snippet",
        "Parser Confidence",
        "Admin Review Required",
    ]
    raw_evidence_text = " ".join(
        str(raw_evidence.cell(row=row, column=column).value or "")
        for row in range(2, raw_evidence.max_row + 1)
        for column in range(1, raw_evidence.max_column + 1)
    )
    assert "MIDLAND CREDIT MANAGEMENT" in raw_evidence_text
    assert "Account Number" in raw_evidence_text
    qa_verification = workbook["QA Verification"]
    qa_headers = [qa_verification.cell(row=1, column=column).value for column in range(1, qa_verification.max_column + 1)]
    assert qa_headers == ["Check ID", "Check Name", "Result", "Severity", "Evidence", "Fix Required"]
    qa_text = " ".join(
        str(qa_verification.cell(row=row, column=column).value or "")
        for row in range(2, qa_verification.max_row + 1)
        for column in range(1, qa_verification.max_column + 1)
    )
    assert "Raw evidence present" in qa_text
    assert "Field evidence present" in qa_text
    assert "Production approval status" in qa_text
    security_audit = workbook["Security Audit Summary"]
    security_audit_text = " ".join(
        str(security_audit.cell(row=row, column=column).value or "")
        for row in range(1, security_audit.max_row + 1)
        for column in range(1, security_audit.max_column + 1)
    )
    assert "Pre-Scan Health Check Result" in security_audit_text
    assert "Safe Mode Enabled" in security_audit_text
    assert "Scan Allowed" in security_audit_text
    parser_qa = workbook["Parser QA Warnings"]
    parser_qa_headers = [parser_qa.cell(row=1, column=column).value for column in range(1, parser_qa.max_column + 1)]
    assert "Warning Type" in parser_qa_headers
    assert "Admin Action" in parser_qa_headers
    desktop_dashboard = workbook["Desktop Dashboard"]
    dashboard_text = " ".join(
        str(desktop_dashboard.cell(row=row, column=column).value or "")
        for row in range(1, desktop_dashboard.max_row + 1)
        for column in range(1, desktop_dashboard.max_column + 1)
    )
    assert "Health score" in dashboard_text
    assert "Findings" in dashboard_text
    assert "Upload proof" in dashboard_text
    desktop_workbox = workbook["Desktop Staff Workbox"]
    workbox_headers = [desktop_workbox.cell(row=1, column=column).value for column in range(1, desktop_workbox.max_column + 1)]
    assert "Priority Score" in workbox_headers
    assert "Recommended Letter Types" in workbox_headers
    desktop_matrix = workbook["Desktop Field Matrix"]
    matrix_headers = [desktop_matrix.cell(row=1, column=column).value for column in range(1, desktop_matrix.max_column + 1)]
    assert matrix_headers == ["Account Name", "Field", "Label", "Equifax", "Experian", "TransUnion", "Differs"]
    matrix_text = " ".join(
        str(desktop_matrix.cell(row=row, column=column).value or "")
        for row in range(2, desktop_matrix.max_row + 1)
        for column in range(1, desktop_matrix.max_column + 1)
    )
    assert "Current balance" in matrix_text
    assert "Date reported or updated" in matrix_text
    raw_dates = workbook["Raw Tradelines With Dates"]
    raw_date_headers = [raw_dates.cell(row=1, column=column).value for column in range(1, raw_dates.max_column + 1)]
    assert "date_opened" in raw_date_headers
    assert "date_of_first_delinquency" in raw_date_headers
    assert "raw_evidence" in raw_date_headers
    dates_audit = workbook["Dates Found Audit"]
    dates_audit_headers = [dates_audit.cell(row=1, column=column).value for column in range(1, dates_audit.max_column + 1)]
    assert "normalized_date" in dates_audit_headers
    assert "context" in dates_audit_headers
    dates_audit_text = " ".join(
        str(dates_audit.cell(row=row, column=column).value or "")
        for row in range(2, dates_audit.max_row + 1)
        for column in range(1, dates_audit.max_column + 1)
    )
    assert "Date Opened" in dates_audit_text
    date_issues = workbook["Date Issues To Dispute"]
    date_issue_headers = [date_issues.cell(row=1, column=column).value for column in range(1, date_issues.max_column + 1)]
    assert "Severity" in date_issue_headers
    assert "What we found in plain English" in date_issue_headers
    draft_letters = workbook["Draft Letters"]
    draft_letter_text = " ".join(
        str(draft_letters.cell(row=row, column=column).value or "")
        for row in range(2, draft_letters.max_row + 1)
        for column in range(1, draft_letters.max_column + 1)
    )
    assert "To: Credit Bureau" in draft_letter_text
    assert "forward the dispute and all relevant information to the furnisher" in draft_letter_text
    assert "To: Furnisher / Collector" in draft_letter_text
    assert "basis for your reporting" in draft_letter_text
    assert "To: Debt Collector / Debt Buyer" in draft_letter_text
    assert "itemized accounting of the balance" in draft_letter_text
    expert = workbook["Metro 2 + FCRA Review"]
    expert_headers = [expert.cell(row=1, column=column).value for column in range(1, expert.max_column + 1)]
    assert "Metro 2 Fields To Review" in expert_headers
    assert "FCRA Sections / Duties" in expert_headers
    requirements = workbook["Metro 2 Requirements"]
    requirement_headers = [requirements.cell(row=1, column=column).value for column in range(1, requirements.max_column + 1)]
    assert "Metro 2 Profile" in requirement_headers
    assert "Required/Core Fields" in requirement_headers
    assert "Missing / Needs Validation" in requirement_headers
    requirements_text = " ".join(
        str(requirements.cell(row=row, column=column).value or "")
        for row in range(2, requirements.max_row + 1)
        for column in range(1, requirements.max_column + 1)
    )
    assert "Collection account" in requirements_text
    assert "Date of First Delinquency" in requirements_text
    assert "K1 Original Creditor Name" in requirements_text
    assert "Original Charge-off Amount" in requirements_text
    assert "Compliance Condition Code" in requirements_text
    assert "official licensed CDIA Metro 2 CRRG" in requirements_text
    guide_notes = workbook["Metro 2 Guide Notes"]
    guide_text = " ".join(
        str(guide_notes.cell(row=row, column=column).value or "")
        for row in range(1, guide_notes.max_row + 1)
        for column in range(1, guide_notes.max_column + 1)
    )
    assert "https://www.collect.org/cv13/Help/howtoreadthemetro2format.html" in guide_text
    assert "Payment History Profile" in guide_text
    assert "Consumer Information Indicator" in guide_text
    assert "official licensed CDIA Metro 2 CRRG" in guide_text
    fcra_compliance = workbook["FCRA Compliance Review"]
    fcra_headers = [fcra_compliance.cell(row=1, column=column).value for column in range(1, fcra_compliance.max_column + 1)]
    assert "FCRA Area" in fcra_headers
    assert "Law Reference" in fcra_headers
    assert "Scanner Signals" in fcra_headers
    assert "Tracking Action" in fcra_headers
    fcra_workbook_text = " ".join(
        str(fcra_compliance.cell(row=row, column=column).value or "")
        for row in range(2, fcra_compliance.max_row + 1)
        for column in range(1, fcra_compliance.max_column + 1)
    )
    assert "Bureau reinvestigation duty" in fcra_workbook_text
    assert "Direct furnisher dispute" in fcra_workbook_text
    assert "DOFD and obsolete information review" in fcra_workbook_text
    assert "FCRA 623" in fcra_workbook_text
    fcra_rights = workbook["FCRA Rights Regulators"]
    fcra_rights_text = " ".join(
        str(fcra_rights.cell(row=row, column=column).value or "")
        for row in range(1, fcra_rights.max_row + 1)
        for column in range(1, fcra_rights.max_column + 1)
    )
    assert "Bureau of Consumer Financial Protection" in fcra_rights_text
    assert "Federal Trade Commission Consumer Response Center" in fcra_rights_text
    assert "state Attorney General" in fcra_rights_text
    assert "Virginia" in fcra_rights_text
    assert "Security freeze" in fcra_rights_text
    assert "Adverse action notice" in fcra_rights_text
    assert "Maryland Office of Financial Regulation" in fcra_rights_text
    assert "100 words" in fcra_rights_text
    bureau_help = workbook["Bureau Help + FDCPA"]
    bureau_help_text = " ".join(
        str(bureau_help.cell(row=row, column=column).value or "")
        for row in range(1, bureau_help.max_row + 1)
        for column in range(1, bureau_help.max_column + 1)
    )
    assert "Equifax" in bureau_help_text
    assert "TransUnion" in bureau_help_text
    assert "Verified and updated" in bureau_help_text
    assert "Remains" in bureau_help_text
    assert "Validation notice" in bureau_help_text
    assert "Contact time and place limits" in bureau_help_text
    field_compliance = workbook["Field Compliance Audit"]
    field_headers = [field_compliance.cell(row=1, column=column).value for column in range(1, field_compliance.max_column + 1)]
    assert "Metro 2 Concept" in field_headers
    assert "FCRA / Reg V Basis" in field_headers
    assert "Verification Ask" in field_headers
    field_workbook_text = " ".join(
        str(field_compliance.cell(row=row, column=column).value or "")
        for row in range(2, field_compliance.max_row + 1)
        for column in range(1, field_compliance.max_column + 1)
    )
    assert "K1 original creditor/source account" in field_workbook_text
    assert "Compliance condition code" in field_workbook_text
    assert "Consumer information indicator" in field_workbook_text
    assert "Date of First Delinquency" in field_workbook_text
    assert "Original Creditor" in field_workbook_text
    assert "FCRA 623(a)(5)" in field_workbook_text
    eoscar = workbook["e-OSCAR Packaging Review"]
    eoscar_headers = [eoscar.cell(row=1, column=column).value for column in range(1, eoscar.max_column + 1)]
    assert "e-OSCAR / ACDV Category" in eoscar_headers
    assert "ACDV Packaging Steps" in eoscar_headers
    assert "Evidence Hint" in eoscar_headers
    eoscar_workbook_text = " ".join(
        str(eoscar.cell(row=row, column=column).value or "")
        for row in range(2, eoscar.max_row + 1)
        for column in range(1, eoscar.max_column + 1)
    )
    assert "ACDV" in eoscar_workbook_text
    assert "AUD" in eoscar_workbook_text
    assert "Do not claim direct e-OSCAR access" in eoscar_workbook_text
    notice_rules = workbook["FCRA Notice Rules"]
    notice_rule_text = " ".join(
        str(notice_rules.cell(row=row, column=2).value or "")
        for row in range(2, notice_rules.max_row + 1)
    )
    assert "customer approval required" in notice_rule_text
    assert "written reinvestigation results" in notice_rule_text
    assert "CFPB credit reporting company dispute letter template" in notice_rule_text
    assert "CFPB furnisher dispute letter template" in notice_rule_text
    assert "specific information being disputed" in notice_rule_text
    dispute_methods = workbook["Dispute Methods"]
    method_text = " ".join(
        str(dispute_methods.cell(row=row, column=1).value or "") + " " +
        str(dispute_methods.cell(row=row, column=2).value or "")
        for row in range(2, dispute_methods.max_row + 1)
    )
    assert "FCRA Bureau Dispute" in method_text
    assert "Direct Furnisher Dispute" in method_text
    assert "FDCPA Debt Validation Request" in method_text
    assert "CFPB / CFPA Complaint Escalation" in method_text
    assert "Metro 2 Field-Level Dispute" in method_text
    dispute_sop = workbook["Dispute SOP"]
    sop_text = " ".join(
        str(dispute_sop.cell(row=row, column=1).value or "") + " " +
        str(dispute_sop.cell(row=row, column=2).value or "") + " " +
        str(dispute_sop.cell(row=row, column=9).value or "")
        for row in range(2, dispute_sop.max_row + 1)
    )
    assert "Round 1" in sop_text
    assert "Round 2" in sop_text
    assert "Round 3" in sop_text
    assert "CFPB" in sop_text
    draft_letters = (tmp_path / "draft_dispute_letters.txt").read_text(encoding="utf-8")
    assert "DRAFT" in draft_letters
    assert "CUSTOMER REVIEW AND APPROVAL REQUIRED" in draft_letters


def test_equifax_help_text_does_not_become_account():
    result = parse_reports({
        "equifax june 29 2026.pdf": {
            "text": EQUIFAX_HELP_SECTION_THAT_IS_NOT_AN_ACCOUNT,
            "bureau": "Equifax",
        },
    })
    data = result_to_dict(result)
    account_names = [item["account_name"] for item in data["tradelines"]]

    assert any("MIDLAND CREDIT MANAGEMENT" in name for name in account_names)
    assert not any("consumer added notices" in name.lower() for name in account_names)
    assert not any("equifax.com/personal/help" in name.lower() for name in account_names)
    assert not any("see if an account is open" in name.lower() for name in account_names)


def test_equifax_duplicate_fragments_collapse_to_best_account_name():
    result = parse_reports({
        "equifax june 29 2026.pdf": {
            "text": EQUIFAX_DUPLICATE_ACCOUNT_WINDOWS,
            "bureau": "Equifax",
        },
    })
    data = result_to_dict(result)
    account_names = [item["account_name"] for item in data["tradelines"]]

    assert account_names == ["MIDLAND CREDIT MANAGEMENT - Closed"]
    assert not any(name == "Prepared for" for name in account_names)
    assert not any("Loan/Account Type" in name for name in account_names)


def test_boilerplate_disclosure_text_is_not_a_tradeline():
    result = parse_reports({"experian_boilerplate.pdf": {"text": BOILERPLATE_SAMPLE, "bureau": "Experian"}})
    data = result_to_dict(result)
    assert data["tradelines"] == []
    assert data["issues"] == []


def test_money_labels_are_not_account_names():
    assert is_bad_account_name("High Balance $2,335")
    assert is_bad_account_name("Credit Limit $500")
    assert is_bad_account_name("Amount Past Due $75")


def test_page_level_bureau_detection_positive_accounts_no_disputes(tmp_path):
    result = parse_reports({
        "mixed_bureau_upload.pdf": {"text": PAGE_LEVEL_POSITIVE_SAMPLE, "bureau": "Experian"},
    })
    data = result_to_dict(result)
    by_account = {item["account_number_masked"]: item for item in data["tradelines"]}

    assert by_account["*1664"]["bureau"] == "Equifax"
    assert by_account["*1664"]["bureau_conflict"] is True
    assert "upload_metadata_conflicts_with_page_header" in by_account["*1664"]["parser_warnings"]
    assert by_account["*4796"]["bureau"] == "Experian"
    assert all(item["source_file_hash"] for item in data["tradelines"])
    assert all(item["raw_block_id"] and item["raw_block_hash"] for item in data["tradelines"])
    assert by_account["*1664"]["is_negative"] is False
    assert by_account["*4796"]["is_negative"] is False
    for item in data["tradelines"]:
        for field_name in ["account_name", "account_number_masked", "account_type", "balance", "status", "date_opened", "date_reported"]:
            assert field_name in item["field_evidence"]
            assert item["field_evidence"][field_name]["raw_line"]
    assert data["issues"] == []
    assert data["recommended_letter_queue"] == []

    write_outputs(result, tmp_path)
    workbook = load_workbook(tmp_path / "credit_vivo_desktop_scanner_output.xlsx", read_only=True)
    comparison = workbook["3 Bureau Comparison"]
    headers = [comparison.cell(row=1, column=column).value for column in range(1, comparison.max_column + 1)]
    assert len(headers) == len(set(headers))

    account_column = headers.index("Account #") + 1
    equifax_column = headers.index("Equifax Account #") + 1
    experian_column = headers.index("Experian Account #") + 1
    transunion_column = headers.index("TransUnion Account #") + 1
    errors_column = headers.index("Errors / Findings") + 1
    action_column = headers.index("Recommended Action") + 1

    rows_by_account = {
        comparison.cell(row=row, column=account_column).value: row
        for row in range(2, comparison.max_row + 1)
    }
    eq_row = rows_by_account["*1664"]
    ex_row = rows_by_account["*4796"]

    assert comparison.cell(row=eq_row, column=equifax_column).value == "*1664"
    assert comparison.cell(row=eq_row, column=experian_column).value in (None, "")
    assert comparison.cell(row=eq_row, column=transunion_column).value in (None, "")
    assert comparison.cell(row=eq_row, column=errors_column).value == "No verified dispute issue detected."
    assert "Missing bureau presence alone is not a dispute issue" in comparison.cell(row=eq_row, column=action_column).value

    assert comparison.cell(row=ex_row, column=equifax_column).value in (None, "")
    assert comparison.cell(row=ex_row, column=experian_column).value == "*4796"
    assert comparison.cell(row=ex_row, column=transunion_column).value in (None, "")
    assert comparison.cell(row=ex_row, column=errors_column).value == "No verified dispute issue detected."

    assert workbook["Side By Side Negative"].max_row == 1
    assert workbook["Draft Letters"].max_row == 1
    assert "Raw Evidence Index" in workbook.sheetnames
    assert "QA Verification" in workbook.sheetnames


def test_possible_skipped_tradeline_goes_to_admin_qa_not_disputes(tmp_path):
    result = parse_reports({
        "possible_skipped.pdf": {"text": POSSIBLE_SKIPPED_TRADELINE_SAMPLE, "bureau": "Equifax"},
    })
    data = result_to_dict(result)

    assert data["tradelines"] == []
    assert data["issues"] == []
    assert data["recommended_letter_queue"] == []
    assert len(data["parser_qa_warnings"]) == 1
    warning = data["parser_qa_warnings"][0]
    assert warning["warning_type"] == "possible_skipped_tradeline"
    assert warning["customer_visible"] is False
    assert warning["creates_dispute_issue"] is False
    assert warning["account_number_guess"] == "*9999"
    assert "Admin QA should review" in warning["admin_action"]

    write_outputs(result, tmp_path)
    workbook = load_workbook(tmp_path / "credit_vivo_desktop_scanner_output.xlsx", read_only=True)
    parser_qa = workbook["Parser QA Warnings"]
    parser_qa_text = " ".join(
        str(parser_qa.cell(row=row, column=column).value or "")
        for row in range(1, parser_qa.max_row + 1)
        for column in range(1, parser_qa.max_column + 1)
    )
    assert "possible_skipped_tradeline" in parser_qa_text
    assert "*9999" in parser_qa_text
    assert workbook["Draft Letters"].max_row == 1


def test_three_bureau_comparison_includes_ungrouped_accounts():
    data = {
        "tradelines": [
            {
                "id": "eq1",
                "bureau": "Equifax",
                "source_filename": "equifax.pdf",
                "account_name": "CREDIT ONE BANK",
                "account_number_masked": "*1111",
                "account_type": "Credit Card",
                "balance": "$100",
                "status": "Pays As Agreed",
            },
            {
                "id": "ex1",
                "bureau": "Experian",
                "source_filename": "experian.pdf",
                "account_name": "CREDIT ONE BANK",
                "account_number_masked": "*2222",
                "account_type": "Credit Card",
                "balance": "$125",
                "status": "Open",
            },
            {
                "id": "tu_only",
                "bureau": "TransUnion",
                "source_filename": "transunion.pdf",
                "account_name": "LVNV FUNDING LLC",
                "account_number_masked": "*3333",
                "account_type": "Collection",
                "balance": "$500",
                "status": "Collection",
            },
        ],
        "issues": [],
        "cross_bureau_groups": [{"group_id": "group-credit-one", "tradeline_ids": ["eq1", "ex1"]}],
    }

    rows = build_three_bureau_comparison_rows(data)
    account_names = [row[0] for row in rows[1:]]

    assert any("CREDIT ONE BANK" in name for name in account_names)
    assert "LVNV FUNDING LLC" in account_names
    assert len(rows) == 3


def test_parser_fills_columns_from_realistic_bureau_snippets():
    result = parse_reports({
        "equifax_realistic.pdf": {"text": REALISTIC_EQUIFAX, "bureau": "Equifax"},
        "experian_realistic.pdf": {"text": REALISTIC_EXPERIAN, "bureau": "Experian"},
        "transunion_realistic.pdf": {"text": REALISTIC_TRANSUNION, "bureau": "TransUnion"},
    })
    data = result_to_dict(result)
    by_bureau = {item["bureau"]: item for item in data["tradelines"]}

    assert by_bureau["Equifax"]["account_name"] == "CREDIT ONE BANK"
    assert by_bureau["Equifax"]["account_type"] == "Credit Card"
    assert by_bureau["Equifax"]["responsibility"] == "Individual"
    assert by_bureau["Equifax"]["credit_limit"] == "$500"
    assert by_bureau["Equifax"]["high_credit_or_original_amount"] == "$696"
    assert by_bureau["Equifax"]["past_due"] == "$60"
    assert by_bureau["Equifax"]["date_last_payment"] == "02/11/2026"

    assert by_bureau["Experian"]["account_name"] == "CREDIT ONE BANK"
    assert by_bureau["Experian"]["balance"] == "$488"
    assert by_bureau["Experian"]["date_reported"] == "03/11/2026"
    assert by_bureau["Experian"]["credit_limit"] == "$500"
    assert by_bureau["Experian"]["high_credit_or_original_amount"] == "$696"
    assert "disputed by consumer" in by_bureau["Experian"]["remarks"].lower()

    assert by_bureau["TransUnion"]["account_name"] == "CREDIT ONE BANK"
    assert by_bureau["TransUnion"]["account_number_masked"].endswith("8171")
    assert by_bureau["TransUnion"]["account_type"] == "Revolving Account"
    assert by_bureau["TransUnion"]["status"] == "Current Account"
    assert by_bureau["TransUnion"]["date_reported"] == "03/11/2026"
    assert by_bureau["TransUnion"]["date_last_payment"] == "02/11/2026"
    assert by_bureau["TransUnion"]["credit_limit"] == "$500"
    assert by_bureau["TransUnion"]["high_credit_or_original_amount"] == "$696"


def test_original_creditor_alone_does_not_merge_collection_with_card():
    result = parse_reports({
        "equifax_macys.pdf": {"bureau": "Equifax", "text": """
--- PAGE 1 ---
Equifax Credit Report

MACYS/CITIBANK NA - Closed
PO BOX 6789, SIOUX FALLS, SD 57117-6789 Date Reported: 12/25/2019 | Balance: $0
Account Number: *0540 | Owner: Individual Account Credit Limit: $100 | High Credit: $99
Loan/Account Type: Charge Account | Status: Pays As Agreed
"""},
        "transunion_midland.pdf": {"bureau": "TransUnion", "text": """
--- PAGE 1 ---
TransUnion Credit Report

MIDLAND CREDIT MANAGEMENT INC
32075****
Date Opened 02/17/2022
Responsibility Individual Account
Account Type Open Account
Loan Type FACTORING COMPANY ACCOUNT
Balance $1,445
Date Updated 03/19/2026
High Balance $1,127
Original Creditor CITIBANK N A
Pay Status >Collection<
Remarks Account information disputed by consumer (FCRA); >PLACED FOR COLLECTION<
"""},
    })
    data = result_to_dict(result)
    assert len(data["tradelines"]) == 2
    assert data["cross_bureau_groups"] == []


def test_dense_adjacent_accounts_stay_isolated_across_three_bureaus():
    reports = {}
    for bureau in ("Experian", "Equifax", "TransUnion"):
        reports[f"{bureau.lower()}-dense.pdf"] = {"bureau": bureau, "text": f"""
--- PAGE 1 ---
{bureau} Credit Report
MIDLAND CREDIT MANAGEMENT
Account Number: 1234567890
Account Type: Collection
Original Creditor: CAPITAL ONE BANK
Balance: $1,234
Status: Collection
Date Opened: 01/10/2021
Remarks: Account placed for collection
CAPITAL ONE
Account Number: 999988881111
Account Type: Credit Card
Balance: $0
Past Due: $0
Status: Charge-off transferred or sold
Date Opened: 05/01/2019
Date of First Delinquency: 10/01/2020
"""}
    data = result_to_dict(parse_reports(reports))
    assert len(data["tradelines"]) == 6
    for bureau in ("Experian", "Equifax", "TransUnion"):
        rows = [item for item in data["tradelines"] if item["bureau"] == bureau]
        assert {item["account_number_masked"] for item in rows} == {"*7890", "*1111"}
        midland = next(item for item in rows if item["account_number_masked"] == "*7890")
        capital_one = next(item for item in rows if item["account_number_masked"] == "*1111")
        assert "charge-off" not in midland["raw_block"].lower()
        assert "midland" not in capital_one["raw_block"].lower()
        assert "past_due" not in capital_one["negative_signals"]
