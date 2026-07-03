from pathlib import Path

from openpyxl import load_workbook

from credit_vivo_proprietary_engine import (
    build_ours_three_bureaus_comparison_rows,
    detect_bureau,
    parse_reports,
    result_to_dict,
    write_outputs,
)
from rules_engine import blocked_compliance_phrases, classify_negative_tradeline, load_scanner_rules


SAMPLE_REPORTS = {
    "experian.pdf": {
        "bureau": "Experian",
        "text": """
--- PAGE 1 ---
Experian Credit Report

MIDLAND CREDIT MANAGEMENT
Account Number: 1234567890
Account Type: Collection
Original Creditor: CAPITAL ONE BANK
Balance: $1,234
Status: Collection
Date Opened: 01/10/2021
Date Reported: 04/01/2026
Remarks: Account placed for collection
Payment History: 30 days past due as of March 2026
""",
    },
    "equifax.pdf": {
        "bureau": "Equifax",
        "text": """
--- PAGE 1 ---
Equifax Credit Report

MIDLAND CREDIT MANAGEMENT
Account Number: 1234567890
Loan/Account Type: Debt Buyer Account
Balance: $1,999
Status: Collection Account
Date Opened: 01/10/2021
Date Reported: 05/01/2026
Date of 1st Delinquency:
Amount Past Due: $1,999
Remarks: Account placed for collection
""",
    },
    "transunion.pdf": {
        "bureau": "TransUnion",
        "text": """
--- PAGE 1 ---
TransUnion Credit Report

MIDLAND CREDIT MANAGEMENT
Account Number: 1234567890
Account Type: Collection
Original Creditor: CAPITAL ONE BANK
Balance $1,234
Pay Status Paid collection
Date Opened 01/10/2021
Date Updated 04/01/2026
Date of First Delinquency 01/01/2021
Remarks Account placed for collection
""",
    },
}


def parsed_data():
    return result_to_dict(parse_reports(SAMPLE_REPORTS))


def test_rule_libraries_load_required_sections():
    rules = load_scanner_rules()
    assert rules["negative_account_rules"]["categories"]
    assert rules["metro2_field_map"]["fields"]
    assert rules["fcra_rules"]["rules"]
    assert rules["eoscar_workflow_rules"]["safe_letter_recommendations"]
    assert "automatic dispute" in rules["compliance_guard_rules"]["blocked_phrases"]


def test_bureau_detection():
    assert detect_bureau("exp.pdf", "Experian credit report") == "Experian"
    assert detect_bureau("eq.pdf", "Equifax Information Services") == "Equifax"
    assert detect_bureau("tu.pdf", "TransUnion LLC") == "TransUnion"


def test_negative_tradeline_classification():
    data = parsed_data()
    collection_item = next(item for item in data["tradelines"] if "MIDLAND" in item["account_name"])
    labels = {match["label"] for match in classify_negative_tradeline(collection_item)}
    assert "Collection" in labels
    assert labels & {"Debt buyer", "Past due", "Duplicate / overlapping debt review"}


def test_metro2_and_fcra_issue_rules_are_emitted():
    data = parsed_data()
    issue_types = {issue["issue_type"] for issue in data["issues"]}
    assert "missing_dofd_review" in issue_types or "negative_outdated_adverse_review" in issue_types
    assert "cross_bureau_balance_mismatch" in issue_types
    assert "original_creditor_missing_or_inconsistent" in issue_types or "negative_collection_review" in issue_types
    assert "duplicate_overlap_review" in issue_types
    assert "removal_obsolescence_date_mismatch_or_missing" in issue_types
    assert "payment_history_inconsistency" in issue_types


def test_compliance_guard_blocked_phrases():
    blocked = blocked_compliance_phrases("This is a guaranteed removal and automatic dispute.")
    assert "guaranteed removal" in blocked
    assert "automatic dispute" in blocked


def test_workbook_includes_v9_sheet_contract(tmp_path: Path):
    result = parse_reports(SAMPLE_REPORTS)
    write_outputs(result, tmp_path)
    workbook = load_workbook(tmp_path / "credit_vivo_desktop_scanner_output.xlsx")
    required_sheets = {
        "Dashboard",
        "Ours 3 Bureaus Comparison",
        "Account_Summary",
        "Identity_Cleanup",
        "Negative_Definitions",
        "Negative_Account_Rules",
        "License_Check",
        "State_License_Links",
        "Dispute_Cycle_Status",
        "Exact_Letters_To_Mail",
        "Escalation_Addresses",
        "Complaint_Packet",
        "Scanner_Skills_Map",
        "FICO_Scenario_Planner",
        "Codex_Build_Task",
        "Read_Me",
        "Ground_Truth_Validation",
        "QA_Verification",
        "Security_Audit_Summary",
        "Production_Gate",
        "Positive_Accounts_Keep",
    }
    assert required_sheets.issubset(set(workbook.sheetnames))

    ours = workbook["Ours 3 Bureaus Comparison"]
    assert ours.cell(row=1, column=1).value == "Three-Bureau Negative Tradeline Forensic Comparison - Experian / Equifax / TransUnion"
    headers = [ours.cell(row=4, column=column).value for column in range(1, 13)]
    assert headers == [
        "Field #",
        "Account Info",
        "Experian",
        "Equifax",
        "TransUnion",
        "Forensic issue / dispute lead",
        "3-CRA Status",
        "AI Error / Inaccuracy Found",
        "Reason / Why It Matters",
        "Dispute / Verification Request",
        "Priority",
        "Evidence / Notes",
    ]
    assert ours.freeze_panes == "A5"
    assert ours.column_dimensions["B"].width == 34
    assert ours.column_dimensions["K"].width == 82
    assert ours.max_row >= 355
    assert any(
        ours.cell(row=row, column=2).value == "Account Type"
        for row in range(5, ours.max_row + 1)
    )
    assert any(
        ours.cell(row=row, column=8).value
        for row in range(5, ours.max_row + 1)
    )
    skills = workbook["Scanner_Skills_Map"]
    skill_headers = [skills.cell(row=4, column=column).value for column in range(1, 8)]
    assert skill_headers == [
        "Skill ID",
        "Skill",
        "Scanner Role",
        "Used For",
        "Output Area",
        "Approval / Safety Gate",
        "Customer Visible",
    ]
    skill_ids = {
        skills.cell(row=row, column=1).value
        for row in range(5, skills.max_row + 1)
        if skills.cell(row=row, column=1).value
    }
    assert {
        "credit_report_parser",
        "workbook_output_qa",
        "creditvivo_compliance_reviewer",
        "dispute_strategy_assistant",
        "creditvivo_product_manager",
        "letter_lifecycle_manager",
        "security_privacy_reviewer",
    }.issubset(skill_ids)
    identity = workbook["Identity_Cleanup"]
    identity_headers = [identity.cell(row=4, column=column).value for column in range(1, 12)]
    assert identity_headers == [
        "Action",
        "Identity Field",
        "Raw Report Value",
        "Keep One Correct Value",
        "Bureau",
        "Source File",
        "Page",
        "Brief Compliance Review",
        "Requested Outcome",
        "Customer Confirmation Needed",
        "Admin Notes",
    ]
    assert workbook["Account_Summary"].max_row >= 15
    assert workbook["Account_Summary"].max_column >= 12
    assert identity.max_row >= 10
    assert workbook["License_Check"].max_row >= 16
    assert workbook["License_Check"].max_column >= 14
    assert workbook["State_License_Links"].max_row >= 13
    assert workbook["Dispute_Cycle_Status"].max_row >= 15
    assert workbook["Escalation_Addresses"].max_row >= 19
    assert workbook["Complaint_Packet"].max_row >= 12
    assert workbook["FICO_Scenario_Planner"].max_row >= 11
    assert workbook["Ground_Truth_Validation"].max_row >= 12
    assert workbook["Security_Audit_Summary"].max_row >= 10
    assert workbook["Production_Gate"].max_row >= 10
    assert workbook["Ground_Truth_Validation"].cell(row=4, column=4).value
    assert workbook["Security_Audit_Summary"].cell(row=4, column=1).value
    assert workbook["Production_Gate"].cell(row=4, column=1).value

    rules = workbook["Negative_Account_Rules"]
    rules_text = " ".join(str(rules.cell(row=row, column=column).value or "") for row in range(1, rules.max_row + 1) for column in range(1, rules.max_column + 1))
    assert "Collection" in rules_text
    assert "Charge-off" in rules_text
    assert "Draft review data only" in rules_text

    exact_letters = workbook["Exact_Letters_To_Mail"]
    exact_letters_text = " ".join(
        str(exact_letters.cell(row=row, column=column).value or "")
        for row in range(1, exact_letters.max_row + 1)
        for column in range(1, exact_letters.max_column + 1)
    )
    assert "Bureau review/dispute letter draft" in exact_letters_text
    assert "Furnisher direct dispute draft" in exact_letters_text
    assert "Debt validation draft" in exact_letters_text
    assert "Method of Verification draft" in exact_letters_text
    assert "Reinvestigation draft" in exact_letters_text
    assert "Documented follow-up packet" in exact_letters_text
    assert "Complaint preparation packet" in exact_letters_text
    assert "Not sent" in exact_letters_text

    assert "CFPB_Packet_Checklist" in workbook.sheetnames
    assert "3B_Comparison_Attachment" in workbook.sheetnames
    assert "Document_Vault" in workbook.sheetnames
    assert "Lob_Tracking" in workbook.sheetnames
    packet_headers = [workbook["CFPB_Packet_Checklist"].cell(row=4, column=column).value for column in range(1, 15)]
    assert "Customer E-Sign Required" in packet_headers
    assert "Mailing Allowed" in packet_headers
    attachment_headers = [workbook["3B_Comparison_Attachment"].cell(row=4, column=column).value for column in range(1, 18)]
    assert attachment_headers[:6] == [
        "Account / Field",
        "Equifax Raw Value",
        "Experian Raw Value",
        "TransUnion Raw Value",
        "Main Issue",
        "License / Authority Status",
    ]

    comparison_text = " ".join(
        str(ours.cell(row=row, column=column).value or "")
        for row in range(1, min(ours.max_row, 80) + 1)
        for column in range(1, ours.max_column + 1)
    )
    assert "Pass / Verify Source Records" in comparison_text
    assert "Compliance Review - Does Not Match" in comparison_text or "Compliance Review - Missing" in comparison_text
    assert "FCRA 607(b)" in comparison_text
    assert "Metro 2:" in comparison_text
    assert "Outcome:" in comparison_text


def test_cfpb_packet_system_preserves_raw_values_and_blocks_mailing(tmp_path: Path):
    result = parse_reports(SAMPLE_REPORTS)
    data = result_to_dict(result)
    packet_system = data["cfpb_packet_system"]
    assert packet_system["security"]["automatic_mailing_enabled"] is False
    assert packet_system["security"]["automatic_complaint_submission_enabled"] is False
    assert packet_system["dispute_packets"]
    assert packet_system["document_vault"]["records"]
    assert any(row["browser_local_storage_allowed"] is False for row in packet_system["document_vault"]["records"])
    assert all(packet["mailing_allowed"] is False for packet in packet_system["dispute_packets"])
    assert all(packet["auto_send"] is False for packet in packet_system["dispute_packets"])
    assert any(packet["packet_type"] == "method_of_verification_request" for packet in packet_system["dispute_packets"])
    assert any(packet["packet_type"] == "reinvestigation_request" for packet in packet_system["dispute_packets"])
    assert any(packet["packet_type"] == "escalation_follow_up" for packet in packet_system["dispute_packets"])
    assert any(packet.get("lob_ready_preview", {}).get("blocked_until") for packet in packet_system["dispute_packets"])
    assert any("customer_esign_required" in packet["packet_gate"]["block_reasons"] for packet in packet_system["dispute_packets"])
    comparison_text = " ".join(
        str(row.get("equifax_raw_value", "")) + " " +
        str(row.get("experian_raw_value", "")) + " " +
        str(row.get("transunion_raw_value", ""))
        for row in packet_system["three_bureau_comparison_attachment"]
    )
    assert "MIDLAND CREDIT MANAGEMENT" in comparison_text
    assert "License/business status review needed." in " ".join(row["license_authority_status"] for row in packet_system["three_bureau_comparison_attachment"])

    write_outputs(result, tmp_path)
    assert (tmp_path / "document_vault" / "document_vault_manifest.json").exists()
    assert (tmp_path / "document_vault" / "three_bureau_comparison_attachment.json").exists()
    assert (tmp_path / "lob_ready_letter_preview_manifest.json").exists()
    assert (tmp_path / "letters").exists()


def test_collection_open_account_type_gets_specific_compliance_review():
    data = {
        "scanner_rules_library": load_scanner_rules(),
        "tradelines": [
            {
                "id": "eq1",
                "bureau": "Equifax",
                "account_name": "MIDLAND CREDIT MANAGEMENT",
                "account_type": "Debt Buyer Account",
                "status": "Collection",
                "raw_block": "MIDLAND CREDIT MANAGEMENT Loan/Account Type: Debt Buyer Account Status: Collection",
                "source_filename": "equifax.pdf",
                "page_start": 4,
            },
            {
                "id": "ex1",
                "bureau": "Experian",
                "account_name": "MIDLAND CREDIT MANAGEMENT",
                "account_type": "Debt Buyer",
                "status": "Collection account",
                "raw_block": "MIDLAND CREDIT MANAGEMENT Account Type Debt Buyer Status Collection account",
                "source_filename": "experian.pdf",
                "page_start": 18,
            },
            {
                "id": "tu1",
                "bureau": "TransUnion",
                "account_name": "MIDLAND CREDIT MANAGEMENT",
                "account_type": "Open Account",
                "status": "Collection",
                "raw_block": "MIDLAND CREDIT MANAGEMENT Account Type Open Account Pay Status Collection",
                "source_filename": "transunion.pdf",
                "page_start": 22,
            },
        ],
        "cross_bureau_groups": [{"group_id": "g1", "tradeline_ids": ["eq1", "ex1", "tu1"]}],
        "issues": [],
    }
    rows = build_ours_three_bureaus_comparison_rows(data)
    assert rows[4][0] == "MIDLAND CREDIT MANAGEMENT"
    account_type_row = next(row for row in rows if len(row) > 1 and row[1] == "Account Type")
    assert account_type_row[2] == "Debt Buyer"
    assert account_type_row[3] == "Debt Buyer Account"
    assert account_type_row[4] == "Open Account"
    assert "Account type" in account_type_row[7] or "Collection/debt-buyer" in account_type_row[7]
    assert "FCRA" in account_type_row[8]
    assert "correct" in account_type_row[9].lower() or "verify" in account_type_row[9].lower()


def test_comparison_group_name_uses_report_visible_bureau_names():
    data = {
        "scanner_rules_library": load_scanner_rules(),
        "tradelines": [
            {
                "id": "eq1",
                "bureau": "Equifax",
                "account_name": "MIDLAND CREDIT MANAGEMENT - Closed",
                "account_type": "Debt Buyer Account",
                "status": "Collection",
                "raw_block": "MIDLAND CREDIT MANAGEMENT - Closed Loan/Account Type: Debt Buyer Account Status: Collection",
                "source_filename": "equifax.pdf",
                "page_start": 4,
            },
            {
                "id": "ex1",
                "bureau": "Experian",
                "account_name": "MIDLAND CREDIT MANAGEMEN",
                "account_type": "Debt Buyer",
                "status": "Collection account",
                "raw_block": "MIDLAND CREDIT MANAGEMEN Account Type Debt Buyer Status Collection account",
                "source_filename": "experian.pdf",
                "page_start": 18,
            },
            {
                "id": "tu1",
                "bureau": "TransUnion",
                "account_name": "MIDLAND CREDIT MANAGEMENT INC",
                "account_type": "Open Account",
                "status": "Collection",
                "raw_block": "MIDLAND CREDIT MANAGEMENT INC Account Type Open Account Pay Status Collection",
                "source_filename": "transunion.pdf",
                "page_start": 22,
            },
        ],
        "cross_bureau_groups": [{"group_id": "g1", "tradeline_ids": ["eq1", "ex1", "tu1"]}],
        "issues": [],
    }
    rows = build_ours_three_bureaus_comparison_rows(data)
    assert rows[4][0] == "MIDLAND CREDIT MANAGEMENT - Closed"
    account_name_row = next(row for row in rows if len(row) > 1 and row[1] == "Account/Furnisher Name")
    assert account_name_row[2] == "MIDLAND CREDIT MANAGEMEN"
    assert account_name_row[3] == "MIDLAND CREDIT MANAGEMENT - Closed"
    assert account_name_row[4] == "MIDLAND CREDIT MANAGEMENT INC"


def test_comparison_bureau_columns_use_raw_report_values_not_normalized_fields():
    data = {
        "scanner_rules_library": load_scanner_rules(),
        "tradelines": [
            {
                "id": "ex1",
                "bureau": "Experian",
                "account_name": "CREDIT ONE BANK",
                "account_type": "Credit card",
                "status": "Open.",
                "status_updated": "2023-09",
                "balance": "$488.00",
                "date_opened": "2022-12-27",
                "date_reported": "2026-03-11",
                "date_of_first_delinquency": "",
                "raw_block": (
                    "CREDIT ONE BANK\n"
                    "Account Type Credit card\n"
                    "Status Open.\n"
                    "Status Updated Sep 2023\n"
                    "Balance $488\n"
                    "Balance Updated 03/11/2026\n"
                    "Date Opened 12/27/2022\n"
                    "Payment History 30 days past due as of Aug 2023\n"
                    "By Apr 2030, this account is scheduled to go to a positive status."
                ),
                "source_filename": "experian.pdf",
                "page_start": 1,
            }
        ],
        "cross_bureau_groups": [{"group_id": "g1", "tradeline_ids": ["ex1"]}],
        "issues": [],
    }
    rows = build_ours_three_bureaus_comparison_rows(data)
    by_field = {row[1]: row for row in rows[4:] if row and isinstance(row[0], int)}
    assert by_field["Date Opened/Assigned"][2] == "12/27/2022"
    assert by_field["Date Reported/Updated"][2] == "03/11/2026"
    assert by_field["Status Updated"][2] == "Sep 2023"
    assert by_field["Current Balance"][2] == "$488"
    assert by_field["Estimated Removal/On Record Until"][2] == "Apr 2030"


def test_identity_cleanup_uses_raw_data_keep_one_delete_extras(tmp_path: Path):
    result = parse_reports({
        "experian_identity.pdf": {
            "bureau": "Experian",
            "text": """
--- PAGE 1 ---
Experian Credit Report
Prepared for: TIM KHAC DO
Consumer Name: TIM K DO
Current Address: 123 MAIN ST OLNEY MD 20832
Previous Address: 999 OLD RD ROCKVILLE MD 20850
Phone: 301-555-1212
Phone: 240-555-3434
DOB: 01/02/1980
SSN: 123-45-6789
Employer: OLD CAFE

MIDLAND CREDIT MANAGEMENT
Account Number: 1234567890
Account Type: Collection
Balance: $100
Status: Collection
Date Opened: 01/01/2020
""",
        }
    })
    data = result_to_dict(result)
    assert data["identity_raw_data"]
    write_outputs(result, tmp_path)
    workbook = load_workbook(tmp_path / "credit_vivo_desktop_scanner_output.xlsx")
    sheet = workbook["Identity_Cleanup"]
    rows = [
        [sheet.cell(row=row, column=column).value for column in range(1, 12)]
        for row in range(5, sheet.max_row + 1)
    ]
    actions = [row[0] for row in rows]
    assert "KEEP" in actions
    assert "DELETE" in actions
    assert any(row[1] == "masked_ssn" and row[2] == "***-**-6789" for row in rows)
    assert not any("123-45-6789" in " ".join(str(value or "") for value in row) for row in rows)
    assert any("Keep one confirmed current value" in row[7] for row in rows)
    assert any("Delete/remove this extra identity/contact value" in row[8] for row in rows)
